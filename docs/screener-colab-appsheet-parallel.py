# ===============================================================================
# SCREENER.IN MULTI-MTF TRACKER -- PARALLEL / Colab  (GAS v4.0 + AI + Dedup)
# Parallel 2 scanners + deduplicated YF/AI fetches for maximum speed
# ===============================================================================

# -- Install (run once per Colab session) --------------------------------------
# !pip install yfinance requests pytz boto3 openpyxl openai beautifulsoup4 -q

import math, json, re, time, io, warnings, logging, os, sys, sqlite3
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import html as _html_mod
import requests, pytz, boto3
import numpy as np
import pandas as pd
import yfinance as yf
import openpyxl
from copy import copy
from openai import OpenAI as AIClient
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from datetime import datetime, timedelta, time as dt_time, timezone
from botocore.client import Config as BotoConfig
from botocore.exceptions import ClientError
from urllib.parse import quote as url_encode, urlencode, urlparse, parse_qs

try:
    from bs4 import BeautifulSoup
except Exception:
    BeautifulSoup = None

# -- Runtime / log tuning ----------------------------------------------------
def _env_flag(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}

def _looks_like_colab_runtime() -> bool:
    return bool(
        os.environ.get("COLAB_RELEASE_TAG")
        or os.environ.get("COLAB_GPU")
        or "google.colab" in sys.modules
    )

def _looks_like_notebook_runtime() -> bool:
    return _looks_like_colab_runtime() or "ipykernel" in sys.modules or bool(os.environ.get("JPY_PARENT_PID"))

COMPACT_RUNTIME_LOGS = _env_flag("TRACKER_COMPACT_LOGS", _looks_like_colab_runtime())
LOG_SCREEN_PAGE_COUNTS = not COMPACT_RUNTIME_LOGS
LOG_NEW_STOCK_EVENTS = not COMPACT_RUNTIME_LOGS
LOG_SCANNER_WRITE_EACH_SHEET = not COMPACT_RUNTIME_LOGS
LOG_AI_PROVIDER_SUCCESS = not COMPACT_RUNTIME_LOGS

def _patch_notebook_warning_noise():
    # Colab / Jupyter on Python 3.12 can flood stderr with jupyter_client
    # datetime.utcnow() deprecation warnings when the notebook emits lots of
    # progress output. Filter the warning and patch the helper when possible.
    warnings.filterwarnings("ignore", category=DeprecationWarning, module=r"jupyter_client\.session")
    warnings.filterwarnings(
        "ignore",
        message=r"datetime\.datetime\.utcnow\(\) is deprecated.*",
        category=DeprecationWarning,
    )
    if not _looks_like_notebook_runtime():
        return
    try:
        import jupyter_client.session as _jc_session

        def _tracker_utcnow():
            return datetime.now(timezone.utc)

        if hasattr(_jc_session, "utcnow"):
            _jc_session.utcnow = _tracker_utcnow
    except Exception:
        pass

# -- Silence noisy yfinance / peewee logs ------------------------------------
logging.getLogger("yfinance").setLevel(logging.CRITICAL)
logging.getLogger("peewee").setLevel(logging.CRITICAL)
logging.getLogger("urllib3").setLevel(logging.WARNING)
warnings.filterwarnings("ignore")
_patch_notebook_warning_noise()

# -----------------------------------------------------------------------------
# ██████████████  CONFIG -- EDIT ONLY THIS SECTION  ████████████████████████████
# -----------------------------------------------------------------------------

TELEGRAM_BOT_TOKEN = "TEST_TELEGRAM_BOT_TOKEN"
TELEGRAM_CHAT_ID   = "TEST_TELEGRAM_CHAT_ID"
SCREENER_COOKIE    = "sessionid=TEST_SESSION_ID"
PROXY_URL          = "https://example.invalid/screener-proxy"
YF_SUFFIX          = ".NS"

S3_ENDPOINT   = "https://s3.example.invalid"
S3_BUCKET     = "test-bucket"
S3_ACCESS_KEY = "TEST_S3_ACCESS_KEY"
S3_SECRET_KEY = "TEST_S3_SECRET_KEY"
S3_REGION     = "test-region-1"
S3_EXCEL_KEY  = "reports/gas_stock_tracker.xlsx"
S3_DASHBOARD_EXCEL_KEY = "reports/gas_stock_tracker_dashboard.xlsx"
S3_DASHBOARD_DB_KEY = "reports/dashboard_snapshots.db"
LOCAL_EXCEL_FILE = "gas_stock_tracker.xlsx"
LOCAL_DASHBOARD_EXCEL_FILE = "gas_stock_tracker_dashboard.xlsx"

# -- AI Decision Layer ---------------------------------------------------------
AI_PRIMARY           = "google"  # "nvidia" or "google"
AI_SECONDARY_ENABLED = False     # False = only primary provider, True = fall back to the other provider

NVIDIA_NIM_API_KEYS = [
    "nvapi-TEST_NVIDIA_NIM_API_KEY",
]

NVIDIA_NIM_MODELS = [
    "z-ai/glm4.7",
]


GEMINI_API_KEYS = [
    "TEST_GEMINI_API_KEY",
]

# Multiple Google API keys -- add as many as you want.
# Used only when Google is primary, or when secondary fallback is enabled.
# Primary and secondary chains stay isolated; keys rotate only within the active provider.
GEMINI_MODELS = [
    "gemini-3.1-flash-lite-preview",
]

# Keep one fixed model per provider by default to reduce variance.
# Multiple keys still scale throughput without changing decision behavior.
AI_TEMPERATURE  = 0.0
AI_TOP_P        = 1.0
AI_MAX_TOKENS   = 256
AI_ENABLED      = False     # set False to skip AI column entirely
AI_DELAY_SEC    = 0.6      # extra delay per AI call
GEMINI_RPM_PER_KEY = 10    # strict rolling per-key cap for Google requests
GEMINI_BATCH_ENABLED = False        # Opt-in only: Gemini Batch API is asynchronous and can take much longer than live requests
GEMINI_BATCH_MIN_REQUESTS = 12      # keep small runs on the simple interactive path
GEMINI_BATCH_CHUNK_SIZE = 40        # inline requests per batch job
GEMINI_BATCH_MAX_INLINE_BYTES = 18_000_000
GEMINI_BATCH_POLL_SEC = 10
GEMINI_BATCH_MAX_WAIT_SEC = 900
AI_TIMEOUT_SEC  = 60       # per-request timeout
AI_MAX_RETRIES  = 2        # retries per key/model combo

# -- Fundamental Decision Layer ----------------------------------------------
FUNDAMENTALS_ENABLED = True
FUNDAMENTAL_FETCH_WORKERS = 2
FUNDAMENTAL_FETCH_TIMEOUT_SEC = 20
FUNDAMENTAL_CACHE_FILE = "fundamental_cache.json"
FUNDAMENTAL_CACHE_VERSION = 1
FUNDAMENTAL_CACHE_MAX_AGE_DAYS = 7
FUNDAMENTAL_FORCE_REFRESH = False
FUNDAMENTAL_USE_PROXY = True
FUNDAMENTAL_PREFER_CONSOLIDATED = True

ROTATION_COOLDOWN_SEC = 60
YF_DELAY_SEC          = 0.35
YF_CACHE_MAX          = 300
OFF_MARKET_SIGNAL_SNAPSHOT_ENABLED = False  # Disabled: always fetch fresh from screener+YF on each Colab session; Days Tracked persists via S3 Excel
SIGNAL_SNAPSHOT_FILE  = "signal_snapshot_cache.json"
SIGNAL_SNAPSHOT_KEEP_DAYS = 10

SIGNAL_PROFILE = "balanced"  # precision | conservative | balanced | aggressive
SIGNAL_ENGINE = "enhanced"   # enhanced | legacy
WALKFORWARD_CACHE_FILE = "signal_walkforward_cache.json"
WALKFORWARD_CACHE_VERSION = 1
WALKFORWARD_LOOKBACK_BARS = 320     # enough for MA200 + BB rank context
WALKFORWARD_EVAL_BARS = 180         # recent bars used for walk-forward validation
WALKFORWARD_MIN_SAMPLES = 5         # below this, quality gate becomes conservative
QUALITY_GATE_MIN_WIN_PROB = 55.0
QUALITY_GATE_MIN_HIST_PRECISION = 55.0
QUALITY_GATE_MIN_EXP_10D = 1.0
QUALITY_GATE_MIN_SCORE = 55.0

BENCHMARK_TICKER = "^NSEI"  # NIFTY 50 broad-market benchmark
SECTOR_PROFILE_TIMEOUT_SEC = 8
SELF_VALIDATION_ENABLED = True
SELF_VALIDATION_SAMPLE_SIZE = 12
SELF_VALIDATION_NUMERIC_TOLERANCE = 0.05
SELF_VALIDATION_MAX_DETAILS = 5

# -- Parallel + Crash-proof settings ------------------------------------------
PARALLEL_WORKERS          = 2      # how many YF/AI fetches to run in parallel
SCREENER_FETCH_WORKERS    = 4      # screener page fetch concurrency (kept separate from YF/AI)
S3_CHECKPOINT_EVERY       = 20      # upload S3 checkpoint every N scanners (not every 1)
S3_CHECKPOINT_SAVE_LOCAL  = False  # periodic checkpoints skip local artifact save for speed
S3_CHECKPOINT_BACKUP      = False  # periodic checkpoints skip backup copy; final save still backs up
S3_UPLOAD_RETRIES         = 3      # retry S3 upload N times before giving up
RELOAD_WB_EVERY           = 3      # reload workbook from S3 every N iterations (prevents RAM bloat)
AI_CACHE_MAX              = 500    # max entries in AI decision cache (LRU evict oldest)
COLAB_RESTART_EVERY       = 500      # restart Colab runtime every N iterations (0 = never restart)
DASHBOARD_DB_ENABLED      = True     # append Dashboard rows to SQLite each iteration
DASHBOARD_DB_FILE         = "dashboard_snapshots.db"
DASHBOARD_DB_TABLE        = "dashboard_snapshots"

SCANNERS = [
     {"id": "goat1",        "name": "GOAT 1",               "url": "https://www.screener.in/screens/3525076/goat1/?page=1",                 "color": "E8F5E9"},
     {"id": "weekly-burst", "name": "Weekly Burst",          "url": "https://www.screener.in/screens/3530492/weekly-burst/?page=1",           "color": "E3F2FD"},
     {"id": "month-ret",    "name": "Month Returns",         "url": "https://www.screener.in/screens/3530540/month-returns/?page=1",          "color": "FFF8E1"},
     {"id": "weekly-max",   "name": "Weekly Max Precision",  "url": "https://www.screener.in/screens/3530554/weekly-screener-maximum-technical-precision/?page=1", "color": "F3E5F5"},
     {"id": "monthly-opus", "name": "Monthly Opus V3",       "url": "https://www.screener.in/screens/3534290/monthly-opus-v3/?page=1",        "color": "E0F7FA"},
     {"id": "opus-tele",    "name": "Opus Tele",             "url": "https://www.screener.in/screens/3535927/opus-tele/?page=1",              "color": "FCE4EC"},
     {"id": "weekly",       "name": "Weekly",                "url": "https://www.screener.in/screens/3536956/weekly/?page=1",                 "color": "E8EAF6"},
     {"id": "daily-scan",   "name": "Daily Scanner",         "url": "https://www.screener.in/screens/3536981/daily-scanner/?page=1",          "color": "E1F5FE"},
     {"id": "daily-must",   "name": "Daily Must Buy",        "url": "https://www.screener.in/screens/3537642/daily-must-buy/?page=1",         "color": "E8F5E9"},
     {"id": "weekly-must",  "name": "Weekly Must Buy",       "url": "https://www.screener.in/screens/3537648/weekly-must-buy/?page=1",        "color": "FFF3E0"},
     {"id": "monthly-must", "name": "Monthly Must Buy",      "url": "https://www.screener.in/screens/3537651/monthly-must-buy/?page=1",       "color": "F1F8E9"},
     {"id": "daily-v2",     "name": "Daily Scanner V2",      "url": "https://www.screener.in/screens/3539839/daily-scanner-v2-less-noise/?page=1", "color": "E0F2F1"},
     {"id": "long-term",    "name": "Long Term",             "url": "https://www.screener.in/screens/3540686/long-term/?page=1",              "color": "FBE9E7"},
     {"id": "lt-fund",      "name": "Long Term Fundamental", "url": "https://www.screener.in/screens/3558022/long-term-fundamendal-built/?page=1", "color": "FBE9E7"},
     {"id": "daily-v2-1",   "name": "Daily V2 Less Noise-1", "url": "https://www.screener.in/screens/3566765/daily-scanner-v2-less-noise-1/?page=1", "color": "FBE9E7"},
     {"id": "good-value",   "name": "Good Value",            "url": "https://www.screener.in/screens/3567890/good-value/?page=1",             "color": "FFF9C4"},
     {"id": "refer-only",   "name": "less fundamendals-up-only-refer-solid-screeers", "url": "https://www.screener.in/screens/3567891/only-up-no-buy-refer-its-on-other-screeners/?page=1", "color": "FBE9E7"},
     {"id": "daily-Max-dervied-from-weekly",   "name": "daily-precision-from-weekly", "url": "https://www.screener.in/screens/3570124/daily-precision-from-weekly/?page=1", "color": "FBE9E7"},
     {"id": "DTECH-001", "name": "DTECH-001-REF-DEARLY-T-AND-DFUND-001", "url": "https://www.screener.in/screens/3580814/dtech-001/?page=1", "color": "FBE9E7"},
     {"id": "DFUND-001", "name": "DFUND-001-REF-DEARLY-F-AND-DTECH-001", "url": "https://www.screener.in/screens/3580836/dfund-001/?page=1", "color": "FBE9E7"},
     {"id": "DCOMB-001", "name": "DCOMB-001-REF-DCOMB-002", "url": "https://www.screener.in/screens/3580842/dcomb-001/?page=1", "color": "FBE9E7"},
     {"id": "DCOMB-002", "name": "DCOMB-002-REF-DCOMB-001", "url": "https://www.screener.in/screens/3580844/dcomb-002/?page=1", "color": "FBE9E7"},
     {"id": "DSCAN-001", "name": "DSCAN-001-REF-DMUST-001-DTECH-001-DFUND-001-AND-DSCAN-001", "url": "https://www.screener.in/screens/3580885/dscan-001/?page=1", "color": "FBE9E7"},
     {"id": "DUP-001",   "name": "DUP-001-REF-DTECH-001", "url": "https://www.screener.in/screens/3580889/dup-001/?page=1", "color": "FBE9E7"},
     {"id": "DEARLY-T",  "name": "DEARLY-T-REF-DTECH-001", "url": "https://www.screener.in/screens/3580894/dearly-t/?page=1", "color": "FBE9E7"},
     {"id": "DEARLY-F",  "name": "DEARLY-F-REF-DFUND-001", "url": "https://www.screener.in/screens/3580899/dearly-f/?page=1", "color": "FBE9E7"},
     {"id": "DEARLY-C",  "name": "DEARLY-C-REF-DCOMB-001-OR-DCOMB-002", "url": "https://www.screener.in/screens/3580901/dearly-c/?page=1", "color": "FBE9E7"},
     {"id": "DMUST-001", "name": "DMUST-001-REF-DTECH-001-DFUND-001-AND-DSCAN-001", "url": "https://www.screener.in/screens/3580905/dmust-001/?page=1", "color": "FBE9E7"},
     {"id": "WTECH-001", "name": "WTECH-001-REF-WEARLY-T-AND-WFUND-001", "url": "https://www.screener.in/screens/3580925/wtech-001/?page=1", "color": "FBE9E7"},
     {"id": "WFUND-001", "name": "WFUND-001-REF-WEARLY-F-AND-WTECH-001", "url": "https://www.screener.in/screens/3580939/wfund-001/?page=1", "color": "FBE9E7"},
     {"id": "WCOMB-001", "name": "WCOMB-001-REF-WCOMB-002", "url": "https://www.screener.in/screens/3580951/wcomb-001/?page=1", "color": "FBE9E7"},
     {"id": "WCOMB-002", "name": "WCOMB-002-REF-WCOMB-001", "url": "https://www.screener.in/screens/3580954/wcomb-002/?page=1", "color": "FBE9E7"},
     {"id": "WSCAN-001", "name": "WSCAN-001-REF-WMUST-001-WTECH-001-WFUND-001-AND-WSCAN-001", "url": "https://www.screener.in/screens/3580955/wscan-001/?page=1", "color": "FBE9E7"},
     {"id": "WUP-001",   "name": "WUP-001-REF-WTECH-001", "url": "https://www.screener.in/screens/3580956/wup-001/?page=1", "color": "FBE9E7"},
     {"id": "WEARLY-T",  "name": "WEARLY-T-REF-WTECH-001", "url": "https://www.screener.in/screens/3580961/wearly-t/?page=1", "color": "FBE9E7"},
     {"id": "WEARLY-F",  "name": "WEARLY-F-REF-WFUND-001", "url": "https://www.screener.in/screens/3580963/wearly-f/?page=1", "color": "FBE9E7"},
     {"id": "WEARLY-C",  "name": "WEARLY-C-REF-WCOMB-001-OR-WCOMB-002", "url": "https://www.screener.in/screens/3580965/wearly-c/?page=1", "color": "FBE9E7"},
     {"id": "WMUST-001", "name": "WMUST-001-REF-WTECH-001-WFUND-001-AND-WSCAN-001", "url": "https://www.screener.in/screens/3580968/wmust-001/?page=1", "color": "FBE9E7"},
     {"id": "MTECH-001", "name": "MTECH-001-REF-MEARLY-T-AND-MFUND-001", "url": "https://www.screener.in/screens/3580973/mtech-001/?page=1", "color": "FBE9E7"},
     {"id": "MFUND-001", "name": "MFUND-001-REF-MEARLY-F-AND-MTECH-001", "url": "https://www.screener.in/screens/3580978/mfund-001/?page=1", "color": "FBE9E7"},
     {"id": "MCOMB-001", "name": "MCOMB-001-REF-MCOMB-002", "url": "https://www.screener.in/screens/3580987/mcomb-001/?page=1", "color": "FBE9E7"},
     {"id": "MCOMB-002", "name": "MCOMB-002-REF-MCOMB-001", "url": "https://www.screener.in/screens/3580991/mcomb-002/?page=1", "color": "FBE9E7"},
     {"id": "MSCAN-001", "name": "MSCAN-001-REF-MMUST-001-MTECH-001-MFUND-001-AND-MSCAN-001", "url": "https://www.screener.in/screens/3580996/mscan-001/?page=1", "color": "FBE9E7"},
     {"id": "MUP-001",   "name": "MUP-001-REF-MTECH-001", "url": "https://www.screener.in/screens/3580999/mup-001/?page=1", "color": "FBE9E7"},
     {"id": "MEARLY-T",  "name": "MEARLY-T-REF-MTECH-001", "url": "https://www.screener.in/screens/3581022/mearly-t/?page=1", "color": "FBE9E7"},
     {"id": "MEARLY-F",  "name": "MEARLY-F-REF-MFUND-001", "url": "https://www.screener.in/screens/3581024/mearly-f/?page=1", "color": "FBE9E7"},
     {"id": "MEARLY-C",  "name": "MEARLY-C-REF-MCOMB-001-OR-MCOMB-002", "url": "https://www.screener.in/screens/3581026/mearly-c/?page=1", "color": "FBE9E7"},
     {"id": "MMUST-001", "name": "MMUST-001-REF-MTECH-001-MFUND-001-AND-MSCAN-001", "url": "https://www.screener.in/screens/3581028/mmust-001/?page=1", "color": "FBE9E7"},
     {"id": "🟡 SCREENER 1-early", "name": "1-early", "url": "https://www.screener.in/screens/3586543/screener-1-trend-birth-earliest-possible/?page=1", "color": "FBE9E7"},
     {"id": "🔴 SCREENER 2— MOMENTUM-IGNITION", "name": "2—MOMENTUM-IGNITION", "url": "https://www.screener.in/screens/3586544/screener-2-momentum-ignition/?page=1", "color": "FBE9E7"},
     {"id": "🟢 SCREENER 3—FULL POWER LOCK", "name": "3—FULL-POWER-LOCK", "url": "https://www.screener.in/screens/3586546/screener-3-full-power-lock/?page=1", "color": "FBE9E7"},
     {"id": "multibagger-1", "name": "multibagger", "url": "https://www.screener.in/screens/3586536/multibagger/", "color": "FBE9E7"},
     {"id": "🟢 Variant-1—Aggressive-SwingTrading", "name": "Variant-1—Aggressive-SwingTrading", "url": "https://www.screener.in/screens/3589722/variant-1-aggressive-swing-trading/?page=1", "color": "FBE9E7"},
 {"id": "🟡 Variant-2—Standard-PositionTrading", "name": "Variant-2—Standard-PositionTrading", "url": "https://www.screener.in/screens/3589724/variant-2-standard-position-trading/?page=1", "color": "FBE9E7"},
 {"id": "🔴 Variant-3—Ultra Strict Breakout Leaders Only", "name": "Variant-3—Ultra Strict Breakout Leaders Only", "url": "https://www.screener.in/screens/3589726/variant-3-ultra-strict-breakout-leaders-only/?page=1", "color": "FBE9E7"},

 ]


IST = pytz.timezone("Asia/Kolkata")

SIGNAL_PROFILES = {
    "precision":    dict(ADX_STRONG=28, ADX_WEAK=20, VOL_HIGH=1.6,  DIST_52W_MAX=1.5,
                         BREAKOUT_MIN=1.0, RSI_OVERSOLD=30, RSI_NEU_MIN=50, RSI_BULL_MIN=58, RSI_OB=74,
                         DI_BULL_GAP=8, NATR_MAX=6.0, NATR_HOT=7.5, BB_SQUEEZE_PCTL=12,
                         HIGH_PRECISION=True, BB_PUSH_MIN=0.9, BB_STRETCHED_MIN=1.02,
                         SELL_RSI_MAX=42, SELL_BB_MAX=0.35, DI_SELL_GAP=8, SELL_BREAKDOWN_MIN=2.0),
    "conservative": dict(ADX_STRONG=25, ADX_WEAK=18, VOL_HIGH=1.5,  DIST_52W_MAX=2,
                         BREAKOUT_MIN=0.75, RSI_OVERSOLD=30, RSI_NEU_MIN=45, RSI_BULL_MIN=55, RSI_OB=78,
                         DI_BULL_GAP=6, NATR_MAX=6.5, NATR_HOT=8.5, BB_SQUEEZE_PCTL=15),
    "balanced":     dict(ADX_STRONG=20, ADX_WEAK=16, VOL_HIGH=1.25, DIST_52W_MAX=3,
                         BREAKOUT_MIN=0.25, RSI_OVERSOLD=32, RSI_NEU_MIN=45, RSI_BULL_MIN=52, RSI_OB=78,
                         DI_BULL_GAP=4, NATR_MAX=8.0, NATR_HOT=10.0, BB_SQUEEZE_PCTL=20),
    "aggressive":   dict(ADX_STRONG=18, ADX_WEAK=14, VOL_HIGH=1.1,  DIST_52W_MAX=5,
                         BREAKOUT_MIN=0,   RSI_OVERSOLD=35, RSI_NEU_MIN=42, RSI_BULL_MIN=50, RSI_OB=80,
                         DI_BULL_GAP=3, NATR_MAX=10.0, NATR_HOT=12.0, BB_SQUEEZE_PCTL=25),
}

FUNDAMENTAL_RAW_FIELDS = [
    "Fundamental Source","Fundamental Updated At","Fundamental Freshness",
    "Market Cap Cr","PE","PB","EV/EBITDA","Dividend Yield%","Book Value","EPS TTM",
    "Sales TTM Cr","Profit TTM Cr","OPM%","NPM%",
    "ROE%","ROCE%","ROE 3Y%","ROE Last Year%",
    "Sales CAGR 3Y%","Sales CAGR 5Y%","Sales Growth TTM%",
    "Profit CAGR 3Y%","Profit CAGR 5Y%","Profit Growth TTM%",
    "Debt/Equity","Borrowings Cr","Interest Coverage","Asset Turnover",
    "CFO TTM Cr","FCF TTM Cr","CFO/PAT","CFO/OP%",
    "Debtor Days","Inventory Days","Cash Conversion Cycle",
    "Promoter Holding%","Promoter Holding Change%","Promoter Pledge%",
    "Pros Count","Cons Count",
]

FUNDAMENTAL_DERIVED_FIELDS = [
    "Profitability Tag","Growth Tag","Valuation Tag","Balance Sheet Tag",
    "Cashflow Tag","Ownership Tag","Fundamental Risk Tag","Fundamental Quality Tag",
    "Fundamental Score","Investability Tag","Early Entry OK",
    "Tech + Fundamental Score","Decision Guardrail","Final Confidence Tag",
    "Final Signal","Final Signal Reason",
]

FUNDAMENTAL_SUMMARY_FIELDS = [
    "Fundamental Score","Fundamental Quality Tag","Growth Tag","Valuation Tag",
    "Balance Sheet Tag","Cashflow Tag","Ownership Tag","Fundamental Risk Tag",
    "Investability Tag","Early Entry OK","Tech + Fundamental Score","Decision Guardrail",
    "Final Confidence Tag","Final Signal",
    "Fundamental Updated At","Fundamental Freshness",
]

FUNDAMENTAL_SCANNER_FIELDS = FUNDAMENTAL_RAW_FIELDS + FUNDAMENTAL_DERIVED_FIELDS

# -- Column schema (AI columns placed right after the signal columns) ---------
HEADERS = [
    "Symbol","Name","First Captured","Last Seen","In Screener?",
    "Capture Price","Current Price","Since Capture%",
    "1D%","1W%","1M%","3M%","6M%","1Y%","2Y%","3Y%",
    "Avg Weekly%","Avg Monthly%","Avg 3M%","Avg 6M%","Avg 1Y%",
    "RSI 14","MA 20","MA 50","MA 200",
    "Signal","Setup Signal","Core Signal",
    "Signal Quality","Signal Regime","Win Prob%","Hist Precision%","Exp 5D%","Exp 10D%","WF Samples",
    "Sector","Industry","Sector Benchmark","RS Tag","RS vs NIFTY 1M%","RS vs NIFTY 3M%","RS vs Sector 1M%","RS vs Sector 3M%","Avg Traded Value 20D Cr","Liquidity Tag",
    "AI Decision","AI Reason","AI Conf%",  # ← 3 new AI columns
    "Last Updated",
    "ADX 14","Vol Ratio 20","MACD Line","MACD Hist","52W High Dist%","20D Breakout%",
    "ATR 14","NATR 14","+DI 14","-DI 14",
] + FUNDAMENTAL_SCANNER_FIELDS
C = {h: i for i, h in enumerate(HEADERS)}

PRICE_HISTORY_HEADERS = [
    "Snapshot At","Scanner","Symbol","Name","In Screener?",
    "Capture Price","Current Price","Since Capture%",
    "1D%","1W%","1M%","3M%","6M%","1Y%",
    "Cam H3","Cam H4","Cam L3","Cam L4",
    "RSI 14","ADX 14","+DI 14","-DI 14","ATR 14","NATR 14","Vol Ratio 20","MACD Line","MACD Hist",
    "52W High Dist%","20D Breakout%","Signal","Setup Signal","Core Signal",
    "Signal Quality","Signal Regime","Win Prob%","Hist Precision%","Exp 5D%","Exp 10D%","WF Samples",
    "Sector","Industry","Sector Benchmark","RS Tag","RS vs NIFTY 1M%","RS vs NIFTY 3M%","RS vs Sector 1M%","RS vs Sector 3M%","Avg Traded Value 20D Cr","Liquidity Tag",
    "AI Decision","AI Conf%",
] + FUNDAMENTAL_SUMMARY_FIELDS

DASHBOARD_HEADERS = [
    "Symbol","Name","In Screener?","Quick Action","Consensus Score","MTF Alignment","Historical MTF",
    "Sector","Industry","Sector Benchmark","RS Tag","RS vs NIFTY 1M%","RS vs NIFTY 3M%","RS vs Sector 1M%","RS vs Sector 3M%","Avg Traded Value 20D Cr","Liquidity Tag",
    "Momentum Rank","Risk Tag","BB Signal","Cam Setup","Volume Buzz","Since Capture Trend",
    "First Captured","Days Tracked","Last Seen",
    "Total Appearances","Unique Scanners","Scanner List","Best Scanner",
    "Capture Price","Current Price","Cam H3","Cam H4","Cam L3","Cam L4",
    "Ideal Enter Price","Possible Sell Value","Stop Loss Value","Since Capture%",
    "1D%","1W%","1M%","3M%","6M%","1Y%",
    "RSI 14","ADX 14","+DI 14","-DI 14","ATR 14","NATR 14",
    "Signal","Setup Signal","Core Signal",
    "Signal Quality","Signal Regime","Win Prob%","Hist Precision%","Exp 5D%","Exp 10D%","WF Samples",
    "AI Decision","AI Conf%",
    "Screener Link","Last Updated","Momentum Tag",
] + FUNDAMENTAL_RAW_FIELDS + FUNDAMENTAL_DERIVED_FIELDS
DC = {h: i for i, h in enumerate(DASHBOARD_HEADERS)}

DASHBOARD_HISTORY_HEADERS = [
    "Snapshot At","Symbol","Name","In Screener?","Quick Action","Consensus Score","MTF Alignment","Historical MTF",
    "Sector","Industry","Sector Benchmark","RS Tag","RS vs NIFTY 1M%","RS vs NIFTY 3M%","RS vs Sector 1M%","RS vs Sector 3M%","Avg Traded Value 20D Cr","Liquidity Tag",
    "Momentum Rank","Risk Tag","Cam Setup","Total Appearances","Unique Scanners",
    "Scanner List","Capture Price","Current Price","Cam H3","Cam H4","Cam L3","Cam L4",
    "Ideal Enter Price","Possible Sell Value","Stop Loss Value","Since Capture%",
    "1D%","1W%","1M%","RSI 14","ADX 14","+DI 14","-DI 14","ATR 14","NATR 14",
    "Signal","Setup Signal","Core Signal",
    "Signal Quality","Signal Regime","Win Prob%","Hist Precision%","Exp 5D%","Exp 10D%","WF Samples",
    "AI Decision","AI Conf%",
] + FUNDAMENTAL_SUMMARY_FIELDS

VALIDATION_HEADERS = [
    "Snapshot At","Iteration","Mode","Checked Rows","Matched Rows","Mismatch Rows","Unresolved Rows",
    "Latest Session","Status","Details"
]

# -- Scanner timeframe classification (for MTF Alignment) -----------------
def _scanner_timeframe(scanner_id: str) -> str:
    """Classify a scanner as Daily/Weekly/Monthly based on its ID."""
    sid = scanner_id.upper()
    if sid.startswith("D") and sid not in ("DAILY-MAX-DERVIED-FROM-WEEKLY",):
        return "D"
    if sid.startswith("W"):
        return "W"
    if sid.startswith("M"):
        return "M"
    # Legacy scanners
    sl = scanner_id.lower()
    if "daily" in sl:
        return "D"
    if "weekly" in sl or "week" in sl:
        return "W"
    if "month" in sl:
        return "M"
    if "long-term" in sl or "lt-" in sl:
        return "M"
    if "good-value" in sl or "refer-only" in sl or "goat" in sl or "opus" in sl:
        return "W"  # default mid-range
    return "D"  # default

# Build scanner ID -> timeframe map
_SCANNER_TF = {sc["id"]: _scanner_timeframe(sc["id"]) for sc in SCANNERS}

SENTINEL = "~NOFOUND"


def _format_mtf_ticks(timeframes):
    return " ".join(
        f"{tf}{chr(0x2705) if tf in timeframes else chr(0x274C)}"
        for tf in ("D", "W", "M")
    )


class Phase1NoStocksError(RuntimeError):
    pass

# Module-level BB/Camarilla caches -- populated in Phase 3, read by update_dashboard()
# Key: symbol -> dict
_bb_data: dict = {}
_cam_data: dict = {}
_price_mtf_data: dict = {}  # symbol -> D/W/M price-based alignment string
_fundamental_cache: dict | None = None
_fundamental_cache_dirty: bool = False
_fundamental_data: dict = {}  # symbol -> normalized fundamental summary
_fundamental_parser_warned: bool = False
_screener_auth_warned: bool = False
_walkforward_cache = None
_walkforward_cache_dirty = False
_walkforward_cache_lock = Lock()
_benchmark_hist_cache: dict = {}
_benchmark_hist_lock = Lock()
_symbol_meta_cache: dict = {}
_symbol_meta_lock = Lock()
_sector_benchmark_ticker_cache: dict = {}

# -----------------------------------------------------------------------------
# AI DECISION ENGINE  (Primary/secondary provider chain)
# -----------------------------------------------------------------------------
# Cache: key=(symbol + rule signal + rounded metric snapshot + profile)
# Avoids reusing stale AI output when the underlying metrics changed.
_ai_cache: dict = {}

# Round-robin counter -- incremented on every AI call to distribute load
_ai_rr_counter: int = 0

# Lazy AI clients -- one per provider tag, created once
_ai_clients: dict = {}  # tag -> AIClient
_gemini_rate_limit_hits: dict = {}
_gemini_rate_limit_lock = Lock()
_ai_disabled_tags: set[str] = set()
_ai_disabled_reasons: dict = {}
_ai_disabled_lock = Lock()

def _ai_tag_disabled(tag: str) -> bool:
    with _ai_disabled_lock:
        return str(tag or "") in _ai_disabled_tags

def _disable_ai_tag(tag: str, label: str, reason):
    tag = str(tag or "").strip()
    if not tag:
        return
    with _ai_disabled_lock:
        if tag in _ai_disabled_tags:
            return
        _ai_disabled_tags.add(tag)
        _ai_disabled_reasons[tag] = str(reason)[:300]
    _ai_clients.pop(tag, None)
    print(f"      [AI] Disabled {label} for this run: permanent API key/auth error")

def _is_permanent_ai_auth_error(exc) -> bool:
    text = str(exc or "").upper()
    markers = (
        "API_KEY_INVALID",
        "API KEY EXPIRED",
        "API KEY HAS EXPIRED",
        "KEY EXPIRED",
        "API KEY NOT FOUND",
        "INVALID API KEY",
        "UNAUTHENTICATED",
        "PERMISSION_DENIED",
        "HTTP 401",
    )
    return any(marker in text for marker in markers)

def _valid_model_list(models) -> list:
    return [str(m).strip() for m in models if str(m or "").strip() and str(m).strip().upper() != "NA"]

def _valid_nvidia_api_keys():
    return [str(k).strip() for k in NVIDIA_NIM_API_KEYS if str(k or "").strip().startswith("nvapi-")]

def _valid_gemini_api_keys():
    return [str(k).strip() for k in GEMINI_API_KEYS if str(k or "").strip() and str(k).strip().upper() != "NA"]

def _ai_key_tail(api_key: str) -> str:
    key = str(api_key or "").strip()
    return f"...{key[-6:]}" if len(key) >= 6 else "***"

def _ai_primary_provider() -> str:
    provider = str(AI_PRIMARY or "").strip().lower()
    return provider if provider in {"nvidia", "google"} else "nvidia"

def _ai_secondary_provider() -> str:
    return "google" if _ai_primary_provider() == "nvidia" else "nvidia"

def _provider_model_list(provider: str) -> list:
    return _valid_model_list(NVIDIA_NIM_MODELS if provider == "nvidia" else GEMINI_MODELS)

def _build_ai_provider_entries(provider: str, rr_seed: int) -> list:
    if provider == "nvidia":
        keys = _valid_nvidia_api_keys()
        models = _provider_model_list("nvidia")
        prefix = "nim"
        label_prefix = "NVIDIA"
    else:
        keys = _valid_gemini_api_keys()
        models = _provider_model_list("google")
        prefix = "gemini"
        label_prefix = "Google"

    if not keys or not models:
        return []

    total = len(keys) * len(models)
    start_combo = rr_seed % total
    entries = []
    for offset in range(total):
        combo_idx = (start_combo + offset) % total
        key_idx = combo_idx % len(keys)
        model_idx = (combo_idx // len(keys)) % len(models)
        model = models[model_idx]
        short_model = model.split("/")[-1]
        key_tail = _ai_key_tail(keys[key_idx])
        tag = f"{prefix}_{key_idx}"
        if _ai_tag_disabled(tag):
            continue
        entries.append((tag, model, f"{label_prefix}-{key_idx+1}{key_tail}({short_model})", provider))
    return entries

def _nvidia_request_options(model: str) -> dict:
    """
    NVIDIA NIM model families do not expose reasoning controls consistently.
    For GLM models, explicitly disable thinking via chat_template_kwargs.
    Also send the generic thinking=disabled hint for routes that honor it.
    """
    lowered = str(model or "").strip().lower()
    extra_body = {"thinking": {"type": "disabled"}}
    if "glm" in lowered:
        extra_body["chat_template_kwargs"] = {"enable_thinking": False}
    return {"extra_body": extra_body}

def _gemini_key_for_tag(tag: str) -> str | None:
    if not tag.startswith("gemini_"):
        return None
    if _ai_tag_disabled(tag):
        return None
    try:
        idx = int(tag.split("_")[1])
    except (IndexError, ValueError):
        return None
    keys = _valid_gemini_api_keys()
    return keys[idx] if idx < len(keys) else None

def _gemini_wait_for_slot(api_key: str):
    """
    Enforce a strict rolling 60-second rate limit per Gemini API key.
    This keeps parallel workers from exceeding the per-key request budget.
    Failed attempts still count because they are real outbound requests.
    """
    rpm = int(GEMINI_RPM_PER_KEY or 0)
    if rpm <= 0 or not api_key:
        return
    window_sec = 60.0
    safety_pad_sec = 0.05
    key_id = str(api_key).strip()
    while True:
        wait_sec = 0.0
        now = time.monotonic()
        with _gemini_rate_limit_lock:
            hits = _gemini_rate_limit_hits.setdefault(key_id, [])
            cutoff = now - window_sec
            while hits and hits[0] <= cutoff:
                hits.pop(0)
            if len(hits) < rpm:
                hits.append(now)
                return
            wait_sec = max(safety_pad_sec, window_sec - (now - hits[0]) + safety_pad_sec)
        time.sleep(wait_sec)

def _gemini_response_state_name(value) -> str:
    if isinstance(value, dict):
        return str(value.get("name") or value.get("state") or "").strip()
    return str(value or "").strip()

def _gemini_batch_state_name(job: dict) -> str:
    if not isinstance(job, dict):
        return ""
    candidates = [
        job.get("state"),
        (job.get("metadata") or {}).get("state"),
        (job.get("response") or {}).get("state"),
    ]
    for candidate in candidates:
        name = _gemini_response_state_name(candidate)
        if name:
            return name
    return ""

def _gemini_extract_text_from_response(response_obj: dict) -> str:
    candidates = (response_obj or {}).get("candidates") or []
    if not candidates:
        raise ValueError(f"No candidates in Gemini response: {str(response_obj)[:300]}")
    first = candidates[0] or {}
    content = first.get("content") or {}
    parts = content.get("parts") or []
    texts = []
    for part in parts:
        if isinstance(part, dict):
            text = part.get("text")
            if text:
                texts.append(str(text))
    raw = "".join(texts).strip()
    if not raw:
        finish_reason = first.get("finishReason")
        raise ValueError(f"Empty Gemini content (finishReason={finish_reason})")
    return raw

def _build_google_generate_content_request(user_msg: str, max_tokens: int) -> dict:
    return {
        "systemInstruction": {
            "parts": [{"text": _AI_SYSTEM_PROMPT}],
        },
        "contents": [
            {
                "role": "user",
                "parts": [
                    {
                        "text": user_msg + "\nReturn one minified JSON object only.",
                    }
                ],
            }
        ],
        "generationConfig": {
            "temperature": AI_TEMPERATURE,
            "topP": AI_TOP_P,
            "maxOutputTokens": max(max_tokens, 384),
            "responseMimeType": "application/json",
            "responseJsonSchema": _AI_RESPONSE_SCHEMA,
            "thinkingConfig": {
                "thinkingLevel": "minimal",
                "includeThoughts": False,
            },
        },
    }

def _gemini_batch_enabled_for_run(request_count: int) -> bool:
    return bool(
        GEMINI_BATCH_ENABLED
        and request_count >= max(1, int(GEMINI_BATCH_MIN_REQUESTS or 0))
        and _ai_primary_provider() == "google"
        and not AI_SECONDARY_ENABLED
        and _valid_gemini_api_keys()
        and _provider_model_list("google")
    )

def _gemini_batch_chunks(contexts: list) -> list:
    max_items = max(1, int(GEMINI_BATCH_CHUNK_SIZE or 1))
    max_bytes = max(1024, int(GEMINI_BATCH_MAX_INLINE_BYTES or 18_000_000))
    chunks = []
    current = []
    current_bytes = 0
    for ctx in contexts:
        request_obj = _build_google_generate_content_request(ctx["user_msg"], AI_MAX_TOKENS)
        inline_item = {
            "request": request_obj,
            "metadata": {"key": ctx["symbol"]},
        }
        encoded = json.dumps(inline_item, separators=(",", ":")).encode("utf-8")
        inline_bytes = len(encoded)
        if current and (len(current) >= max_items or current_bytes + inline_bytes > max_bytes):
            chunks.append(current)
            current = []
            current_bytes = 0
        current.append((ctx, inline_item))
        current_bytes += inline_bytes
    if current:
        chunks.append(current)
    return chunks

def _call_google_batch_generate_content(model: str, api_key: str, inline_requests: list, display_name: str) -> dict:
    _gemini_wait_for_slot(api_key)
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:batchGenerateContent"
    payload = {
        "batch": {
            "display_name": display_name,
            "input_config": {
                "requests": {
                    "requests": inline_requests,
                }
            },
        }
    }
    resp = requests.post(
        url,
        headers={"x-goog-api-key": api_key, "Content-Type": "application/json"},
        json=payload,
        timeout=AI_TIMEOUT_SEC,
    )
    if resp.status_code >= 400:
        try:
            detail = resp.json()
        except Exception:
            detail = resp.text[:500]
        raise RuntimeError(f"HTTP {resp.status_code}: {detail}")
    return resp.json()

def _get_google_batch_job(name: str, api_key: str) -> dict:
    _gemini_wait_for_slot(api_key)
    url = f"https://generativelanguage.googleapis.com/v1beta/{name}"
    resp = requests.get(
        url,
        headers={"x-goog-api-key": api_key, "Content-Type": "application/json"},
        timeout=AI_TIMEOUT_SEC,
    )
    if resp.status_code >= 400:
        try:
            detail = resp.json()
        except Exception:
            detail = resp.text[:500]
        raise RuntimeError(f"HTTP {resp.status_code}: {detail}")
    return resp.json()

def _gemini_batch_inline_responses(job: dict) -> list:
    for key in ("inlinedResponses", "inlined_responses"):
        val = (job or {}).get(key)
        if isinstance(val, list):
            return val
    response = (job or {}).get("response") or {}
    dest = (job or {}).get("dest") or {}
    for key in ("inlinedResponses", "inlined_responses"):
        val = response.get(key)
        if isinstance(val, list):
            return val
        val = dest.get(key)
        if isinstance(val, list):
            return val
    return []

def _call_google_generate_content(model: str, api_key: str, user_msg: str, max_tokens: int) -> str:
    _gemini_wait_for_slot(api_key)
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    payload = _build_google_generate_content_request(user_msg, max_tokens)
    resp = requests.post(
        url,
        params={"key": api_key},
        json=payload,
        timeout=AI_TIMEOUT_SEC,
    )
    if resp.status_code >= 400:
        try:
            detail = resp.json()
        except Exception:
            detail = resp.text[:500]
        raise RuntimeError(f"HTTP {resp.status_code}: {detail}")

    return _gemini_extract_text_from_response(resp.json())

def _get_ai_client(tag: str) -> "AIClient | None":
    """Get or create an AI client by tag: 'nim_0', 'nim_1', 'gemini_0', ..."""
    global _ai_clients
    if _ai_tag_disabled(tag):
        return None
    if tag in _ai_clients:
        return _ai_clients[tag]

    client = None
    if tag.startswith("nim_"):
        idx = int(tag.split("_")[1])
        keys = _valid_nvidia_api_keys()
        if idx < len(keys):
            key = keys[idx]
            if key:
                client = AIClient(
                    base_url="https://integrate.api.nvidia.com/v1",
                    api_key=key,
                )
    elif tag.startswith("gemini_"):
        idx = int(tag.split("_")[1])
        keys = _valid_gemini_api_keys()
        if idx < len(keys):
            key = keys[idx]
            if key:
                client = AIClient(
                    base_url="https://generativelanguage.googleapis.com/v1beta/openai/",
                    api_key=key,
                )

    if client:
        _ai_clients[tag] = client
    return client

_AI_CACHE_METRIC_KEYS = (
    "Current Price", "MA 20", "MA 50", "MA 200",
    "RSI 14", "ADX 14", "+DI 14", "-DI 14",
    "MACD Line", "MACD Hist", "Vol Ratio 20",
    "ATR 14", "NATR 14", "52W High Dist%", "20D Breakout%",
    "1D%", "1W%", "1M%", "3M%",
    "BB %B", "BB Width", "BB Width Pctl", "BB Squeeze",
    "Cam H3", "Cam H4", "Cam L3", "Cam L4",
    "Setup Signal", "Signal Quality", "Signal Regime",
    "Win Prob%", "Hist Precision%", "Exp 5D%", "Exp 10D%", "WF Samples",
    "Fundamental Score", "Fundamental Quality Tag", "Fundamental Risk Tag",
    "Investability Tag", "Early Entry OK", "Growth Tag", "Valuation Tag",
    "Balance Sheet Tag", "Cashflow Tag", "Ownership Tag",
    "Market Cap Cr", "PE", "PB", "EV/EBITDA", "Sales Growth TTM%",
    "Profit Growth TTM%", "Debt/Equity", "CFO/PAT", "Promoter Holding%",
    "Promoter Pledge%", "Tech + Fundamental Score", "Decision Guardrail",
    "Final Confidence Tag", "Final Signal",
)

def _ai_norm_metric(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        try:
            fval = float(value)
        except (TypeError, ValueError):
            return None
        return round(fval, 4) if math.isfinite(fval) else None
    text = str(value).strip()
    return text or None

def _ai_cache_key(symbol: str, rule_signal: str, m: dict) -> str:
    snapshot = {
        "symbol": symbol,
        "rule_signal": rule_signal,
        "signal_profile": SIGNAL_PROFILE,
        "metrics": {k: _ai_norm_metric(m.get(k)) for k in _AI_CACHE_METRIC_KEYS},
    }
    return json.dumps(snapshot, sort_keys=True, separators=(",", ":"))

def _ai_bb_context(m: dict) -> str:
    pctb = _ai_norm_metric(m.get("BB %B"))
    width = _ai_norm_metric(m.get("BB Width"))
    width_pctl = _ai_norm_metric(m.get("BB Width Pctl"))
    squeeze = m.get("BB Squeeze")
    if pctb is None and width is None and width_pctl is None and squeeze is None:
        return "unavailable"

    states = []
    if squeeze is True:
        states.append("squeeze active")
    elif squeeze is False:
        states.append("squeeze inactive")

    if pctb is not None:
        if pctb >= 1.0:
            states.append("upper-band stretch")
        elif pctb <= 0.0:
            states.append("lower-band oversold")
        elif pctb >= 0.8:
            states.append("upper-half push")
        elif pctb <= 0.2:
            states.append("lower-half weakness")
        else:
            states.append("mid-band")

    return (
        f"%B={pctb if pctb is not None else 'N/A'}, "
        f"width={width if width is not None else 'N/A'}, "
        f"width_pctl={width_pctl if width_pctl is not None else 'N/A'}, "
        f"squeeze={'Yes' if squeeze is True else 'No' if squeeze is False else 'N/A'}"
        + (f" | {'; '.join(states)}" if states else "")
    )

def _ai_camarilla_context(m: dict) -> str:
    price = _ai_norm_metric(m.get("Current Price"))
    h3 = _ai_norm_metric(m.get("Cam H3"))
    h4 = _ai_norm_metric(m.get("Cam H4"))
    l3 = _ai_norm_metric(m.get("Cam L3"))
    l4 = _ai_norm_metric(m.get("Cam L4"))
    if all(v is None for v in (h3, h4, l3, l4)):
        return "unavailable"

    state = "inside range"
    if price is not None:
        if h4 is not None and price >= h4:
            state = "at or above H4 resistance"
        elif h3 is not None and price >= h3:
            state = "between H3 and H4 breakout zone"
        elif l4 is not None and price <= l4:
            state = "at or below L4 breakdown zone"
        elif l3 is not None and price <= l3:
            state = "between L4 and L3 reversal zone"

    return (
        f"H3={h3 if h3 is not None else 'N/A'}, "
        f"H4={h4 if h4 is not None else 'N/A'}, "
        f"L3={l3 if l3 is not None else 'N/A'}, "
        f"L4={l4 if l4 is not None else 'N/A'} | {state}"
    )

# -- AI Decision Criteria (designed for Indian equity technical analysis) ------
# The AI evaluates 6 dimensions independently:
#   1. TREND  -- MA stack alignment (price vs MA20/50/200)
#   2. MOMENTUM -- RSI zone + ADX trend strength
#   3. ACCELERATION -- MACD line & histogram direction
#   4. VOLUME -- volume ratio vs 20-day average
#   5. POSITION -- 52W high distance + 20D breakout level
#   6. RETURNS -- 1M / 3M short-term price performance
# Then cross-validates against the rule-based GAS signal to detect divergence
# Output: 7-tier scale  STRONG BUY > BUY > ACCUMULATE > HOLD > REDUCE > SELL > STRONG SELL

_AI_SYSTEM_PROMPT_LEGACY_REFERENCE = """You are a senior quantitative analyst specializing in Indian equity markets (NSE/BSE).
You are a deterministic scoring engine, not a storyteller.
Generate an AI trading decision by following ALL 6 steps exactly. Do NOT skip, assume, interpolate, or guess.
Use only the supplied metrics. Never invent missing values, hidden context, news, fundamentals, sector narratives, management quality, macro views, targets, or sentiment.
If a statement cannot be tied directly to an explicit supplied metric, do not say it.
Be deterministic: if the inputs are unchanged, return the same decision and same scoring logic.
Perform the 6 scoring steps internally. Never print the working, never say "Step 1/2/3", and never restate the rubric in the answer.
Treat the rule-based signal as a cross-check, not as ground truth.
Use Bollinger and Camarilla context as tie-breakers and conviction modifiers, not as primary trend replacements.

----------------------------------------------------------------------
SOURCE-OF-TRUTH RULES  (anti-hallucination, mandatory)
----------------------------------------------------------------------
1. Only the metrics in the user message are real. Everything else is unknown.
2. Never infer missing values from ticker, company name, prior knowledge, common patterns, or market intuition.
3. Missing data is never bullish by itself. If important inputs are missing or mixed, bias toward lower conviction.
4. Do not mention anything about earnings, news, management, valuation, sector outlook, FII/DII flow, options, or candle patterns unless those exact metrics are present in the input. They are not present here, so never mention them.
5. Do not output future certainty. This is a technical ranking, not a guarantee.
6. The reason, bull, and bear text must be grounded in supplied metrics only.
7. Use only these decisions: STRONG BUY, BUY, ACCUMULATE, HOLD, REDUCE, SELL, STRONG SELL.

----------------------------------------------------------------------
MISSING-DATA SAFETY RULES  (mandatory)
----------------------------------------------------------------------
- If Current Price is missing, do not answer. Return HOLD with confidence 30 and reason "Insufficient verified inputs".
- If two or more of these are missing: MA 50, RSI 14, ADX 14, MACD Line, Vol Ratio 20, 20D Breakout%, 52W High Dist%, then cap the best possible decision at HOLD.
- If MA 200 is missing, never output BUY or STRONG BUY. Cap bullish output at ACCUMULATE.
- If BB or Camarilla data is missing, skip only that context adjustment. Do not invent a substitute.
- If walk-forward fields are weak/missing or Signal Quality is weak, reduce conviction; do not compensate with narrative language.

----------------------------------------------------------------------
STEP 1 -- SCORE EACH DIMENSION  (-2 · -1 · 0 · +1 · +2)
----------------------------------------------------------------------

DIM 1 · TREND (MA Stack)
  +2 -> Price > MA20 > MA50 > MA200  (perfect bull stack)
  +1 -> Price > MA50 AND Price > MA200  (bullish but not perfect)
   0 -> Price > MA50, MA200 missing  OR  mixed signals
  -1 -> Price < MA50  (lost mid-term trend)
  -2 -> Price < MA20 < MA50 < MA200  (full bear stack)

DIM 2 · MOMENTUM (RSI 14)
  +2 -> 55 <= RSI <= 70  (healthy bullish, room to run)
  +1 -> 50 <= RSI < 55  (neutral-bullish)
   0 -> 40 <= RSI < 50  (neutral)  OR  RSI < 32 above MA50 (oversold bounce watch)
  -1 -> 32 <= RSI < 40  OR  RSI > 78  (bearish momentum OR overbought trap)
  -2 -> RSI < 32 AND Price < MA50  (oversold in downtrend -- no bounce case)
  NOTE: RSI < 32 above MA50 = 0 (oversold opportunity); below MA50 = -2 (downtrend)

DIM 3 · TREND STRENGTH (ADX 14)
  +2 -> ADX > 25  (strong directional trend)
  +1 -> 20 <= ADX <= 25  (developing trend)
   0 -> 16 <= ADX < 20  (weak trend forming)
  -1 -> ADX < 16  (no trend -- choppy, penalise all directional calls)

DIM 4 · MACD ACCELERATION
  +2 -> MACD Line > 0  AND  Histogram positive AND increasing  (accelerating bull)
  +1 -> MACD Line > 0  AND  Histogram positive but flat/shrinking  (fading momentum)
   0 -> MACD Line > 0  BUT  Histogram negative  (bullish line, losing steam)
  -1 -> MACD Line <= 0  (bearish)
  NOTE: Histogram shrinking toward zero = early warning even if line still positive

DIM 5 · VOLUME PARTICIPATION
  +2 -> Vol Ratio > 1.5  (strong institutional interest)
  +1 -> 1.2 <= Vol Ratio <= 1.5  (above-average participation)
   0 -> 0.8 <= Vol Ratio < 1.2  (average)
  -1 -> Vol Ratio < 0.8  (weak -- move lacks conviction)
  PENALTY: If 1D% < 0 AND Vol Ratio > 1.2 -> subtract 1 extra point (distribution day)

DIM 6 · PRICE POSITION (52W High + Breakout)
  +2 -> 20D Breakout > 0%  AND  52W dist > 5%  (breaking out with room to run)
  +1 -> 20D Breakout > 0%  (breakout regardless of 52W distance)
   0 -> 20D Breakout near 0%  AND  52W dist 5-20%  (consolidating)
  -1 -> Near 52W high (dist < 3%) without breakout  OR  52W dist > 25% (deep hole)
  HARD CAP: If 52W dist < 2% (AT resistance) -> DIM 6 score = 0 max regardless

----------------------------------------------------------------------
STEP 2 -- RETURN MOMENTUM BONUS  (adjust total ±1, apply max one of each)
----------------------------------------------------------------------
  +1 if  1D% > 0  AND  1W% > 1D%  AND  1M% > 1W%  (accelerating bull momentum)
  +1 if  3M% > 0  AND  1M% > 3M%  (pace picking up vs medium term)
  -1 if  1D% < 0  AND  1W% < 1D%  AND  1M% < 1W%  (accelerating bear momentum)
  -1 if  3M% > 0  AND  1M% < 0  (reversal warning inside medium-term uptrend)

----------------------------------------------------------------------
STEP 3 -- CONTEXT ADJUSTMENT  (Bollinger + Camarilla, max +1 and max -1)
----------------------------------------------------------------------
  +1 if Bollinger squeeze is active AND breakout > 0 AND Vol Ratio > 1.2
  +1 if Price > Cam H3 AND ADX >= 20 AND MACD Line > 0
  -1 if BB %B >= 1.0 AND RSI > 78
  -1 if Price >= Cam H4 without breakout confirmation
  -1 if Price <= Cam L3 AND Trend dim <= 0
  If Bollinger/Camarilla data is missing, skip that part.

----------------------------------------------------------------------
STEP 4 -- DECISION MAPPING  (total_score = sum of 6 dims + bonus + context)
----------------------------------------------------------------------
  >=  +8  ->  STRONG BUY    (overwhelming evidence, max conviction)
  +5 to +7  ->  BUY         (clear edge, enter with normal sizing)
  +2 to +4  ->  ACCUMULATE  (bullish bias, build on dips)
  -1 to +1  ->  HOLD        (balanced, no new entry)
  -2 to -4  ->  REDUCE      (bearish lean, trim position)
  -5 to -7  ->  SELL        (clear downtrend, exit)
  <=  -8  ->  STRONG SELL   (severe breakdown, exit immediately)

OVERRIDE RULES (apply after mapping):
  - ADX < 16 -> force HOLD bias unless score >= +6 or <= -6 (no trend = no conviction)
  - MA200 missing -> cap at ACCUMULATE / REDUCE (insufficient history)
  - RSI > 78 AND Trend dim <= 0 -> bias SELL (overbought in weak trend)
  - Breakout > 0  AND  Vol Ratio > 1.25  AND  ADX > 25 -> bias BUY (power breakout)
  - BSE-only stock (symbol "BSE:...") AND ADX < 20 -> cap at ACCUMULATE/REDUCE (thin liquidity)
  - If Rule Signal is BREAKOUT/STRONG BUY but your decision is HOLD/REDUCE/SELL, explain the disagreement explicitly

----------------------------------------------------------------------
STEP 5 -- CONFIDENCE CALCULATION
----------------------------------------------------------------------
  base_confidence = abs(total_score) × 10   [raw 0-120, will be adjusted]
  Add:  +10 if ADX > 25   (confirmed trend)
        +10 if Vol Ratio > 1.25   (volume confirms move)
        + 8 if Bollinger/Camarilla context supports your decision
        +10 if Rule Signal AGREES with your decision
  Subtract:  -15 if ADX < 16   (directionless market)
             -15 if Rule Signal DISAGREES with your decision
             -10 if BSE-only AND ADX < 20
             - 8 if RSI > 78 AND near 52W high (overbought + resistance)
             - 8 if Price >= Cam H4 without breakout confirmation
  Clamp final confidence to [30, 95]

----------------------------------------------------------------------
STEP 6 -- REQUIRED OUTPUT FORMAT
----------------------------------------------------------------------
reason FORMAT (strictly follow this pattern):
  "Trend {score}, Momentum {score}, ADX {score}, MACD {score}, Vol {score}, Pos {score}, Ctx {score} -> {one-line conclusion}"
  If decision conflicts with Rule Signal, APPEND:
  " | AI differs: {specific reason e.g. ADX too weak for rule BUY}"

bull: strongest single bullish fact in <=10 words
bear: biggest single risk or red flag in <=10 words

----------------------------------------------------------------------
FINAL OUTPUT -- strict JSON only, no extra text:
- Output one minified JSON object only.
- First character must be { and last character must be }.
- Do not use markdown fences.
- Do not output step-by-step reasoning, headings, bullet points, or commentary.
- Do not repeat the rubric.
{"score":6,"decision":"BUY","confidence":78,"reason":"Trend +2, Momentum +1, ADX +2, MACD +1, Vol 0, Pos 0, Ctx +1 -> strong trend confirmed","bull":"Perfect MA stack with ADX 24 trending","bear":"Volume below average, lacks institutional push"}"""

_AI_SYSTEM_PROMPT = """You are a deterministic Indian-equity technical + fundamental scoring engine.
Use only supplied fields. Never use outside knowledge, news, unsupplied fundamentals, sector stories, or guesses.
Fundamental fields, when present in the payload, are explicit inputs and may be used only as quality/risk modifiers.
Do all scoring internally. Never print steps, headings, markdown, or commentary.
Return exactly one minified JSON object with keys: score, decision, confidence, reason, bull, bear.
Allowed decisions: STRONG BUY, BUY, ACCUMULATE, HOLD, REDUCE, SELL, STRONG SELL.

Scoring rules:
- Trend: +2 if price>ma20>ma50>ma200; +1 if price>ma50 and price>ma200; 0 if mixed or ma200 missing; -1 if price<ma50; -2 if price<ma20<ma50<ma200.
- RSI: +2 if 55-70; +1 if 50-54.99; 0 if 40-49.99 or rsi<32 while above ma50; -1 if 32-39.99 or rsi>78; -2 if rsi<32 and price<ma50.
- ADX: +2 if >25; +1 if 20-25; 0 if 16-19.99; -1 if <16.
- MACD: +2 if macd_line>0 and macd_hist>0 and accelerating; +1 if macd_line>0 and macd_hist>0; 0 if macd_line>0 and macd_hist<=0; -1 if macd_line<=0.
- Volume: +2 if vol_ratio>1.5; +1 if 1.2-1.5; 0 if 0.8-1.19; -1 if <0.8; subtract 1 more if 1d<0 and vol_ratio>1.2.
- Position: +2 if breakout20d>0 and dist52w>5; +1 if breakout20d>0; 0 if breakout20d near 0 and dist52w 5-20; -1 if dist52w<3 without breakout or dist52w>25; cap this dimension at 0 if dist52w<2.
- Bonus: +1 if 1d>0 and 1w>1d and 1m>1w; +1 if 3m>0 and 1m>3m; -1 if 1d<0 and 1w<1d and 1m<1w; -1 if 3m>0 and 1m<0.
- Context: +1 if bb_squeeze and breakout20d>0 and vol_ratio>1.2; +1 if price>cam_h3 and adx>=20 and macd_line>0; -1 if bb_pctb>=1 and rsi>78; -1 if price>=cam_h4 without breakout; -1 if price<=cam_l3 and trend<=0.

Decision map:
- score>=8 STRONG BUY
- 5..7 BUY
- 2..4 ACCUMULATE
- -1..1 HOLD
- -4..-2 REDUCE
- -7..-5 SELL
- <=-8 STRONG SELL

Safety caps:
- If current_price is missing: HOLD, confidence 30, reason "Insufficient verified inputs".
- If two or more key fields are missing among ma50,rsi14,adx14,macd_line,vol_ratio20,breakout20d,dist52w: cap best decision at HOLD.
- If ma200 missing: cap bullish output at ACCUMULATE.
- If adx<16: force HOLD bias unless score>=6 or score<=-6.
- If rsi>78 and trend<=0: bias bearish.
- If breakout20d>0 and vol_ratio>1.25 and adx>25: bias bullish.
- If supplied Fundamental Risk Tag is HIGH or Investability Tag is AVOID: cap bullish output at ACCUMULATE unless technical score is extremely strong.
- If supplied Fundamental Score is >=70 with LOW/MED fundamental risk, it may improve BUY/ACCUMULATE confidence but must not override missing or bearish technical inputs.
- If supplied Fundamental Score is missing or UNKNOWN, ignore fundamentals and do not invent a business-quality opinion.
- If supplied Decision Guardrail starts with BLOCK: cap best decision at HOLD.
- If supplied Decision Guardrail starts with CAUTION: cap best decision at ACCUMULATE unless supplied Final Signal is HIGH CONVICTION BUY.
- If supplied Final Confidence Tag is LOW or BLOCKED: do not output BUY or STRONG BUY.

Confidence:
- start at abs(score)*10, then adjust up for strong adx/volume/context/rule agreement/fundamental support and down for weak adx/rule disagreement/overbought at resistance/fundamental risk. Clamp 30..95.

Text rules:
- reason format: "Trend X, Momentum X, ADX X, MACD X, Vol X, Pos X, Ctx X -> conclusion"
- bull and bear must each be <=10 words and grounded in supplied fields only.
- Output one minified JSON object only. First character must be { and last character must be }. No prose before or after."""

_AI_RESPONSE_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "score": {"type": "integer", "minimum": -20, "maximum": 20},
        "decision": {
            "type": "string",
            "enum": ["STRONG BUY", "BUY", "ACCUMULATE", "HOLD", "REDUCE", "SELL", "STRONG SELL"],
        },
        "confidence": {"type": "integer", "minimum": 30, "maximum": 95},
        "reason": {"type": "string", "maxLength": 220},
        "bull": {"type": "string", "maxLength": 60},
        "bear": {"type": "string", "maxLength": 60},
    },
    "required": ["score", "decision", "confidence", "reason", "bull", "bear"],
}

def _ai_request_payload(symbol: str, name: str, m: dict, rule_signal: str, bb_context: str, cam_context: str, exchange_context: str, di_gap):
    return {
        "symbol": symbol,
        "name": name,
        "rule_signal": rule_signal,
        "signal_profile": SIGNAL_PROFILE,
        "exchange_context": exchange_context,
        "signals": {
            "final": m.get("Signal"),
            "setup": m.get("Setup Signal"),
            "core": m.get("Core Signal"),
            "quality": m.get("Signal Quality"),
            "regime": m.get("Signal Regime"),
        },
        "walkforward": {
            "win_prob_pct": m.get("Win Prob%"),
            "hist_precision_pct": m.get("Hist Precision%"),
            "exp_5d_pct": m.get("Exp 5D%"),
            "exp_10d_pct": m.get("Exp 10D%"),
            "samples": m.get("WF Samples"),
        },
        "metrics": {
            "current_price": m.get("Current Price"),
            "ma20": m.get("MA 20"),
            "ma50": m.get("MA 50"),
            "ma200": m.get("MA 200"),
            "rsi14": m.get("RSI 14"),
            "adx14": m.get("ADX 14"),
            "plus_di14": m.get("+DI 14"),
            "minus_di14": m.get("-DI 14"),
            "di_gap": di_gap,
            "macd_line": m.get("MACD Line"),
            "macd_hist": m.get("MACD Hist"),
            "vol_ratio20": m.get("Vol Ratio 20"),
            "atr14": m.get("ATR 14"),
            "natr14": m.get("NATR 14"),
            "dist52w": m.get("52W High Dist%"),
            "breakout20d": m.get("20D Breakout%"),
            "ret_1d": m.get("1D%"),
            "ret_1w": m.get("1W%"),
            "ret_1m": m.get("1M%"),
            "ret_3m": m.get("3M%"),
            "bb_pctb": m.get("BB %B"),
            "bb_width": m.get("BB Width"),
            "bb_width_pctl": m.get("BB Width Pctl"),
            "bb_squeeze": m.get("BB Squeeze"),
            "cam_h3": m.get("Cam H3"),
            "cam_h4": m.get("Cam H4"),
            "cam_l3": m.get("Cam L3"),
            "cam_l4": m.get("Cam L4"),
        },
        "contexts": {
            "bollinger": bb_context,
            "camarilla": cam_context,
        },
        "fundamentals": {
            "source": m.get("Fundamental Source"),
            "freshness": m.get("Fundamental Freshness"),
            "score": m.get("Fundamental Score"),
            "quality": m.get("Fundamental Quality Tag"),
            "risk": m.get("Fundamental Risk Tag"),
            "investability": m.get("Investability Tag"),
            "early_entry_ok": m.get("Early Entry OK"),
            "profitability": m.get("Profitability Tag"),
            "growth": m.get("Growth Tag"),
            "valuation": m.get("Valuation Tag"),
            "balance_sheet": m.get("Balance Sheet Tag"),
            "cashflow": m.get("Cashflow Tag"),
            "ownership": m.get("Ownership Tag"),
            "market_cap_cr": m.get("Market Cap Cr"),
            "pe": m.get("PE"),
            "pb": m.get("PB"),
            "ev_ebitda": m.get("EV/EBITDA"),
            "roe_pct": m.get("ROE%"),
            "roce_pct": m.get("ROCE%"),
            "sales_growth_ttm_pct": m.get("Sales Growth TTM%"),
            "profit_growth_ttm_pct": m.get("Profit Growth TTM%"),
            "debt_equity": m.get("Debt/Equity"),
            "cfo_pat": m.get("CFO/PAT"),
            "promoter_holding_pct": m.get("Promoter Holding%"),
            "promoter_pledge_pct": m.get("Promoter Pledge%"),
        },
        "combined": {
            "tech_fundamental_score": m.get("Tech + Fundamental Score"),
            "decision_guardrail": m.get("Decision Guardrail"),
            "final_confidence_tag": m.get("Final Confidence Tag"),
            "final_signal": m.get("Final Signal"),
            "final_signal_reason": m.get("Final Signal Reason"),
        },
    }

def _ai_repair_non_json_response(text: str) -> dict | None:
    flat = re.sub(r"\s+", " ", str(text or "")).strip()
    if not flat:
        return None

    valid = ("STRONG BUY", "STRONG SELL", "ACCUMULATE", "REDUCE", "SELL", "BUY", "HOLD")
    decision = None
    explicit_patterns = (
        r"(?:final\s+)?(?:decision|recommendation|verdict|output)\s*[:=-]?\s*(STRONG BUY|STRONG SELL|ACCUMULATE|REDUCE|SELL|BUY|HOLD)\b",
        r"^\s*(STRONG BUY|STRONG SELL|ACCUMULATE|REDUCE|SELL|BUY|HOLD)\b",
    )
    for pattern in explicit_patterns:
        matches = re.findall(pattern, flat, flags=re.I)
        if matches:
            decision = matches[-1].upper()
            break
    if decision not in valid:
        return None

    def _extract_number(patterns, default=None):
        for pattern in patterns:
            match = re.search(pattern, flat, flags=re.I)
            if match:
                try:
                    return float(match.group(1))
                except (TypeError, ValueError):
                    continue
        return default

    score = _extract_number((
        r"(?:total\s+score|score)\s*[:=-]?\s*(-?\d+(?:\.\d+)?)",
    ))
    confidence = _extract_number((
        r"confidence\s*[:=-]?\s*(\d+(?:\.\d+)?)",
    ), default=50.0)

    bull = ""
    bear = ""
    reason = ""
    bull_match = re.search(r"\bbull\s*[:=-]\s*(.+?)(?:\s+\bbear\b|$)", flat, flags=re.I)
    bear_match = re.search(r"\bbear\s*[:=-]\s*(.+)$", flat, flags=re.I)
    reason_match = re.search(r"(?:reason|conclusion)\s*[:=-]\s*(.+?)(?:\s+\b(?:bull|bear|confidence|score|decision)\b|$)", flat, flags=re.I)
    if bull_match:
        bull = bull_match.group(1).strip()[:60]
    if bear_match:
        bear = bear_match.group(1).strip()[:60]
    if reason_match:
        reason = reason_match.group(1).strip()[:220]

    return {
        "decision": decision,
        "score": int(round(score)) if score is not None else 0,
        "confidence": int(round(confidence)) if confidence is not None else 50,
        "reason": reason,
        "bull": bull,
        "bear": bear,
    }

def _ai_cache_store(cache_key: str, result: dict):
    _ai_cache[cache_key] = result
    while len(_ai_cache) > AI_CACHE_MAX:
        _ai_cache.pop(next(iter(_ai_cache)))

def _ai_build_context(symbol: str, name: str, m: dict, rule_signal: str) -> dict | None:
    if not AI_ENABLED:
        return None
    if not m.get("Current Price") or rule_signal in ("No Data", "Error", "Pending...", "Symbol Not Found"):
        return None

    cache_key = _ai_cache_key(symbol, rule_signal, m)
    cached = _ai_cache.get(cache_key)
    if cached:
        return {
            "symbol": symbol,
            "name": name,
            "cache_key": cache_key,
            "cached": cached,
        }

    pdi = m.get("+DI 14")
    mdi = m.get("-DI 14")
    di_gap = None
    if pdi is not None and mdi is not None:
        try:
            di_gap = round(float(pdi) - float(mdi), 2)
        except (TypeError, ValueError):
            di_gap = None

    bb_context = _ai_bb_context(m)
    cam_context = _ai_camarilla_context(m)
    exchange_context = "BSE-only / thinner liquidity risk" if symbol.startswith("BSE:") else "NSE / standard liquidity"
    payload = _ai_request_payload(symbol, name, m, rule_signal, bb_context, cam_context, exchange_context, di_gap)
    user_msg = (
        "Evaluate the stock strictly from this payload. "
        "Use only these fields. Return one minified JSON object only.\n"
        + json.dumps(payload, sort_keys=True, separators=(",", ":"))
    )
    return {
        "symbol": symbol,
        "name": name,
        "cache_key": cache_key,
        "cached": None,
        "user_msg": user_msg,
    }

def _ai_result_from_raw(raw: str) -> dict:
    cleaned = re.sub(r'^```(?:json)?\s*', '', str(raw or "").strip())
    cleaned = re.sub(r'\s*```\s*$', '', cleaned)
    json_str = None
    start = cleaned.find('{')
    if start >= 0:
        depth = 0
        for i in range(start, len(cleaned)):
            if cleaned[i] == '{':
                depth += 1
            elif cleaned[i] == '}':
                depth -= 1
            if depth == 0:
                json_str = cleaned[start:i+1]
                break
        if json_str is None:
            partial = cleaned[start:]
            in_str = False
            escaped = False
            for ch in partial:
                if escaped:
                    escaped = False
                    continue
                if ch == "\\":
                    escaped = True
                    continue
                if ch == '"':
                    in_str = not in_str
            if in_str:
                partial += '"'
            opens = partial.count('{') - partial.count('}')
            partial += '}' * max(opens, 1)
            json_str = partial

    if not json_str:
        obj = _ai_repair_non_json_response(cleaned)
        if obj is None:
            raise ValueError(f"No JSON in response: {str(raw)[:200]}")
    else:
        try:
            obj = json.loads(json_str)
        except json.JSONDecodeError:
            dec_m = re.search(r'"decision"\s*:\s*"([^"]+)"', json_str)
            sco_m = re.search(r'"score"\s*:\s*([-\d.]+)', json_str)
            con_m = re.search(r'"confidence"\s*:\s*([-\d.]+)', json_str)
            if dec_m:
                obj = {
                    "decision": dec_m.group(1),
                    "score": int(float(sco_m.group(1))) if sco_m else 0,
                    "confidence": int(float(con_m.group(1))) if con_m else 50,
                    "reason": "", "bull": "", "bear": "",
                }
            else:
                obj = _ai_repair_non_json_response(cleaned)
                if obj is None:
                    raise ValueError(f"Cannot parse JSON: {json_str[:200]}")

    valid = {"STRONG BUY","BUY","ACCUMULATE","HOLD","REDUCE","SELL","STRONG SELL"}
    decision = str(obj.get("decision", "HOLD")).strip().upper()
    if decision not in valid:
        decision = "HOLD"
    bull = str(obj.get("bull", "")).strip()[:60]
    bear = str(obj.get("bear", "")).strip()[:60]
    reason_core = str(obj.get("reason", "")).strip()
    reason_parts = []
    if reason_core:
        reason_parts.append(reason_core)
    if bull:
        reason_parts.append(f"Bull: {bull}")
    if bear:
        reason_parts.append(f"Bear: {bear}")
    return {
        "decision": decision,
        "reason": " | ".join(reason_parts)[:220],
        "confidence": min(95, max(30, int(float(obj.get("confidence", 50))))),
        "score": int(float(obj.get("score", 0))) if obj.get("score") is not None else None,
    }

def _collect_ai_parallel(need_ai: dict, sym_to_name: dict) -> dict:
    results = {}
    _ai_done = [0]
    _ai_total = len(need_ai)

    def _fetch_one_ai(item):
        sym, m = item
        name = sym_to_name.get(sym, "")
        try:
            return sym, _ai_signal(sym, name, m, m.get("Signal", ""))
        except Exception as e:
            print(f"      [FAIL] AI {sym}: {e}")
            return sym, {}

    with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as pool:
        futures = {pool.submit(_fetch_one_ai, item): item[0] for item in need_ai.items()}
        for f in as_completed(futures):
            sym, ai = f.result()
            _ai_done[0] += 1
            if ai:
                results[sym] = ai
            if _ai_done[0] % 25 == 0 or _ai_done[0] == _ai_total:
                print(f"    [AI] AI progress: {_ai_done[0]}/{_ai_total} ({len(results)} OK)")
    return results

def _collect_ai_google_batch(need_ai: dict, sym_to_name: dict) -> dict:
    global _ai_rr_counter

    results = {}
    contexts = []
    cached_hits = 0
    for sym, m in need_ai.items():
        ctx = _ai_build_context(sym, sym_to_name.get(sym, ""), m, m.get("Signal", ""))
        if not ctx:
            continue
        if ctx.get("cached"):
            results[sym] = ctx["cached"]
            cached_hits += 1
        else:
            contexts.append(ctx)

    if cached_hits:
        print(f"    [AI] Cache hits reused: {cached_hits}")
    if not contexts:
        return results

    rr_seed = _ai_rr_counter
    _ai_rr_counter += max(1, len(contexts))
    provider_entries = _build_ai_provider_entries("google", rr_seed)
    if not provider_entries:
        return results

    chunks = _gemini_batch_chunks(contexts)
    print(f"    [AI] Gemini batch mode: {len(contexts)} live requests in {len(chunks)} batch chunk(s)")
    batch_ok = 0
    unresolved = []
    terminal_states = {"JOB_STATE_SUCCEEDED", "JOB_STATE_FAILED", "JOB_STATE_CANCELLED", "JOB_STATE_EXPIRED"}

    for chunk_idx, chunk in enumerate(chunks, start=1):
        tag, model, label, _provider = provider_entries[(chunk_idx - 1) % len(provider_entries)]
        if _ai_tag_disabled(tag):
            unresolved.extend(ctx for ctx, _inline in chunk)
            continue
        api_key = _gemini_key_for_tag(tag)
        if not api_key:
            unresolved.extend(ctx for ctx, _inline in chunk)
            continue

        inline_requests = [inline for _ctx, inline in chunk]
        display_name = f"tracker-ai-{int(time.time())}-{chunk_idx}"
        try:
            job = _call_google_batch_generate_content(model, api_key, inline_requests, display_name)
            job_name = str(job.get("name") or "").strip()
            if not job_name:
                raise RuntimeError(f"Batch job missing name: {str(job)[:300]}")

            started = time.time()
            state = _gemini_batch_state_name(job)
            while state not in terminal_states:
                if time.time() - started >= max(10, int(GEMINI_BATCH_MAX_WAIT_SEC or 0)):
                    raise TimeoutError(f"Batch wait exceeded {GEMINI_BATCH_MAX_WAIT_SEC}s ({job_name})")
                time.sleep(max(1, int(GEMINI_BATCH_POLL_SEC or 10)))
                job = _get_google_batch_job(job_name, api_key)
                state = _gemini_batch_state_name(job)

            if state != "JOB_STATE_SUCCEEDED":
                raise RuntimeError(f"{job_name} ended with state {state or 'UNKNOWN'}")

            inline_responses = _gemini_batch_inline_responses(job)
            if not inline_responses:
                raise RuntimeError(f"{job_name} returned no inline responses")

            index_map = {ctx["symbol"]: ctx for ctx, _inline in chunk}
            ordered_contexts = [ctx for ctx, _inline in chunk]
            chunk_seen = set()
            for idx, item in enumerate(inline_responses):
                key = None
                if isinstance(item, dict):
                    metadata = item.get("metadata") or item.get("requestMetadata") or item.get("inputMetadata") or {}
                    key = metadata.get("key") or item.get("key")
                ctx = index_map.get(str(key)) if key else None
                if ctx is None and idx < len(ordered_contexts):
                    ctx = ordered_contexts[idx]
                if ctx is None:
                    continue
                chunk_seen.add(ctx["symbol"])
                if isinstance(item, dict) and item.get("error"):
                    unresolved.append(ctx)
                    continue
                response_obj = None
                if isinstance(item, dict):
                    response_obj = item.get("response") or item.get("inlineResponse")
                    if response_obj is None and item.get("candidates"):
                        response_obj = item
                if not response_obj:
                    unresolved.append(ctx)
                    continue
                raw = _gemini_extract_text_from_response(response_obj)
                result = _ai_result_from_raw(raw)
                _ai_cache_store(ctx["cache_key"], result)
                results[ctx["symbol"]] = result
                batch_ok += 1
            for ctx, _inline in chunk:
                if ctx["symbol"] not in chunk_seen and ctx["symbol"] not in results:
                    unresolved.append(ctx)
            print(f"    [AI] Batch chunk {chunk_idx}/{len(chunks)} via {label} ({model}): {batch_ok}/{len(contexts)} OK")
        except Exception as e:
            if _is_permanent_ai_auth_error(e):
                _disable_ai_tag(tag, label, e)
            print(f"      [AI] Batch {label} failed (chunk {chunk_idx}/{len(chunks)}): {e}")
            unresolved.extend(ctx for ctx, _inline in chunk)

    if unresolved:
        fallback_items = OrderedDict()
        for ctx in unresolved:
            sym = ctx["symbol"]
            if sym in results or sym in fallback_items:
                continue
            fallback_items[sym] = need_ai[sym]
        if fallback_items:
            print(f"    [AI] Falling back to interactive mode for {len(fallback_items)} symbol(s)")
            results.update(_collect_ai_parallel(fallback_items, sym_to_name))
    return results

def _ai_signal(symbol: str, name: str, m: dict, rule_signal: str) -> dict:
    """
    Call the configured AI provider chain with round-robin load-balancing.
    The primary provider always runs first; the secondary provider is used only
    when AI_SECONDARY_ENABLED is True.
    Returns dict with keys: decision, reason, confidence (or empty on failure).
    """
    global _ai_rr_counter

    ctx = _ai_build_context(symbol, name, m, rule_signal)
    if not ctx:
        return {}
    if ctx.get("cached"):
        return ctx["cached"]

    primary_provider = _ai_primary_provider()
    secondary_provider = _ai_secondary_provider()
    user_msg = ctx["user_msg"]

    # ── Build provider chain: primary first, optional secondary second ───────
    # Each provider rotates across its own keys × models without mixing chains.
    rr_seed = _ai_rr_counter
    _ai_rr_counter += 1

    primary_entries = _build_ai_provider_entries(primary_provider, rr_seed)
    secondary_entries = _build_ai_provider_entries(secondary_provider, rr_seed) if AI_SECONDARY_ENABLED else []
    _apis = primary_entries + secondary_entries

    if not _apis:
        return {}

    last_idx = len(_apis) - 1
    for api_idx, (tag, model, label, provider) in enumerate(_apis):
        if _ai_tag_disabled(tag):
            continue
        extra_kwargs = {}
        provider_max_tokens = min(AI_MAX_TOKENS, 180) if provider == "nvidia" else AI_MAX_TOKENS
        api_client = None
        gemini_key = None
        if provider == "nvidia":
            api_client = _get_ai_client(tag)
            if not api_client:
                continue
            extra_kwargs["response_format"] = {"type": "json_object"}
            extra_kwargs.update(_nvidia_request_options(model))
        elif provider == "google":
            gemini_key = _gemini_key_for_tag(tag)
            if not gemini_key:
                continue
        for attempt in range(AI_MAX_RETRIES):
            try:
                if provider == "google":
                    raw = _call_google_generate_content(model, gemini_key, user_msg, provider_max_tokens)
                else:
                    resp = api_client.chat.completions.create(
                        model=model,
                        messages=[
                            {"role": "system", "content": _AI_SYSTEM_PROMPT},
                            {"role": "user",   "content": user_msg + "\nRespond ONLY with valid minified JSON. No prose. No markdown. No steps."},
                        ],
                        temperature=AI_TEMPERATURE,
                        top_p=AI_TOP_P,
                        max_tokens=provider_max_tokens,
                        timeout=AI_TIMEOUT_SEC,
                        **extra_kwargs,
                    )
                    raw = (resp.choices[0].message.content or "").strip()
                result = _ai_result_from_raw(raw)
                _ai_cache_store(ctx["cache_key"], result)
                if api_idx > 0 and LOG_AI_PROVIDER_SUCCESS:
                    print(f"      [AI] AI OK via {label} ({model})")
                return result
            except Exception as e:
                if _is_permanent_ai_auth_error(e):
                    _disable_ai_tag(tag, label, e)
                    print(f"      [AI] AI {label} failed ({symbol}): permanent API key/auth error")
                    if api_idx < last_idx:
                        next_label = _apis[api_idx + 1][2]
                        next_model = _apis[api_idx + 1][1]
                        print(f"      [RELOAD] Falling back to {next_label} ({next_model})...")
                    break
                if attempt < AI_MAX_RETRIES - 1:
                    time.sleep(2 ** attempt)
                else:
                    print(f"      [AI] AI {label} failed ({symbol}): {e}")
                    if api_idx < last_idx:
                        next_label = _apis[api_idx + 1][2]
                        next_model = _apis[api_idx + 1][1]
                        print(f"      [RELOAD] Falling back to {next_label} ({next_model})...")
    return {}

def _ai_style(decision: str):
    """Excel fill + font for AI Decision cell."""
    d = str(decision).upper()
    if d == "STRONG BUY":  return PatternFill("solid",fgColor="1B5E20"), Font(bold=True,color="FFFFFF")
    if d == "BUY":         return PatternFill("solid",fgColor="4CAF50"), Font(bold=True,color="FFFFFF")
    if d == "ACCUMULATE":  return PatternFill("solid",fgColor="A5D6A7"), Font(bold=True,color="1B5E20")
    if d == "HOLD":        return PatternFill("solid",fgColor="FFF9C4"), Font(color="F57F17")
    if d == "REDUCE":      return PatternFill("solid",fgColor="FFCC80"), Font(bold=True,color="E65100")
    if d == "SELL":        return PatternFill("solid",fgColor="EF9A9A"), Font(bold=True,color="B71C1C")
    if d == "STRONG SELL": return PatternFill("solid",fgColor="B71C1C"), Font(bold=True,color="FFFFFF")
    return None, None

def _ai_conf_color(conf: int | None):
    """Excel fill for AI Confidence% cell -- green->yellow->red gradient."""
    if conf is None: return None
    if conf >= 80: return PatternFill("solid", fgColor="C8E6C9")
    if conf >= 65: return PatternFill("solid", fgColor="FFF9C4")
    return PatternFill("solid", fgColor="FFCDD2")

# -----------------------------------------------------------------------------
# HELPERS
# -----------------------------------------------------------------------------
def ist_now(): return datetime.now(IST)
def fmt_dt(dt=None): return (dt or ist_now()).strftime("%Y-%m-%d %H:%M:%S")
_signal_snapshot_cache = None
_signal_snapshot_dirty = False

def _prev_business_day(d):
    cur = d - timedelta(days=1)
    while cur.weekday() >= 5:
        cur -= timedelta(days=1)
    return cur

def _completed_market_session_key(dt=None):
    if not OFF_MARKET_SIGNAL_SNAPSHOT_ENABLED:
        return None
    now = dt or ist_now()
    if now.tzinfo is None:
        now = IST.localize(now)
    now = now.astimezone(IST)
    open_time = dt_time(9, 15)
    close_time = dt_time(15, 30)
    today = now.date()
    if now.weekday() >= 5:
        return _prev_business_day(today).isoformat()
    if now.time() < open_time:
        return _prev_business_day(today).isoformat()
    if now.time() >= close_time:
        return today.isoformat()
    return None

def _load_signal_snapshot_cache():
    global _signal_snapshot_cache
    if _signal_snapshot_cache is not None:
        return _signal_snapshot_cache
    cache = {"version": 1, "sessions": {}}
    try:
        if os.path.exists(SIGNAL_SNAPSHOT_FILE):
            with open(SIGNAL_SNAPSHOT_FILE, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict) and isinstance(data.get("sessions"), dict):
                cache = data
    except Exception as e:
        print(f"  [WARN] Signal snapshot cache load failed: {e}")
    _signal_snapshot_cache = cache
    return _signal_snapshot_cache

def _prune_signal_snapshot_cache():
    cache = _load_signal_snapshot_cache()
    sessions = cache.setdefault("sessions", {})
    keys = sorted(k for k in sessions.keys() if isinstance(k, str))
    while len(keys) > SIGNAL_SNAPSHOT_KEEP_DAYS:
        old = keys.pop(0)
        sessions.pop(old, None)

def _save_signal_snapshot_cache():
    global _signal_snapshot_dirty
    if not _signal_snapshot_dirty:
        return
    cache = _load_signal_snapshot_cache()
    _prune_signal_snapshot_cache()
    tmp = SIGNAL_SNAPSHOT_FILE + ".tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(cache, fh, ensure_ascii=True, separators=(",", ":"))
        os.replace(tmp, SIGNAL_SNAPSHOT_FILE)
        _signal_snapshot_dirty = False
    except Exception as e:
        print(f"  [WARN] Signal snapshot cache save failed: {e}")
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass

def _rebuild_snapshot_signal(metrics):
    m = dict(metrics) if isinstance(metrics, dict) else {}
    if not m:
        return None
    price = m.get("Current Price")
    if price is None:
        return m
    if "Signal Regime" not in m or not m.get("Signal Regime"):
        m["Signal Regime"] = _classify_signal_regime(m)
    bundle = _signal_bundle_from_metrics(m)
    for key, value in bundle.items():
        m[key] = value
    return m

def _get_signal_snapshot_metrics(session_key, symbol):
    if not session_key:
        return None
    cache = _load_signal_snapshot_cache()
    session = cache.get("sessions", {}).get(session_key, {})
    metrics = session.get(symbol)
    return _rebuild_snapshot_signal(metrics) if isinstance(metrics, dict) else None

def _put_signal_snapshot_metrics(session_key, symbol, metrics):
    global _signal_snapshot_dirty
    if not session_key or not symbol or not isinstance(metrics, dict):
        return
    cache = _load_signal_snapshot_cache()
    sessions = cache.setdefault("sessions", {})
    session = sessions.setdefault(session_key, {})
    session[symbol] = dict(metrics)
    _signal_snapshot_dirty = True

def rnd(val, digits=2):
    if val is None or (isinstance(val,float) and math.isnan(val)): return None
    return round(float(val), digits)
def _proxy_url(target): return f"{PROXY_URL}?url={url_encode(target, safe='')}"
def _clean_sym(raw): return _html_mod.unescape(raw).strip().upper()
def _clean_name(raw): return _html_mod.unescape(raw).strip()

def _fundamental_cache_path():
    try:
        return os.path.join(_script_base_dir(), FUNDAMENTAL_CACHE_FILE)
    except Exception:
        return FUNDAMENTAL_CACHE_FILE

def _load_fundamental_cache():
    global _fundamental_cache
    if _fundamental_cache is not None:
        return _fundamental_cache
    cache = {"version": FUNDAMENTAL_CACHE_VERSION, "symbols": {}}
    path = _fundamental_cache_path()
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict) and isinstance(data.get("symbols"), dict):
                cache = data
    except Exception as e:
        print(f"  [WARN] Fundamental cache load failed: {e}")
    _fundamental_cache = cache
    return _fundamental_cache

def _save_fundamental_cache():
    global _fundamental_cache_dirty
    if not _fundamental_cache_dirty:
        return
    cache = _load_fundamental_cache()
    path = _fundamental_cache_path()
    tmp = path + ".tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(cache, fh, ensure_ascii=True, separators=(",", ":"))
        os.replace(tmp, path)
        _fundamental_cache_dirty = False
    except Exception as e:
        print(f"  [WARN] Fundamental cache save failed: {e}")
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass

_FUNDAMENTAL_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36"
)

def _fundamental_parser_available():
    global _fundamental_parser_warned
    if BeautifulSoup is not None:
        return True
    if not _fundamental_parser_warned:
        print("  [WARN] Fundamental parser unavailable: beautifulsoup4 is not installed.")
        print("  [WARN] Install with: pip install beautifulsoup4")
        print("  [WARN] Technical scan will continue; fundamentals will be marked Parser Missing.")
        _fundamental_parser_warned = True
    return False

def _company_clean_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()

def _company_text_or_empty(node) -> str:
    return _company_clean_text(node.get_text(" ", strip=True)) if node else ""

def _company_visible_lines(soup) -> list[str]:
    lines = []
    for line in soup.get_text("\n", strip=True).splitlines():
        cleaned = _company_clean_text(line)
        if cleaned:
            lines.append(cleaned)
    return lines

def _company_html_visible_lines(raw_html: str) -> list[str]:
    text = re.sub(r"<[^>]+>", "\n", raw_html or "")
    text = _html_mod.unescape(text)
    lines = []
    for line in text.splitlines():
        cleaned = _company_clean_text(line)
        if cleaned:
            lines.append(cleaned)
    return lines

def _company_lines_between(lines: list[str], start: str, stop_markers: list[str], limit: int | None = None) -> list[str]:
    wanted = _company_clean_text(start).lower()
    start_idx = -1
    for idx, line in enumerate(lines):
        if line.lower() == wanted:
            start_idx = idx + 1
            break
    if start_idx < 0:
        return []
    out = []
    stop_lower = [_company_clean_text(x).lower() for x in stop_markers]
    for line in lines[start_idx:]:
        lowered = line.lower()
        if any(lowered.startswith(marker) for marker in stop_lower if marker):
            break
        out.append(line)
        if limit and len(out) >= limit:
            break
    return out

def _company_parse_table(table) -> dict:
    if table is None:
        return {"headers": [], "rows": []}
    rows = []
    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if cells:
            rows.append([_company_clean_text(td.get_text(" ", strip=True)) for td in cells])
    if not rows:
        return {"headers": [], "rows": []}

    headers = rows[0]
    body = rows[1:]
    if len(rows) >= 2 and len(rows[1]) > len(rows[0]):
        headers = rows[1]
        body = rows[2:]
    return {"headers": headers, "rows": body}

def _company_find_heading(soup, name: str):
    wanted = _company_clean_text(name).lower()
    for tag in soup.find_all(["h1", "h2", "h3"]):
        text = _company_clean_text(tag.get_text(" ", strip=True)).lower()
        if text == wanted:
            return tag
    return None

def _company_section_root(heading):
    if not heading:
        return None
    return heading.find_parent("section") or heading.parent

def _company_extract_key_value_ratios(soup) -> dict[str, str]:
    ratios: dict[str, str] = {}
    top = soup.find(id="top-ratios")
    if top:
        for li in top.find_all("li"):
            full_text = _company_text_or_empty(li)
            if not full_text:
                continue
            key = ""
            name_node = li.find(class_=re.compile(r"\bname\b"))
            if name_node:
                key = _company_text_or_empty(name_node)
            if not key:
                parts = li.find_all(["span", "div"], recursive=False)
                if parts:
                    key = _company_text_or_empty(parts[0])
            if not key:
                match = re.match(r"([A-Za-z /%&.\-]+)\s+(.+)", full_text)
                key = _company_clean_text(match.group(1)) if match else ""
            value = full_text
            if key and value.startswith(key):
                value = _company_clean_text(value[len(key):])
            if key and value:
                ratios[key] = value
    if ratios:
        return ratios

    fallback_keys = [
        "Market Cap", "Current Price", "High / Low", "Stock P/E", "Book Value",
        "Dividend Yield", "ROCE", "ROE", "Face Value",
    ]
    lines = _company_visible_lines(soup)
    lower_to_key = {k.lower(): k for k in fallback_keys}
    for idx, line in enumerate(lines[:-1]):
        key = lower_to_key.get(line.lower())
        if key and key not in ratios:
            ratios[key] = lines[idx + 1]
    text = soup.get_text("\n", strip=True)
    for key in fallback_keys:
        if key in ratios:
            continue
        match = re.search(rf"{re.escape(key)}\s+([^\n]+)", text)
        if match:
            ratios[key] = _company_clean_text(match.group(1))
    return ratios

def _company_extract_list_section(soup, section_name: str) -> list[str]:
    heading = _company_find_heading(soup, section_name)
    root = _company_section_root(heading)
    if root:
        items = [
            _company_clean_text(li.get_text(" ", strip=True))
            for li in root.find_all("li")
            if _company_clean_text(li.get_text(" ", strip=True))
        ]
        if items:
            return items

    lines = _company_visible_lines(soup)
    stops = ["Cons", "Peer comparison", "Quarterly Results", "* The pros and cons", "The pros and cons are machine generated"]
    if _company_clean_text(section_name).lower() == "cons":
        stops = ["Peer comparison", "Quarterly Results", "* The pros and cons", "The pros and cons are machine generated"]
    items = []
    for line in _company_lines_between(lines, section_name, stops, limit=10):
        lowered = line.lower()
        if "machine generated" in lowered or line in ("Pros", "Cons"):
            break
        cleaned = line.lstrip("* ").strip()
        if cleaned:
            items.append(cleaned)
    return items

def _company_extract_table_after_heading(soup, heading_name: str) -> dict:
    heading = _company_find_heading(soup, heading_name)
    if not heading:
        return {"headers": [], "rows": []}
    return _company_parse_table(heading.find_next("table"))

def _company_extract_growth_metrics(soup, raw_html: str) -> dict[str, dict[str, str]]:
    labels = [
        "Compounded Sales Growth",
        "Compounded Profit Growth",
        "Stock Price CAGR",
        "Return on Equity",
    ]
    all_lines = _company_visible_lines(soup) or _company_html_visible_lines(raw_html)
    metrics: dict[str, dict[str, str]] = {}
    for idx, line in enumerate(all_lines):
        if line not in labels:
            continue
        entries = {}
        cursor = idx + 1
        while cursor + 1 < len(all_lines):
            key_line = _company_clean_text(all_lines[cursor])
            if key_line in labels or key_line in ("Balance Sheet", "Cash Flows", "Ratios", "Shareholding Pattern"):
                break
            val_line = _company_clean_text(all_lines[cursor + 1])
            if key_line.endswith(":"):
                entries[key_line[:-1]] = val_line
                cursor += 2
                continue
            if ":" in key_line:
                k, v = key_line.split(":", 1)
                entries[_company_clean_text(k)] = _company_clean_text(v) or val_line
                cursor += 1 if _company_clean_text(v) else 2
                continue
            break
        if entries:
            metrics[line] = entries
    return metrics

def _company_extract_shareholding(soup) -> dict[str, dict]:
    heading = _company_find_heading(soup, "Shareholding Pattern")
    if not heading:
        return {}
    tables = []
    cur = heading
    while True:
        cur = cur.find_next(["table", "h2"])
        if cur is None or getattr(cur, "name", None) == "h2":
            break
        if getattr(cur, "name", None) == "table":
            tables.append(cur)
        if len(tables) >= 2:
            break
    out = {}
    if tables:
        out["quarterly"] = _company_parse_table(tables[0])
    if len(tables) > 1:
        out["yearly"] = _company_parse_table(tables[1])
    return out

def _fetch_screener_company_html(url: str, cookie: str = "", proxy_url: str = "", timeout: int = 30) -> str:
    if not _fundamental_parser_available():
        raise RuntimeError("beautifulsoup4 missing")
    headers = {"User-Agent": _FUNDAMENTAL_USER_AGENT}
    if cookie:
        headers = {**headers, "Cookie": cookie}
    attempts = [(url, headers)]
    if proxy_url and cookie:
        attempts.append((
            f"{proxy_url}?url={url_encode(url, safe='')}",
            {"User-Agent": _FUNDAMENTAL_USER_AGENT, "X-Screener-Cookie": cookie},
        ))

    errors = []
    for target, hdrs in attempts:
        try:
            resp = requests.get(target, headers=hdrs, timeout=timeout)
            if resp.status_code == 200 and "screener" in resp.text.lower():
                return resp.text
            errors.append(f"{target} -> HTTP {resp.status_code}")
        except Exception as exc:
            errors.append(f"{target} -> {exc}")
    raise RuntimeError("Failed to fetch Screener company page. " + " | ".join(errors))

def _parse_screener_company_page(html: str, page_url: str) -> dict:
    if not _fundamental_parser_available():
        raise RuntimeError("beautifulsoup4 missing")
    soup = BeautifulSoup(html, "html.parser")
    return {
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "source": {
            "url": page_url,
            "website": "https://www.screener.in/",
            "type": "Screener company page",
        },
        "company": {
            "company_name": _company_text_or_empty(soup.find("h1")),
            "top_ratios": _company_extract_key_value_ratios(soup),
            "pros": _company_extract_list_section(soup, "Pros"),
            "cons": _company_extract_list_section(soup, "Cons"),
        },
        "financials": {
            "quarterly_results": _company_extract_table_after_heading(soup, "Quarterly Results"),
            "profit_loss": _company_extract_table_after_heading(soup, "Profit & Loss"),
            "balance_sheet": _company_extract_table_after_heading(soup, "Balance Sheet"),
            "cash_flows": _company_extract_table_after_heading(soup, "Cash Flows"),
            "ratios": _company_extract_table_after_heading(soup, "Ratios"),
            "growth_metrics": _company_extract_growth_metrics(soup, html),
        },
        "shareholding": _company_extract_shareholding(soup),
    }

def _fund_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float, np.integer, np.floating)):
        fv = float(v)
        return fv if math.isfinite(fv) else None
    s = _html_mod.unescape(str(v)).strip()
    if not s:
        return None
    s = s.replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
    s = s.replace("₹", "")
    s = s.replace("Cr.", "").replace("Cr", "").replace("%", "").strip()
    if s == "—":
        return None
    if s in {"-", "--", "—", "N/A", "NA"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", s)
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None

def _fund_round(v, digits=2):
    fv = _fund_float(v)
    return round(fv, digits) if fv is not None else None

def _fund_norm_label(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(label or "").lower())

def _fund_table_row(table: dict, row_label: str):
    target = _fund_norm_label(row_label)
    for row in (table or {}).get("rows") or []:
        if not row:
            continue
        label = _fund_norm_label(row[0])
        if not label:
            continue
        if label == target or label.startswith(target) or target.startswith(label):
            return row
    return None

def _fund_table_value(table: dict, row_label: str, prefer_ttm: bool = True):
    row = _fund_table_row(table, row_label)
    headers = (table or {}).get("headers") or []
    if not row:
        return None
    candidates = []
    if prefer_ttm:
        for idx, h in enumerate(headers):
            if str(h).strip().upper() == "TTM" and idx < len(row):
                candidates.append(row[idx])
    for value in reversed(row[1:]):
        candidates.append(value)
    for value in candidates:
        if str(value or "").strip() not in ("", "-", "--"):
            return value
    return None

def _fund_table_first_value(table: dict, row_labels, prefer_ttm: bool = True):
    for label in row_labels:
        value = _fund_table_value(table, label, prefer_ttm=prefer_ttm)
        if str(value or "").strip() not in ("", "-", "--", "—", "â€”"):
            return value
    return None

def _fund_growth_value(parsed: dict, section: str, period: str):
    growth = ((parsed or {}).get("financials") or {}).get("growth_metrics") or {}
    values = growth.get(section) or {}
    return _fund_round(values.get(period))

def _fund_shareholding_value(parsed: dict, row_label: str, back: int = 0):
    table = ((parsed or {}).get("shareholding") or {}).get("quarterly") or {}
    row = _fund_table_row(table, row_label)
    if not row:
        return None
    values = [v for v in row[1:] if str(v or "").strip() not in ("", "-", "--")]
    if not values or back >= len(values):
        return None
    return _fund_round(values[-1 - back])

def _fund_ratio(parsed: dict, key: str):
    ratios = ((parsed or {}).get("company") or {}).get("top_ratios") or {}
    return _fund_round(ratios.get(key))

def _fund_tag(score, strong=75, good=60, mixed=45):
    if score is None:
        return "UNKNOWN"
    if score >= strong:
        return "STRONG"
    if score >= good:
        return "GOOD"
    if score >= mixed:
        return "MIXED"
    return "WEAK"

_FUNDAMENTAL_SCORE_INPUT_FIELDS = (
    "Market Cap Cr", "PE", "PB", "EV/EBITDA",
    "Sales TTM Cr", "Profit TTM Cr", "OPM%", "NPM%", "ROE%", "ROCE%",
    "Sales CAGR 3Y%", "Sales Growth TTM%", "Profit CAGR 3Y%", "Profit Growth TTM%",
    "Debt/Equity", "Borrowings Cr", "Interest Coverage",
    "CFO TTM Cr", "FCF TTM Cr", "CFO/PAT", "CFO/OP%",
    "Promoter Holding%", "Promoter Holding Change%", "Promoter Pledge%",
)

def _fundamental_data_coverage(f: dict) -> int:
    return sum(1 for key in _FUNDAMENTAL_SCORE_INPUT_FIELDS if _fund_float((f or {}).get(key)) is not None)

def _unknown_fundamental_tags(freshness: str | None = None) -> dict:
    data = {
        "Fundamental Score": None, "Fundamental Quality Tag": "UNKNOWN",
        "Profitability Tag": "UNKNOWN", "Growth Tag": "UNKNOWN", "Valuation Tag": "UNKNOWN",
        "Balance Sheet Tag": "UNKNOWN", "Cashflow Tag": "UNKNOWN", "Ownership Tag": "UNKNOWN",
        "Fundamental Risk Tag": "UNKNOWN", "Investability Tag": "UNKNOWN", "Early Entry OK": "No",
    }
    if freshness:
        data["Fundamental Freshness"] = freshness
    return data

def _fund_component_scores(f):
    roe, roce, opm, npm = (_fund_float(f.get(k)) for k in ("ROE%", "ROCE%", "OPM%", "NPM%"))
    sales3, profit3, sales_ttm, profit_ttm = (
        _fund_float(f.get(k)) for k in ("Sales CAGR 3Y%", "Profit CAGR 3Y%", "Sales Growth TTM%", "Profit Growth TTM%")
    )
    de, ic, debtor = (_fund_float(f.get(k)) for k in ("Debt/Equity", "Interest Coverage", "Debtor Days"))
    cfo_pat, fcf, cfo_op = (_fund_float(f.get(k)) for k in ("CFO/PAT", "FCF TTM Cr", "CFO/OP%"))
    prom, prom_chg, pledge = (_fund_float(f.get(k)) for k in ("Promoter Holding%", "Promoter Holding Change%", "Promoter Pledge%"))
    pe, pb, evebitda = (_fund_float(f.get(k)) for k in ("PE", "PB", "EV/EBITDA"))

    profitability = 50
    if roe is not None:
        profitability += 20 if roe >= 20 else 10 if roe >= 15 else -10 if roe < 8 else 0
    if roce is not None:
        profitability += 25 if roce >= 20 else 12 if roce >= 15 else -12 if roce < 8 else 0
    if opm is not None:
        profitability += 10 if opm >= 15 else -8 if opm < 8 else 0
    if npm is not None:
        profitability += 10 if npm >= 8 else -8 if npm < 3 else 0

    growth = 50
    for val in (sales3, profit3, sales_ttm, profit_ttm):
        if val is None:
            continue
        growth += 12 if val >= 20 else 6 if val >= 10 else -10 if val < 0 else 0

    balance = 60
    if de is not None:
        balance += 15 if de <= 0.5 else 8 if de <= 1.0 else -18 if de > 2.0 else -8
    if ic is not None:
        balance += 15 if ic >= 5 else 5 if ic >= 2 else -18
    if debtor is not None:
        balance += 8 if debtor <= 90 else -12 if debtor > 180 else -5 if debtor > 120 else 0

    cashflow = 50
    if cfo_pat is not None:
        cashflow += 25 if cfo_pat >= 0.8 else 10 if cfo_pat >= 0.5 else -15 if cfo_pat < 0.25 else 0
    if fcf is not None:
        cashflow += 15 if fcf > 0 else -15
    if cfo_op is not None:
        cashflow += 10 if cfo_op >= 80 else -10 if cfo_op < 40 else 0

    ownership = 55
    if prom is not None:
        ownership += 15 if prom >= 50 else 8 if prom >= 35 else -10 if prom < 20 else 0
    if prom_chg is not None:
        ownership += 8 if prom_chg > 0 else -12 if prom_chg <= -5 else -5 if prom_chg < 0 else 0
    if pledge is not None:
        ownership += 10 if pledge == 0 else -20 if pledge > 10 else -8

    valuation = 55
    if pe is not None:
        valuation += 12 if 8 <= pe <= 35 else -15 if pe > 70 or pe <= 0 else -5 if pe > 50 else 0
    if pb is not None:
        valuation += 8 if pb <= 5 else -15 if pb > 12 else -6 if pb > 8 else 0
    if evebitda is not None:
        valuation += 8 if evebitda <= 20 else -12 if evebitda > 40 else -4 if evebitda > 30 else 0

    clamp = lambda x: max(0, min(100, round(float(x), 1)))
    return {
        "profitability": clamp(profitability),
        "growth": clamp(growth),
        "balance": clamp(balance),
        "cashflow": clamp(cashflow),
        "ownership": clamp(ownership),
        "valuation": clamp(valuation),
    }

def _derive_fundamental_tags(f: dict) -> dict:
    if not f or f.get("Fundamental Freshness") in ("Missing", "Parse Error", "Parser Missing"):
        return _unknown_fundamental_tags()
    coverage = _fundamental_data_coverage(f)
    anchor_fields = ("Sales TTM Cr", "Profit TTM Cr", "ROE%", "ROCE%", "Market Cap Cr")
    financial_anchor_fields = (
        "Sales TTM Cr", "Profit TTM Cr", "OPM%", "Sales CAGR 3Y%", "Profit CAGR 3Y%",
        "Debt/Equity", "CFO TTM Cr", "Promoter Holding%",
    )
    has_anchor = any(_fund_float(f.get(k)) is not None for k in anchor_fields)
    has_financial_anchor = any(_fund_float(f.get(k)) is not None for k in financial_anchor_fields)
    if coverage < 8 or not has_anchor or not has_financial_anchor:
        return _unknown_fundamental_tags("Insufficient")
    scores = _fund_component_scores(f)
    total = (
        scores["profitability"] * 0.25 + scores["growth"] * 0.25 +
        scores["balance"] * 0.20 + scores["cashflow"] * 0.15 +
        scores["ownership"] * 0.10 + scores["valuation"] * 0.05
    )
    total = round(total, 1)

    risk_flags = 0
    if scores["balance"] < 45: risk_flags += 2
    if scores["cashflow"] < 45: risk_flags += 2
    if scores["valuation"] < 40: risk_flags += 1
    if scores["ownership"] < 45: risk_flags += 1
    if (_fund_float(f.get("Debt/Equity")) or 0) > 2: risk_flags += 2
    pledge = _fund_float(f.get("Promoter Pledge%"))
    if pledge is not None and pledge > 10: risk_flags += 2
    debtor = _fund_float(f.get("Debtor Days"))
    if debtor is not None and debtor > 180: risk_flags += 1

    if risk_flags >= 4:
        fund_risk = "HIGH"
    elif risk_flags >= 2:
        fund_risk = "MED"
    else:
        fund_risk = "LOW"

    quality = "ELITE" if total >= 82 else "STRONG" if total >= 70 else "GOOD" if total >= 60 else "MIXED" if total >= 45 else "WEAK"
    if total >= 75 and fund_risk == "LOW":
        invest = "INVEST GRADE"
    elif total >= 65 and fund_risk != "HIGH":
        invest = "ACCUMULATE QUALITY"
    elif total >= 50:
        invest = "TRADE ONLY"
    else:
        invest = "AVOID"
    early = "Yes" if total >= 70 and fund_risk != "HIGH" and scores["growth"] >= 60 and scores["profitability"] >= 60 else "Watch" if total >= 60 and fund_risk != "HIGH" else "No"

    return {
        "Fundamental Score": total,
        "Fundamental Quality Tag": quality,
        "Profitability Tag": _fund_tag(scores["profitability"]),
        "Growth Tag": _fund_tag(scores["growth"]),
        "Valuation Tag": _fund_tag(scores["valuation"], strong=70, good=58, mixed=42),
        "Balance Sheet Tag": _fund_tag(scores["balance"]),
        "Cashflow Tag": _fund_tag(scores["cashflow"]),
        "Ownership Tag": _fund_tag(scores["ownership"]),
        "Fundamental Risk Tag": fund_risk,
        "Investability Tag": invest,
        "Early Entry OK": early,
    }

def _normalize_company_fundamentals(parsed: dict, symbol: str, name: str, url: str) -> dict:
    financials = (parsed or {}).get("financials") or {}
    pl = financials.get("profit_loss") or {}
    bs = financials.get("balance_sheet") or {}
    cf = financials.get("cash_flows") or {}
    ratios_table = financials.get("ratios") or {}

    sales_ttm = _fund_round(_fund_table_first_value(pl, ("Sales", "Revenue", "Interest Earned")))
    profit_ttm = _fund_round(_fund_table_first_value(pl, ("Net Profit", "Profit after tax", "PAT")))
    op_profit = _fund_round(_fund_table_first_value(pl, ("Operating Profit", "Financing Profit")))
    dep = _fund_round(_fund_table_value(pl, "Depreciation"))
    interest = _fund_round(_fund_table_value(pl, "Interest"))
    pbt = _fund_round(_fund_table_first_value(pl, ("Profit before tax", "Profit before Tax", "PBT")))
    eps_ttm = _fund_round(_fund_table_value(pl, "EPS in Rs"))
    opm = _fund_round(_fund_table_value(pl, "OPM %"))
    borrowings = _fund_round(_fund_table_value(bs, "Borrowings", prefer_ttm=False))
    equity = _fund_round(_fund_table_value(bs, "Equity Capital", prefer_ttm=False))
    reserves = _fund_round(_fund_table_value(bs, "Reserves", prefer_ttm=False))
    total_assets = _fund_round(_fund_table_value(bs, "Total Assets", prefer_ttm=False))
    cfo = _fund_round(_fund_table_value(cf, "Cash from Operating Activity"))
    fcf = _fund_round(_fund_table_value(cf, "Free Cash Flow"))
    cfo_op = _fund_round(_fund_table_value(cf, "CFO/OP"))
    debtor_days = _fund_round(_fund_table_value(ratios_table, "Debtor Days", prefer_ttm=False))
    inventory_days = _fund_round(_fund_table_value(ratios_table, "Inventory Days", prefer_ttm=False))
    ccc = _fund_round(_fund_table_value(ratios_table, "Cash Conversion Cycle", prefer_ttm=False))

    market_cap = _fund_ratio(parsed, "Market Cap")
    price = _fund_ratio(parsed, "Current Price")
    book_value = _fund_ratio(parsed, "Book Value")
    pe = _fund_ratio(parsed, "Stock P/E")
    pb = round(price / book_value, 2) if price and book_value else None
    ebitda = (op_profit or 0) + (dep or 0) if op_profit is not None or dep is not None else None
    ev_ebitda = round(((market_cap or 0) + (borrowings or 0)) / ebitda, 2) if ebitda and ebitda > 0 and market_cap else None
    debt_equity = round(borrowings / (equity + reserves), 2) if borrowings is not None and equity is not None and reserves is not None and (equity + reserves) else None
    interest_coverage = round(((pbt or 0) + (interest or 0)) / interest, 2) if interest and interest > 0 and pbt is not None else None
    npm = round(profit_ttm / sales_ttm * 100, 2) if profit_ttm is not None and sales_ttm else None
    asset_turnover = round(sales_ttm / total_assets, 2) if sales_ttm is not None and total_assets else None
    cfo_pat = round(cfo / profit_ttm, 2) if cfo is not None and profit_ttm else None
    prom = _fund_shareholding_value(parsed, "Promoters")
    prom_old = _fund_shareholding_value(parsed, "Promoters", back=4)
    prom_change = round(prom - prom_old, 2) if prom is not None and prom_old is not None else None
    pledge = _fund_shareholding_value(parsed, "Pledged")

    updated_at = fmt_dt(ist_now())
    summary = {
        "Fundamental Source": url,
        "Fundamental Updated At": updated_at,
        "Fundamental Freshness": "Fresh",
        "Market Cap Cr": market_cap,
        "PE": pe,
        "PB": pb,
        "EV/EBITDA": ev_ebitda,
        "Dividend Yield%": _fund_ratio(parsed, "Dividend Yield"),
        "Book Value": book_value,
        "EPS TTM": eps_ttm,
        "Sales TTM Cr": sales_ttm,
        "Profit TTM Cr": profit_ttm,
        "OPM%": opm,
        "NPM%": npm,
        "ROE%": _fund_ratio(parsed, "ROE"),
        "ROCE%": _fund_ratio(parsed, "ROCE"),
        "ROE 3Y%": _fund_growth_value(parsed, "Return on Equity", "3 Years"),
        "ROE Last Year%": _fund_growth_value(parsed, "Return on Equity", "Last Year"),
        "Sales CAGR 3Y%": _fund_growth_value(parsed, "Compounded Sales Growth", "3 Years"),
        "Sales CAGR 5Y%": _fund_growth_value(parsed, "Compounded Sales Growth", "5 Years"),
        "Sales Growth TTM%": _fund_growth_value(parsed, "Compounded Sales Growth", "TTM"),
        "Profit CAGR 3Y%": _fund_growth_value(parsed, "Compounded Profit Growth", "3 Years"),
        "Profit CAGR 5Y%": _fund_growth_value(parsed, "Compounded Profit Growth", "5 Years"),
        "Profit Growth TTM%": _fund_growth_value(parsed, "Compounded Profit Growth", "TTM"),
        "Debt/Equity": debt_equity,
        "Borrowings Cr": borrowings,
        "Interest Coverage": interest_coverage,
        "Asset Turnover": asset_turnover,
        "CFO TTM Cr": cfo,
        "FCF TTM Cr": fcf,
        "CFO/PAT": cfo_pat,
        "CFO/OP%": cfo_op,
        "Debtor Days": debtor_days,
        "Inventory Days": inventory_days,
        "Cash Conversion Cycle": ccc,
        "Promoter Holding%": prom,
        "Promoter Holding Change%": prom_change,
        "Promoter Pledge%": pledge,
        "Pros Count": len(((parsed or {}).get("company") or {}).get("pros") or []),
        "Cons Count": len(((parsed or {}).get("company") or {}).get("cons") or []),
    }
    summary.update(_derive_fundamental_tags(summary))
    return summary

def _build_company_url(symbol: str, meta: dict | None = None):
    meta = meta or {}
    slug = str(meta.get("screener_slug") or "").strip().strip("/")
    bse_code = str(meta.get("bseCode") or "").strip()
    sym = str(symbol or "").strip()
    if sym.startswith("BSE:"):
        sym = sym[4:]
    raw = slug or bse_code or sym
    if not raw:
        return ""
    raw = str(raw).strip().strip("/").split("/")[0]
    suffix = "/consolidated/" if FUNDAMENTAL_PREFER_CONSOLIDATED else "/"
    raw_path = raw if slug else raw.upper()
    return f"https://www.screener.in/company/{url_encode(raw_path, safe='')}{suffix}"

def _fundamental_cache_fresh(entry: dict):
    if FUNDAMENTAL_FORCE_REFRESH or not isinstance(entry, dict):
        return False
    fetched_at = str(entry.get("fetched_at") or "")
    try:
        dt = datetime.fromisoformat(fetched_at.replace("Z", "+00:00"))
        age = datetime.now(timezone.utc) - dt.astimezone(timezone.utc)
        return age.days < FUNDAMENTAL_CACHE_MAX_AGE_DAYS
    except Exception:
        return False

def _fetch_one_fundamental(symbol: str, meta: dict):
    global _fundamental_cache_dirty
    cache = _load_fundamental_cache()
    cached = (cache.get("symbols") or {}).get(symbol)
    if _fundamental_cache_fresh(cached):
        data = dict(cached.get("data") or {})
        data["Fundamental Freshness"] = "Fresh" if data.get("Fundamental Freshness") == "Fresh" else data.get("Fundamental Freshness")
        data.update(_derive_fundamental_tags(data))
        return symbol, data

    url = _build_company_url(symbol, meta)
    if not url:
        data = {"Fundamental Freshness": "Missing", "Investability Tag": "UNKNOWN", "Early Entry OK": "No"}
        data.update(_derive_fundamental_tags(data))
        return symbol, data

    if not _fundamental_parser_available():
        data = {"Fundamental Source": url, "Fundamental Freshness": "Parser Missing", "Investability Tag": "UNKNOWN", "Early Entry OK": "No"}
        data.update(_derive_fundamental_tags(data))
        return symbol, data

    try:
        cookie = SCREENER_COOKIE if FUNDAMENTAL_USE_PROXY else ""
        proxy = PROXY_URL if FUNDAMENTAL_USE_PROXY else ""
        candidate_urls = [url]
        if "/consolidated/" in url:
            candidate_urls.append(url.replace("/consolidated/", "/"))
        last_error = None
        data = None
        used_url = url
        for candidate_url in candidate_urls:
            try:
                html = _fetch_screener_company_html(candidate_url, cookie=cookie, proxy_url=proxy, timeout=FUNDAMENTAL_FETCH_TIMEOUT_SEC)
                parsed = _parse_screener_company_page(html, candidate_url)
                data = _normalize_company_fundamentals(parsed, symbol, meta.get("name") or "", candidate_url)
                used_url = candidate_url
                break
            except Exception as inner:
                last_error = inner
        if data is None:
            raise last_error or RuntimeError("no fundamental data parsed")
        cache.setdefault("symbols", {})[symbol] = {
            "fetched_at": datetime.now(timezone.utc).isoformat(),
            "url": used_url,
            "data": data,
        }
        _fundamental_cache_dirty = True
        return symbol, data
    except Exception as e:
        if isinstance(cached, dict) and isinstance(cached.get("data"), dict):
            data = dict(cached["data"])
            data["Fundamental Freshness"] = "Stale"
            data.update(_derive_fundamental_tags(data))
            return symbol, data
        data = {"Fundamental Source": url, "Fundamental Freshness": "Parse Error", "Investability Tag": "UNKNOWN", "Early Entry OK": "No"}
        data.update(_derive_fundamental_tags(data))
        print(f"      [WARN] Fundamentals {symbol}: {e}")
        return symbol, data

def _collect_fundamentals_parallel(symbol_meta: dict):
    _fundamental_data.clear()
    if not FUNDAMENTALS_ENABLED or not symbol_meta:
        return {}
    _load_fundamental_cache()
    if not _fundamental_parser_available():
        print("  [WARN] Phase 2B fundamentals skipped: built-in parser dependency unavailable. Technical scan will continue.")
        for sym, meta in sorted(symbol_meta.items()):
            data = {
                "Fundamental Source": _build_company_url(sym, meta),
                "Fundamental Freshness": "Parser Missing",
                "Investability Tag": "UNKNOWN",
                "Early Entry OK": "No",
            }
            data.update(_derive_fundamental_tags(data))
            _fundamental_data[sym] = data
        return _fundamental_data
    done = 0
    total = len(symbol_meta)
    workers = max(1, min(FUNDAMENTAL_FETCH_WORKERS, total))
    print(f"\n  [+] Phase 2B: Fetching fundamentals for {total} unique symbols (parallel {workers})...")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_fetch_one_fundamental, sym, meta): sym for sym, meta in sorted(symbol_meta.items())}
        for fut in as_completed(futures):
            try:
                sym, data = fut.result()
            except Exception as e:
                sym = futures.get(fut, "")
                data = {"Fundamental Freshness": "Parse Error", "Fundamental Source": "", "Investability Tag": "UNKNOWN", "Early Entry OK": "No"}
                data.update(_derive_fundamental_tags(data))
                print(f"      [WARN] Fundamentals {sym}: {e}")
            done += 1
            if data:
                _fundamental_data[sym] = data
            if done % 25 == 0 or done == total:
                print(f"    [UP] Fundamentals progress: {done}/{total} ({len(_fundamental_data)} OK)")
    _save_fundamental_cache()
    return _fundamental_data

def _signal_bullish_level(signal):
    s = str(signal or "").upper()
    if "BREAKOUT" in s or "STRONG BUY" in s:
        return 4
    if "BUY" in s:
        return 3
    if "PULLBACK" in s or "OVERSOLD" in s:
        return 2
    if "HOLD" in s:
        return 1
    if "SELL" in s or "WEAK" in s:
        return -2
    return 0

def _technical_decision_score(m: dict):
    level = _signal_bullish_level(m.get("Signal"))
    score = 50 + level * 10
    win_prob = _fund_float(m.get("Win Prob%"))
    exp10 = _fund_float(m.get("Exp 10D%"))
    rs_tag = str(m.get("RS Tag") or "").upper()
    liq = str(m.get("Liquidity Tag") or "").upper()
    quality = str(m.get("Signal Quality") or "").upper()
    regime = str(m.get("Signal Regime") or "").upper()
    if win_prob is not None:
        score += (win_prob - 55) * 0.4
    if exp10 is not None:
        score += exp10 * 2
    if "HIGH" in quality: score += 8
    if "REJECT" in quality: score -= 18
    if regime == "TRENDING": score += 5
    if regime == "HIGH-VOL": score -= 8
    if "STRONG VS BOTH" in rs_tag or "RS LEADER" in rs_tag: score += 7
    if "LAGGING" in rs_tag or "WEAK RS" in rs_tag: score -= 10
    if "ILLIQUID" in liq or "THIN" in liq: score -= 10
    return round(max(0, min(100, score)), 1)

def _final_signal_guardrails(m: dict, mtf_count: int | None, fund_score, fund_risk: str, invest: str):
    hard_blocks = []
    cautions = []
    quality = str(m.get("Signal Quality") or "").upper()
    regime = str(m.get("Signal Regime") or "").upper()
    rs_tag = str(m.get("RS Tag") or "").upper()
    liq = str(m.get("Liquidity Tag") or "").upper()
    fresh = str(m.get("Fundamental Freshness") or "").upper()
    win_prob = _fund_float(m.get("Win Prob%"))
    hist_precision = _fund_float(m.get("Hist Precision%"))
    exp5 = _fund_float(m.get("Exp 5D%"))
    exp10 = _fund_float(m.get("Exp 10D%"))
    wf_samples = _fund_float(m.get("WF Samples"))
    ai_dec = str(m.get("AI Decision") or "").upper()
    ai_conf = _fund_float(m.get("AI Conf%"))
    rsi = _fund_float(m.get("RSI 14"))
    dist52 = _fund_float(m.get("52W High Dist%"))
    natr = _fund_float(m.get("NATR 14"))

    if "REJECT" in quality:
        hard_blocks.append("signal quality rejected")
    if "ILLIQUID" in liq or "THIN" in liq:
        hard_blocks.append("liquidity weak")
    if fund_risk == "HIGH" or invest == "AVOID":
        hard_blocks.append("fundamental risk high")
    if "LAGGING" in rs_tag:
        hard_blocks.append("relative strength lagging")
    elif "WEAK RS" in rs_tag:
        cautions.append("relative strength weak")

    if win_prob is not None:
        if win_prob < 50:
            hard_blocks.append("win probability below 50")
        elif win_prob < 55:
            cautions.append("win probability below ideal")
    if hist_precision is not None and (wf_samples is None or wf_samples >= 20):
        if hist_precision < 45:
            hard_blocks.append("historical precision weak")
        elif hist_precision < 52:
            cautions.append("historical precision not strong")
    elif wf_samples is not None and wf_samples < 20:
        cautions.append("walk-forward sample small")

    if exp5 is not None and exp10 is not None and exp5 < 0 and exp10 < 0:
        hard_blocks.append("expected returns negative")
    elif (exp5 is not None and exp5 < 0) or (exp10 is not None and exp10 < 0):
        cautions.append("one expected-return window negative")

    if regime == "HIGH-VOL":
        if exp10 is not None and exp10 <= 0:
            hard_blocks.append("high-vol regime without positive 10D expectancy")
        else:
            cautions.append("high-vol regime")
    elif regime == "CHOPPY":
        cautions.append("choppy regime")

    if mtf_count is not None and mtf_count < 2:
        cautions.append("less than 2 MTF confirmations")
    if fresh not in ("", "FRESH", "STALE") and fund_score is None:
        cautions.append("fundamentals unavailable or insufficient")
    if rsi is not None and rsi > 76 and dist52 is not None and dist52 < 3:
        cautions.append("stretched near 52W high")
    if natr is not None and natr >= 9:
        cautions.append("volatility elevated")
    if ai_dec in ("REDUCE", "SELL", "STRONG SELL") and (ai_conf is None or ai_conf >= 70):
        hard_blocks.append("AI bearish confirmation")

    def _dedupe(items):
        out = []
        for item in items:
            if item and item not in out:
                out.append(item)
        return out

    return _dedupe(hard_blocks), _dedupe(cautions)

def _downgrade_final_signal(final: str, hard_blocks: list[str], cautions: list[str], fund_risk: str, invest: str):
    bullish = {"HIGH CONVICTION BUY", "BUY", "EARLY ENTRY", "ACCUMULATE"}
    if final not in bullish:
        return final, None
    if hard_blocks:
        if fund_risk == "HIGH" or invest == "AVOID":
            return "AVOID", "Strict guardrail blocked bullish signal"
        return "WATCH", "Strict guardrail blocked bullish signal"
    if cautions:
        if final == "HIGH CONVICTION BUY":
            return "BUY", "Caution guardrail reduced conviction"
        if final == "BUY":
            return "ACCUMULATE", "Caution guardrail reduced conviction"
        if final in ("EARLY ENTRY", "ACCUMULATE"):
            return "WATCH", "Caution guardrail requires more confirmation"
    return final, None

def _final_confidence_tag(final: str, combined, hard_blocks: list[str], cautions: list[str]):
    if hard_blocks:
        return "BLOCKED"
    if final in ("AVOID", "WATCH") and cautions:
        return "LOW"
    score = _fund_float(combined)
    if final == "HIGH CONVICTION BUY" and not cautions and score is not None and score >= 75:
        return "VERY HIGH"
    if final in ("HIGH CONVICTION BUY", "BUY", "EARLY ENTRY", "ACCUMULATE") and not cautions and (score is None or score >= 65):
        return "HIGH"
    if final in ("BUY", "EARLY ENTRY", "ACCUMULATE") or cautions:
        return "MED"
    if final in ("TRADE ONLY", "WAIT FOR TECHNICALS", "TECHNICAL ONLY"):
        return "MED"
    return "LOW"

def _apply_combined_final_signal(m: dict, mtf_count: int | None = None):
    tech_score = _technical_decision_score(m)
    fund_score = _fund_float(m.get("Fundamental Score"))
    fund_risk = str(m.get("Fundamental Risk Tag") or "").upper()
    invest = str(m.get("Investability Tag") or "").upper()
    early_ok = str(m.get("Early Entry OK") or "").upper()
    signal_level = _signal_bullish_level(m.get("Signal"))
    ai_dec = str(m.get("AI Decision") or "").upper()
    ai_adj = 0
    if ai_dec in ("STRONG BUY", "BUY"): ai_adj = 5
    elif ai_dec == "ACCUMULATE": ai_adj = 2
    elif ai_dec in ("REDUCE", "SELL", "STRONG SELL"): ai_adj = -7

    if fund_score is None:
        combined = max(0, min(100, tech_score + ai_adj))
    else:
        combined = max(0, min(100, tech_score * 0.55 + fund_score * 0.45 + ai_adj))
    combined = round(combined, 1)

    if fund_score is None:
        final = "TECHNICAL ONLY"
        reason = "Fundamentals unavailable; using technical signal only"
    elif signal_level >= 3 and fund_score >= 75 and fund_risk != "HIGH" and (mtf_count is None or mtf_count >= 2):
        final = "HIGH CONVICTION BUY"
        reason = "Bullish technicals plus strong fundamentals"
    elif signal_level >= 3 and fund_score >= 65 and fund_risk != "HIGH":
        final = "BUY"
        reason = "Technical setup supported by acceptable fundamentals"
    elif signal_level >= 1 and fund_score >= 70 and fund_risk != "HIGH" and early_ok in ("YES", "WATCH"):
        final = "EARLY ENTRY"
        reason = "Fundamentals strong while technicals are improving"
    elif signal_level >= 3 and (fund_score < 50 or invest == "AVOID"):
        final = "TRADE ONLY"
        reason = "Technical setup exists but fundamentals are weak"
    elif signal_level <= -1 and fund_score >= 70:
        final = "WAIT FOR TECHNICALS"
        reason = "Fundamentals strong but price structure is not ready"
    elif fund_risk == "HIGH" or invest == "AVOID":
        final = "AVOID"
        reason = "Fundamental risk is high"
    elif combined >= 60:
        final = "ACCUMULATE"
        reason = "Combined technical and fundamental score is constructive"
    else:
        final = "WATCH"
        reason = "Mixed setup; wait for better confirmation"

    hard_blocks, cautions = _final_signal_guardrails(m, mtf_count, fund_score, fund_risk, invest)
    downgraded_final, downgrade_reason = _downgrade_final_signal(final, hard_blocks, cautions, fund_risk, invest)
    if downgraded_final != final:
        drivers = hard_blocks if hard_blocks else cautions
        reason = f"{downgrade_reason}: {', '.join(drivers[:3])}"
        final = downgraded_final

    m["Tech + Fundamental Score"] = combined
    if hard_blocks:
        m["Decision Guardrail"] = "BLOCK - " + "; ".join(hard_blocks[:4])
    elif cautions:
        m["Decision Guardrail"] = "CAUTION - " + "; ".join(cautions[:4])
    else:
        m["Decision Guardrail"] = "PASS - STRICT"
    m["Final Confidence Tag"] = _final_confidence_tag(final, combined, hard_blocks, cautions)
    m["Final Signal"] = final
    m["Final Signal Reason"] = reason
    return m

def _valid_query_word(w: str) -> bool:
    w = w.strip(".")
    if not w or len(w) <= 1: return False
    if re.fullmatch(r'[^a-zA-Z0-9]+', w): return False
    return True

def _name_for_search(name: str) -> str:
    clean = re.sub(
        r'\b(ltd\.?|limited|inds\.?|industries|pvt\.?|private|corp\.?|'
        r'corporation|enterprises|holdings|technologies|tech|'
        r'pharmaceuticals?|pharma|chemicals?|textiles?|text\.?|'
        r'finance|financial|capital|investment|infra|infrastructure|'
        r'spinning|woven|fabrics?|steel|power|energy|agro|agri|'
        r'global|india|indian|retail|jewel(?:s|lers?)?|overseas|'
        r'fun\s*world|world|group|associates?|ventures?)\b',
        '', name, flags=re.IGNORECASE
    )
    clean = re.sub(r'[&.,()\[\]/\\0-9]', ' ', clean)
    words = [w for w in clean.split() if _valid_query_word(w)]
    return ' '.join(words).strip()

# -----------------------------------------------------------------------------
# WASABI S3
# -----------------------------------------------------------------------------
def _s3():
    return boto3.client(
        "s3", endpoint_url=S3_ENDPOINT,
        aws_access_key_id=S3_ACCESS_KEY, aws_secret_access_key=S3_SECRET_KEY,
        config=BotoConfig(signature_version="s3v4"), region_name=S3_REGION,
    )

# Sentinel: returned by s3_download_excel when file genuinely doesn't exist (first run)
_S3_FIRST_RUN = "__FIRST_RUN__"

def s3_download_excel(retries=3):
    """Download Excel from S3 with retry logic.
    Returns:
      - workbook on success
      - _S3_FIRST_RUN string if file doesn't exist (NoSuchKey)
      - None on transient/network failure (after retries)
    """
    for attempt in range(1, retries + 1):
        try:
            obj  = _s3().get_object(Bucket=S3_BUCKET, Key=S3_EXCEL_KEY)
            data = obj["Body"].read()
            wb   = openpyxl.load_workbook(io.BytesIO(data))
            print(f"  [DL] Downloaded existing Excel from S3 ({len(data)//1024} KB)")
            return wb
        except ClientError as e:
            code = e.response.get("Error",{}).get("Code","")
            if code in ("NoSuchKey","404"):
                print("  📂 No existing Excel on S3 -- first run, creating fresh workbook")
                return _S3_FIRST_RUN  # genuinely doesn't exist
            print(f"  [WARN]  S3 ClientError [{code}] (attempt {attempt}/{retries}): {e}")
        except Exception as e:
            print(f"  [WARN]  S3 download failed (attempt {attempt}/{retries}): {e}")
        if attempt < retries:
            wait = 5 * attempt
            print(f"  [WAIT] Retrying S3 download in {wait}s...")
            time.sleep(wait)
    print("  [FAIL] S3 download failed after all retries.")
    return None  # transient failure -- do NOT create fresh workbook

def _s3_get_existing_size():
    """Get the size in bytes of the current S3 Excel file (0 if not found)."""
    try:
        resp = _s3().head_object(Bucket=S3_BUCKET, Key=S3_EXCEL_KEY)
        return resp.get("ContentLength", 0)
    except Exception:
        return 0

def s3_backup_before_upload():
    """Archive current S3 Excel to a timestamped backup key before overwriting."""
    ts = ist_now().strftime("%Y-%m-%d_%H%M%S")
    backup_key = f"reports/backups/gas_stock_tracker_{ts}.xlsx"
    try:
        _s3().copy_object(
            Bucket=S3_BUCKET,
            CopySource={"Bucket": S3_BUCKET, "Key": S3_EXCEL_KEY},
            Key=backup_key,
        )
        print(f"  🗂️  Backed up to s3://{S3_BUCKET}/{backup_key}")
        return True
    except ClientError as e:
        code = e.response.get("Error",{}).get("Code","")
        if code in ("NoSuchKey","404"):
            return True  # nothing to backup on first run
        print(f"  [WARN]  Backup failed: {e}")
        return False
    except Exception as e:
        print(f"  [WARN]  Backup failed: {e}")
        return False

def s3_upload_excel(wb, save_local: bool = True, backup: bool = True):
    try:
        data = _serialize_workbook_bytes(wb)
    except Exception as e:
        print(f"  [FAIL] Excel serialise failed: {e}"); return False

    if save_local:
        _save_local_artifact(data, LOCAL_EXCEL_FILE, "Excel")

    # Safety check: refuse to upload tiny file if existing is much larger
    new_size = len(data)
    existing_size = _s3_get_existing_size()
    if new_size < 5120 and existing_size > 51200:  # < 5KB new vs > 50KB existing
        print(f"  🛡️  SAFETY: Refusing upload -- new file ({new_size}B) is suspiciously small "
              f"vs existing ({existing_size}B). Possible data wipe prevented!")
        return False

    # Backup before overwriting
    if backup:
        s3_backup_before_upload()

    for attempt in range(1, S3_UPLOAD_RETRIES + 1):
        try:
            _s3().put_object(Bucket=S3_BUCKET, Key=S3_EXCEL_KEY, Body=data,
                             ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            print(f"  📤 Uploaded Excel ({len(data)//1024} KB) -> s3://{S3_BUCKET}/{S3_EXCEL_KEY}")
            return True
        except Exception as e:
            print(f"  [WARN]  S3 upload attempt {attempt}/{S3_UPLOAD_RETRIES} failed: {e}")
            if attempt < S3_UPLOAD_RETRIES: time.sleep(5 * attempt)
    print("  [FAIL] S3 upload gave up after all retries."); return False

def s3_upload_dashboard_excel(wb):
    """Build, save, and upload a dashboard-only workbook."""
    try:
        print("  [STAT] Building dashboard-only workbook...")
        dash_wb = build_dashboard_only_workbook(wb)
        data = _serialize_workbook_bytes(dash_wb)
        print(f"  [STAT] Dashboard-only workbook ready ({len(data)//1024} KB)")
    except Exception as e:
        print(f"  [FAIL] Dashboard Excel serialise failed: {e}")
        return False

    _save_local_artifact(data, LOCAL_DASHBOARD_EXCEL_FILE, "Dashboard Excel")

    try:
        _s3_backup_key_before_upload(S3_DASHBOARD_EXCEL_KEY)
    except Exception:
        pass

    for attempt in range(1, S3_UPLOAD_RETRIES + 1):
        try:
            _s3().put_object(
                Bucket=S3_BUCKET,
                Key=S3_DASHBOARD_EXCEL_KEY,
                Body=data,
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            print(f"  📤 Uploaded Dashboard Excel ({len(data)//1024} KB) -> s3://{S3_BUCKET}/{S3_DASHBOARD_EXCEL_KEY}")
            return True
        except Exception as e:
            print(f"  [WARN]  Dashboard Excel upload attempt {attempt}/{S3_UPLOAD_RETRIES} failed: {e}")
            if attempt < S3_UPLOAD_RETRIES:
                time.sleep(5 * attempt)
    print("  [FAIL] Dashboard Excel upload gave up after all retries.")
    return False

def s3_generate_presigned_url(expiry_seconds=86400, key=None):
    """Generate a presigned download URL for an S3 object (default 24h expiry)."""
    try:
        url = _s3().generate_presigned_url(
            "get_object",
            Params={"Bucket": S3_BUCKET, "Key": key or S3_EXCEL_KEY},
            ExpiresIn=expiry_seconds,
        )
        return url
    except Exception as e:
        print(f"  [WARN]  Presigned URL generation failed: {e}")
        return None

def _script_base_dir():
    """Best-effort base directory for local artifacts."""
    if "__file__" in globals():
        try:
            return os.path.dirname(os.path.abspath(__file__))
        except Exception:
            pass
    return os.getcwd()

def _local_output_path(filename: str) -> str:
    return os.path.join(_script_base_dir(), filename)

def _save_local_artifact(data: bytes, filename: str, label: str) -> str | None:
    path = _local_output_path(filename)
    try:
        with open(path, "wb") as f:
            f.write(data)
        print(f"  💾 Saved {label} -> {path}")
        return path
    except Exception as e:
        print(f"  [WARN]  Local save failed for {label}: {e}")
        return None

def _serialize_workbook_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def _dashboard_db_local_path():
    return os.path.join(_script_base_dir(), DASHBOARD_DB_FILE)

def _db_clean_value(v):
    """Normalize worksheet values for DB storage / JSON payload."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (np.integer, int)):
        return int(v)
    if isinstance(v, (np.floating, float)):
        fv = float(v)
        return None if math.isnan(fv) else fv
    return str(v)

def _as_float(v):
    try:
        if v is None:
            return None
        fv = float(v)
        return None if math.isnan(fv) else fv
    except Exception:
        return None

def _as_int(v):
    try:
        if v is None:
            return None
        return int(float(v))
    except Exception:
        return None

def append_dashboard_snapshot_to_db(wb, iteration: int, snapshot_at: str | None = None):
    """Append current Dashboard sheet rows into SQLite (history-preserving)."""
    dash_ws = ensure_dashboard_sheet(wb)
    db_path = _dashboard_db_local_path()
    db_dir = os.path.dirname(db_path)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)

    snap_ts = snapshot_at or fmt_dt()
    table = DASHBOARD_DB_TABLE
    inserted = 0

    conn = sqlite3.connect(db_path, timeout=30)
    try:
        cur = conn.cursor()
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {table} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_at TEXT NOT NULL,
                iteration INTEGER NOT NULL,
                row_no INTEGER NOT NULL,
                symbol TEXT,
                name TEXT,
                in_screener TEXT,
                quick_action TEXT,
                consensus_score REAL,
                mtf_alignment TEXT,
                historical_mtf TEXT,
                sector TEXT,
                industry TEXT,
                sector_benchmark TEXT,
                rs_tag TEXT,
                rs_vs_nifty_1m REAL,
                rs_vs_nifty_3m REAL,
                rs_vs_sector_1m REAL,
                rs_vs_sector_3m REAL,
                avg_traded_value_20d_cr REAL,
                liquidity_tag TEXT,
                momentum_rank INTEGER,
                momentum_tag TEXT,
                risk_tag TEXT,
                bb_signal TEXT,
                cam_setup TEXT,
                cam_h3 REAL,
                cam_h4 REAL,
                cam_l3 REAL,
                cam_l4 REAL,
                ideal_enter_price REAL,
                possible_sell_value REAL,
                stop_loss_value REAL,
                volume_buzz TEXT,
                since_capture_trend TEXT,
                signal TEXT,
                setup_signal TEXT,
                core_signal TEXT,
                signal_quality TEXT,
                signal_regime TEXT,
                win_prob REAL,
                hist_precision REAL,
                exp_5d REAL,
                exp_10d REAL,
                wf_samples INTEGER,
                ai_decision TEXT,
                ai_conf REAL,
                payload_json TEXT NOT NULL
            )
        """)
        existing_cols = {str(r[1]) for r in cur.execute(f"PRAGMA table_info({table})").fetchall()}
        for col_name, col_type in (
            ("in_screener", "TEXT"),
            ("historical_mtf", "TEXT"),
            ("sector", "TEXT"),
            ("industry", "TEXT"),
            ("sector_benchmark", "TEXT"),
            ("rs_tag", "TEXT"),
            ("rs_vs_nifty_1m", "REAL"),
            ("rs_vs_nifty_3m", "REAL"),
            ("rs_vs_sector_1m", "REAL"),
            ("rs_vs_sector_3m", "REAL"),
            ("avg_traded_value_20d_cr", "REAL"),
            ("liquidity_tag", "TEXT"),
            ("momentum_tag", "TEXT"),
            ("cam_setup", "TEXT"),
            ("cam_h3", "REAL"),
            ("cam_h4", "REAL"),
            ("cam_l3", "REAL"),
            ("cam_l4", "REAL"),
            ("ideal_enter_price", "REAL"),
            ("possible_sell_value", "REAL"),
            ("stop_loss_value", "REAL"),
            ("setup_signal", "TEXT"),
            ("core_signal", "TEXT"),
            ("signal_quality", "TEXT"),
            ("signal_regime", "TEXT"),
            ("win_prob", "REAL"),
            ("hist_precision", "REAL"),
            ("exp_5d", "REAL"),
            ("exp_10d", "REAL"),
            ("wf_samples", "INTEGER"),
        ):
            if col_name not in existing_cols:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_type}")
        cur.execute(f"CREATE INDEX IF NOT EXISTS idx_{table}_symbol_ts ON {table}(symbol, snapshot_at)")
        cur.execute(f"CREATE INDEX IF NOT EXISTS idx_{table}_iteration ON {table}(iteration)")

        for ri in range(2, dash_ws.max_row + 1):
            row = {}
            for h in DASHBOARD_HEADERS:
                row[h] = _db_clean_value(dash_ws.cell(row=ri, column=DC[h] + 1).value)

            symbol = str(row.get("Symbol") or "").strip()
            name = str(row.get("Name") or "").strip()
            if not symbol and not name:
                continue

            payload = json.dumps(row, ensure_ascii=False)
            cur.execute(
                f"""
                INSERT INTO {table} (
                    snapshot_at, iteration, row_no, symbol, name,
                    in_screener,
                    quick_action, consensus_score, mtf_alignment, historical_mtf,
                    sector, industry, sector_benchmark, rs_tag, rs_vs_nifty_1m, rs_vs_nifty_3m, rs_vs_sector_1m, rs_vs_sector_3m, avg_traded_value_20d_cr, liquidity_tag,
                    momentum_rank, momentum_tag,
                    risk_tag, bb_signal, cam_setup, cam_h3, cam_h4, cam_l3, cam_l4,
                    ideal_enter_price, possible_sell_value, stop_loss_value,
                    volume_buzz, since_capture_trend,
                    signal, setup_signal, core_signal,
                    signal_quality, signal_regime, win_prob, hist_precision, exp_5d, exp_10d, wf_samples,
                    ai_decision, ai_conf, payload_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snap_ts,
                    int(iteration),
                    int(ri),
                    symbol or None,
                    name or None,
                    row.get("In Screener?"),
                    row.get("Quick Action"),
                    _as_float(row.get("Consensus Score")),
                    row.get("MTF Alignment"),
                    row.get("Historical MTF"),
                    row.get("Sector"),
                    row.get("Industry"),
                    row.get("Sector Benchmark"),
                    row.get("RS Tag"),
                    _as_float(row.get("RS vs NIFTY 1M%")),
                    _as_float(row.get("RS vs NIFTY 3M%")),
                    _as_float(row.get("RS vs Sector 1M%")),
                    _as_float(row.get("RS vs Sector 3M%")),
                    _as_float(row.get("Avg Traded Value 20D Cr")),
                    row.get("Liquidity Tag"),
                    _as_int(row.get("Momentum Rank")),
                    row.get("Momentum Tag"),
                    row.get("Risk Tag"),
                    row.get("BB Signal"),
                    row.get("Cam Setup"),
                    _as_float(row.get("Cam H3")),
                    _as_float(row.get("Cam H4")),
                    _as_float(row.get("Cam L3")),
                    _as_float(row.get("Cam L4")),
                    _as_float(row.get("Ideal Enter Price")),
                    _as_float(row.get("Possible Sell Value")),
                    _as_float(row.get("Stop Loss Value")),
                    row.get("Volume Buzz"),
                    row.get("Since Capture Trend"),
                    row.get("Signal"),
                    row.get("Setup Signal"),
                    row.get("Core Signal"),
                    row.get("Signal Quality"),
                    row.get("Signal Regime"),
                    _as_float(row.get("Win Prob%")),
                    _as_float(row.get("Hist Precision%")),
                    _as_float(row.get("Exp 5D%")),
                    _as_float(row.get("Exp 10D%")),
                    _as_int(row.get("WF Samples")),
                    row.get("AI Decision"),
                    _as_float(row.get("AI Conf%")),
                    payload,
                ),
            )
            inserted += 1

        conn.commit()
        return inserted, db_path
    finally:
        conn.close()

def _s3_backup_key_before_upload(key: str):
    """Backup an existing S3 object key to reports/backups/<name>_<ts>.<ext>."""
    ts = ist_now().strftime("%Y-%m-%d_%H%M%S")
    base = os.path.basename(key)
    name, ext = os.path.splitext(base)
    ext = ext or ".bak"
    backup_key = f"reports/backups/{name}_{ts}{ext}"
    try:
        _s3().copy_object(
            Bucket=S3_BUCKET,
            CopySource={"Bucket": S3_BUCKET, "Key": key},
            Key=backup_key,
        )
        print(f"  🗂️  DB backup -> s3://{S3_BUCKET}/{backup_key}")
        return True
    except ClientError as e:
        code = e.response.get("Error",{}).get("Code","")
        if code in ("NoSuchKey","404"):
            return True
        print(f"  [WARN]  DB backup failed [{code}]: {e}")
        return False
    except Exception as e:
        print(f"  [WARN]  DB backup failed: {e}")
        return False

def s3_upload_dashboard_db(db_path: str):
    """Upload local dashboard SQLite DB to S3 with retry + backup."""
    if not db_path or not os.path.exists(db_path):
        print(f"  [WARN]  Dashboard DB not found: {db_path}")
        return False

    try:
        with open(db_path, "rb") as f:
            data = f.read()
    except Exception as e:
        print(f"  [FAIL] Dashboard DB read failed: {e}")
        return False

    if not data:
        print("  [WARN]  Dashboard DB is empty -- skipping S3 upload.")
        return False

    _s3_backup_key_before_upload(S3_DASHBOARD_DB_KEY)

    for attempt in range(1, S3_UPLOAD_RETRIES + 1):
        try:
            _s3().put_object(
                Bucket=S3_BUCKET,
                Key=S3_DASHBOARD_DB_KEY,
                Body=data,
                ContentType="application/x-sqlite3",
            )
            print(f"  📤 Uploaded Dashboard DB ({len(data)//1024} KB) -> s3://{S3_BUCKET}/{S3_DASHBOARD_DB_KEY}")
            return True
        except Exception as e:
            print(f"  [WARN]  DB upload attempt {attempt}/{S3_UPLOAD_RETRIES} failed: {e}")
            if attempt < S3_UPLOAD_RETRIES:
                time.sleep(5 * attempt)
    print("  [FAIL] Dashboard DB upload gave up after all retries.")
    return False

# -----------------------------------------------------------------------------
# WORKBOOK / SHEET MANAGEMENT
# -----------------------------------------------------------------------------
def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

_THIN_BORDER = _thin_border()
_ALIGN_CENTER = Alignment(horizontal="center")
_ALIGN_CENTER_WRAP = Alignment(horizontal="center", wrap_text=True)
_ALIGN_LEFT_WRAP = Alignment(horizontal="left", wrap_text=True)
_BLANK_FILL = PatternFill()
_DEFAULT_FONT = Font()
_ROW_FILL_CACHE = {}

def _row_fill(color: str | None):
    if not color:
        return None
    fill = _ROW_FILL_CACHE.get(color)
    if fill is None:
        fill = PatternFill("solid", fgColor=color)
        _ROW_FILL_CACHE[color] = fill
    return fill

def ensure_scanner_sheet(wb, scanner):
    name = scanner["name"][:31]
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    for ci in range(1, min(max(ws.max_column or 0, len(HEADERS)) + 5, 250)):
        cell = ws.cell(row=1, column=ci)
        if ci > len(HEADERS):
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.alignment = Alignment()
            cell.border = Border()
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        # AI columns get a distinct purple header
        if h in ("AI Decision","AI Reason","AI Conf%"):
            c.fill = PatternFill("solid", fgColor="4A148C")
        else:
            c.fill = PatternFill("solid", fgColor="37474F")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[1].height = 32
    widths = {
        "Symbol":10,"Name":28,"First Captured":16,"Last Seen":16,"In Screener?":10,
        "Capture Price":13,"Current Price":13,"Since Capture%":12,
        "1D%":8,"1W%":8,"1M%":8,"3M%":8,"6M%":8,"1Y%":8,"2Y%":8,"3Y%":8,
        "Avg Weekly%":11,"Avg Monthly%":11,"Avg 3M%":10,"Avg 6M%":10,"Avg 1Y%":10,
        "RSI 14":9,"MA 20":11,"MA 50":11,"MA 200":11,
        "Signal":22,"Setup Signal":22,"Core Signal":22,
        "Signal Quality":18,"Signal Regime":14,"Win Prob%":10,"Hist Precision%":13,"Exp 5D%":10,"Exp 10D%":10,"WF Samples":10,
        "Sector":18,"Industry":22,"Sector Benchmark":16,"RS Tag":16,
        "RS vs NIFTY 1M%":12,"RS vs NIFTY 3M%":12,"RS vs Sector 1M%":12,"RS vs Sector 3M%":12,
        "Avg Traded Value 20D Cr":16,"Liquidity Tag":12,
        "AI Decision":16,"AI Reason":40,"AI Conf%":10,
        "Last Updated":16,"ADX 14":9,"Vol Ratio 20":12,
        "MACD Line":11,"MACD Hist":11,"52W High Dist%":14,"20D Breakout%":14,
        "ATR 14":10,"NATR 14":10,"+DI 14":9,"-DI 14":9,
    }
    widths.update({
        "Fundamental Score":12, "Fundamental Quality Tag":16, "Fundamental Risk Tag":16,
        "Investability Tag":18, "Early Entry OK":14, "Tech + Fundamental Score":18,
        "Decision Guardrail":34, "Final Confidence Tag":18,
        "Final Signal":20, "Final Signal Reason":34,
        "Market Cap Cr":13, "PE":9, "PB":9, "EV/EBITDA":11,
        "Sales Growth TTM%":15, "Profit Growth TTM%":16,
        "Debt/Equity":12, "CFO/PAT":10, "Promoter Holding%":16, "Promoter Pledge%":15,
        "Fundamental Source":32, "Fundamental Updated At":18, "Fundamental Freshness":16,
    })
    for h, w in widths.items():
        if h in C:
            ws.column_dimensions[get_column_letter(C[h]+1)].width = w
    ws.freeze_panes = "A2"
    return ws

def ensure_price_history_sheet(wb):
    ws = wb["Price History"] if "Price History" in wb.sheetnames else wb.create_sheet("Price History", 0)
    for ci, h in enumerate(PRICE_HISTORY_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A237E")
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border()
    for ci in range(len(PRICE_HISTORY_HEADERS) + 1, min(max(ws.max_column or 0, len(PRICE_HISTORY_HEADERS)) + 5, 250)):
        cell = ws.cell(row=1, column=ci)
        cell.value = None
        cell.fill = PatternFill()
        cell.font = Font()
        cell.alignment = Alignment()
        cell.border = Border()
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 25
    for header, width in {
        "Sector":18,"Industry":22,"Sector Benchmark":16,"RS Tag":16,
        "RS vs NIFTY 1M%":12,"RS vs NIFTY 3M%":12,"RS vs Sector 1M%":12,"RS vs Sector 3M%":12,
        "Avg Traded Value 20D Cr":16,"Liquidity Tag":12,
        "Fundamental Score":12,"Fundamental Quality Tag":16,"Growth Tag":12,"Valuation Tag":13,
        "Balance Sheet Tag":16,"Cashflow Tag":14,"Ownership Tag":14,"Fundamental Risk Tag":16,
        "Investability Tag":18,"Early Entry OK":14,"Tech + Fundamental Score":18,
        "Decision Guardrail":34,"Final Confidence Tag":18,"Final Signal":20,
        "Fundamental Updated At":18,"Fundamental Freshness":16,
    }.items():
        if header in PRICE_HISTORY_HEADERS:
            ws.column_dimensions[get_column_letter(PRICE_HISTORY_HEADERS.index(header) + 1)].width = width
    return ws

def ensure_dashboard_sheet(wb):
    """Create or return the Dashboard summary sheet. Always refreshes headers."""
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
    else:
        ws = wb.create_sheet("Dashboard", 0)  # place at front

    # Header color groups
    _action_cols = {"Quick Action","Consensus Score","MTF Alignment","Historical MTF","RS Tag","Momentum Rank","Momentum Tag",
                    "Risk Tag","BB Signal","Cam Setup","Volume Buzz","Since Capture Trend",
                    "Ideal Enter Price","Possible Sell Value","Stop Loss Value",
                    "Signal Quality","Signal Regime","Win Prob%","Hist Precision%","Exp 5D%","Exp 10D%","WF Samples",
                    "Fundamental Score","Fundamental Quality Tag","Fundamental Risk Tag","Investability Tag",
                    "Early Entry OK","Tech + Fundamental Score","Decision Guardrail","Final Confidence Tag",
                    "Final Signal","Final Signal Reason"}
    _ai_cols = {"AI Decision","AI Conf%"}
    _meta_cols = {"Best Scanner","Screener Link","Days Tracked","Sector","Industry","Sector Benchmark","Liquidity Tag",
                  "Fundamental Source","Fundamental Updated At","Fundamental Freshness"}
    _fund_cols = set(FUNDAMENTAL_RAW_FIELDS + FUNDAMENTAL_DERIVED_FIELDS)

    # Always rewrite headers (fixes old-schema sheets)
    # First clear any extra columns beyond current header count
    for ci in range(1, min(max(ws.max_column or 0, len(DASHBOARD_HEADERS)) + 5, 200)):
        cell = ws.cell(row=1, column=ci)
        cell.value = None
        cell.fill = PatternFill()
        cell.font = Font()
        cell.alignment = Alignment()
        cell.border = Border()

    # Write current headers
    for ci, h in enumerate(DASHBOARD_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        if h in _action_cols:
            c.fill = PatternFill("solid", fgColor="E65100")  # orange accent
        elif h in ("Total Appearances", "Unique Scanners"):
            c.fill = PatternFill("solid", fgColor="1565C0")
        elif h in _ai_cols:
            c.fill = PatternFill("solid", fgColor="4A148C")
        elif h in _meta_cols:
            c.fill = PatternFill("solid", fgColor="00695C")  # teal
        elif h in _fund_cols:
            c.fill = PatternFill("solid", fgColor="33691E")  # fundamentals green
        else:
            c.fill = PatternFill("solid", fgColor="0D47A1")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[1].height = 36

    # Column widths
    widths = {
        "Symbol":12,"Name":28,"In Screener?":11,"Quick Action":16,"Consensus Score":15,
        "MTF Alignment":14,"Historical MTF":14,
        "Sector":18,"Industry":22,"Sector Benchmark":16,"RS Tag":16,
        "RS vs NIFTY 1M%":12,"RS vs NIFTY 3M%":12,"RS vs Sector 1M%":12,"RS vs Sector 3M%":12,
        "Avg Traded Value 20D Cr":16,"Liquidity Tag":12,
        "Momentum Rank":14,"Momentum Tag":16,"Risk Tag":10,"BB Signal":14,"Cam Setup":20,"Volume Buzz":12,
        "Since Capture Trend":16,"First Captured":16,"Days Tracked":12,
        "Last Seen":16,"Total Appearances":16,"Unique Scanners":15,
        "Scanner List":40,"Best Scanner":20,"Capture Price":13,"Current Price":13,
        "Cam H3":11,"Cam H4":11,"Cam L3":11,"Cam L4":11,
        "Ideal Enter Price":15,"Possible Sell Value":16,"Stop Loss Value":14,
        "Since Capture%":13,"1D%":8,"1W%":8,"1M%":8,"3M%":8,"6M%":8,"1Y%":8,
        "RSI 14":9,"ADX 14":9,"+DI 14":9,"-DI 14":9,"ATR 14":10,"NATR 14":10,
        "Signal":22,"Setup Signal":22,"Core Signal":22,
        "Signal Quality":18,"Signal Regime":14,"Win Prob%":10,"Hist Precision%":13,"Exp 5D%":10,"Exp 10D%":10,"WF Samples":10,
        "AI Decision":16,"AI Conf%":10,
        "Screener Link":38,"Last Updated":17,
    }
    widths.update({
        "Fundamental Source":32, "Fundamental Updated At":18, "Fundamental Freshness":16,
        "Market Cap Cr":13, "PE":9, "PB":9, "EV/EBITDA":11, "Dividend Yield%":14,
        "Sales TTM Cr":13, "Profit TTM Cr":13, "OPM%":9, "NPM%":9,
        "ROE%":9, "ROCE%":9, "Sales Growth TTM%":15, "Profit Growth TTM%":16,
        "Debt/Equity":12, "Interest Coverage":15, "CFO/PAT":10, "CFO/OP%":10,
        "Promoter Holding%":16, "Promoter Pledge%":15,
        "Profitability Tag":15, "Growth Tag":12, "Valuation Tag":13, "Balance Sheet Tag":16,
        "Cashflow Tag":14, "Ownership Tag":14, "Fundamental Risk Tag":16,
        "Fundamental Quality Tag":18, "Fundamental Score":14, "Investability Tag":20,
        "Early Entry OK":14, "Tech + Fundamental Score":18, "Decision Guardrail":34,
        "Final Confidence Tag":18, "Final Signal":22, "Final Signal Reason":40,
    })
    for h, w in widths.items():
        if h in DC:
            ws.column_dimensions[get_column_letter(DC[h]+1)].width = w

    # Auto-filter on header row (so filter dropdowns appear on row 1)
    last_col = get_column_letter(len(DASHBOARD_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    ws.freeze_panes = "A2"
    return ws

def ensure_dashboard_history_sheet(wb):
    """Create or return the Dashboard History append-only log sheet. Always refreshes headers."""
    if "Dashboard History" in wb.sheetnames:
        ws = wb["Dashboard History"]
    else:
        ws = wb.create_sheet("Dashboard History", 1)  # right after Dashboard

    # Always rewrite headers (fixes old-schema sheets)
    for ci in range(1, min(max(ws.max_column or 0, len(DASHBOARD_HISTORY_HEADERS)) + 5, 200)):
        cell = ws.cell(row=1, column=ci)
        cell.value = None; cell.fill = PatternFill(); cell.font = Font()
        cell.alignment = Alignment(); cell.border = Border()

    for ci, h in enumerate(DASHBOARD_HISTORY_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="004D40")  # teal
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border()

    # Column widths
    hist_widths = {
        "Snapshot At":18, "Symbol":12, "Name":28, "In Screener?":11, "Quick Action":16,
        "Consensus Score":15, "MTF Alignment":14, "Historical MTF":14,
        "Sector":18,"Industry":22,"Sector Benchmark":16,"RS Tag":16,
        "RS vs NIFTY 1M%":12,"RS vs NIFTY 3M%":12,"RS vs Sector 1M%":12,"RS vs Sector 3M%":12,
        "Avg Traded Value 20D Cr":16,"Liquidity Tag":12,
        "Momentum Rank":14,
        "Risk Tag":10, "Cam Setup":20, "Total Appearances":16, "Unique Scanners":15,
        "Scanner List":40, "Capture Price":13, "Current Price":13,
        "Cam H3":11, "Cam H4":11, "Cam L3":11, "Cam L4":11,
        "Ideal Enter Price":15, "Possible Sell Value":16, "Stop Loss Value":14,
        "Since Capture%":13, "1D%":8, "1W%":8, "1M%":8,
        "RSI 14":9, "ADX 14":9, "+DI 14":9, "-DI 14":9, "ATR 14":10, "NATR 14":10,
        "Signal":22, "Setup Signal":22, "Core Signal":22,
        "Signal Quality":18, "Signal Regime":14, "Win Prob%":10, "Hist Precision%":13, "Exp 5D%":10, "Exp 10D%":10, "WF Samples":10,
        "AI Decision":16, "AI Conf%":10,
    }
    hist_widths.update({
        "Fundamental Score":14, "Fundamental Quality Tag":18, "Growth Tag":12, "Valuation Tag":13,
        "Balance Sheet Tag":16, "Cashflow Tag":14, "Ownership Tag":14, "Fundamental Risk Tag":16,
        "Investability Tag":20, "Early Entry OK":14, "Tech + Fundamental Score":18,
        "Decision Guardrail":34, "Final Confidence Tag":18, "Final Signal":22,
        "Fundamental Updated At":18, "Fundamental Freshness":16,
    })
    dhc = {h: i for i, h in enumerate(DASHBOARD_HISTORY_HEADERS)}
    for h, w in hist_widths.items():
        if h in dhc:
            ws.column_dimensions[get_column_letter(dhc[h]+1)].width = w

    # Auto-filter on header row
    last_col = get_column_letter(len(DASHBOARD_HISTORY_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    ws.freeze_panes = "A2"
    return ws

def ensure_validation_sheet(wb):
    if "Validation" in wb.sheetnames:
        ws = wb["Validation"]
    else:
        insert_at = min(len(wb.sheetnames), 2)
        ws = wb.create_sheet("Validation", insert_at)
    for ci in range(1, min(max(ws.max_column or 0, len(VALIDATION_HEADERS)) + 5, 80)):
        cell = ws.cell(row=1, column=ci)
        cell.value = None
        cell.fill = PatternFill()
        cell.font = Font()
        cell.alignment = Alignment()
        cell.border = Border()
    for ci, h in enumerate(VALIDATION_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="6A1B9A")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _thin_border()
    widths = {
        "Snapshot At":18,"Iteration":10,"Mode":14,"Checked Rows":12,"Matched Rows":12,
        "Mismatch Rows":13,"Unresolved Rows":14,"Latest Session":18,"Status":12,"Details":80,
    }
    vh = {h: i for i, h in enumerate(VALIDATION_HEADERS)}
    for h, w in widths.items():
        if h in vh:
            ws.column_dimensions[get_column_letter(vh[h] + 1)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(VALIDATION_HEADERS))}1"
    return ws

def build_dashboard_only_workbook(source_wb):
    """Create a workbook containing only the Dashboard sheet."""
    src_ws = ensure_dashboard_sheet(source_wb)
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Dashboard"

    max_col = len(DASHBOARD_HEADERS)
    last_row = 1
    for ri in range(src_ws.max_row, 1, -1):
        sym_val = src_ws.cell(row=ri, column=DC["Symbol"] + 1).value
        name_val = src_ws.cell(row=ri, column=DC["Name"] + 1).value
        if str(sym_val or "").strip() or str(name_val or "").strip():
            last_row = ri
            break

    for ri in range(1, last_row + 1):
        for ci in range(1, max_col + 1):
            cell = src_ws.cell(row=ri, column=ci)
            tgt = out_ws.cell(row=ri, column=ci, value=cell.value)
            if ri == 1 or cell.value not in (None, ""):
                if cell.has_style:
                    tgt._style = copy(cell._style)
                if cell.number_format:
                    tgt.number_format = cell.number_format
                if cell.alignment:
                    tgt.alignment = copy(cell.alignment)
                if cell.font:
                    tgt.font = copy(cell.font)
                if cell.fill:
                    tgt.fill = copy(cell.fill)
                if cell.border:
                    tgt.border = copy(cell.border)
                if cell.hyperlink:
                    try:
                        tgt._hyperlink = copy(cell.hyperlink)
                    except Exception:
                        pass

    for ci in range(1, max_col + 1):
        col_letter = get_column_letter(ci)
        dim = src_ws.column_dimensions.get(col_letter)
        if not dim:
            continue
        out_dim = out_ws.column_dimensions[col_letter]
        out_dim.width = dim.width
        out_dim.hidden = dim.hidden
        out_dim.bestFit = dim.bestFit

    for ri in range(1, last_row + 1):
        dim = src_ws.row_dimensions.get(ri)
        if not dim:
            continue
        out_dim = out_ws.row_dimensions[ri]
        out_dim.height = dim.height
        out_dim.hidden = dim.hidden

    out_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter and src_ws.auto_filter.ref:
        last_col = get_column_letter(max_col)
        out_ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    return out_wb

def _validation_values_match(expected, actual) -> bool:
    if expected in (None, "") and actual in (None, ""):
        return True
    ef = _as_float(expected)
    af = _as_float(actual)
    if ef is not None and af is not None:
        return abs(ef - af) <= SELF_VALIDATION_NUMERIC_TOLERANCE
    return str(expected or "").strip() == str(actual or "").strip()

def append_validation_result(wb, result: dict):
    ws = ensure_validation_sheet(wb)
    ri = ws.max_row + 1
    row = [
        result.get("Snapshot At"),
        result.get("Iteration"),
        result.get("Mode"),
        result.get("Checked Rows"),
        result.get("Matched Rows"),
        result.get("Mismatch Rows"),
        result.get("Unresolved Rows"),
        result.get("Latest Session"),
        result.get("Status"),
        result.get("Details"),
    ]
    for ci, value in enumerate(row, 1):
        cell = ws.cell(row=ri, column=ci, value=value)
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", wrap_text=(ci == len(row)))
    status = str(result.get("Status") or "").upper()
    if status == "PASS":
        ws.cell(row=ri, column=VALIDATION_HEADERS.index("Status") + 1).fill = PatternFill("solid", fgColor="C8E6C9")
    elif status in ("WARN", "FAIL"):
        ws.cell(row=ri, column=VALIDATION_HEADERS.index("Status") + 1).fill = PatternFill("solid", fgColor="FFCDD2")

def _select_validation_sample(eligible_rows, sample_size):
    if sample_size <= 0 or not eligible_rows:
        return []
    if len(eligible_rows) <= sample_size:
        return list(eligible_rows)
    picks = []
    last_idx = len(eligible_rows) - 1
    for pos in range(sample_size):
        idx = round(pos * last_idx / max(sample_size - 1, 1))
        item = eligible_rows[idx]
        if item not in picks:
            picks.append(item)
    cursor = 0
    while len(picks) < sample_size and cursor < len(eligible_rows):
        item = eligible_rows[cursor]
        if item not in picks:
            picks.append(item)
        cursor += 1
    return picks[:sample_size]

def run_post_run_validation(wb, iteration: int) -> dict:
    snap_ts = fmt_dt()
    result = {
        "Snapshot At": snap_ts,
        "Iteration": iteration,
        "Mode": f"fresh-sample-{SELF_VALIDATION_SAMPLE_SIZE}",
        "Checked Rows": 0,
        "Matched Rows": 0,
        "Mismatch Rows": 0,
        "Unresolved Rows": 0,
        "Latest Session": "",
        "Status": "SKIP",
        "Details": "Validation disabled or no eligible active rows.",
    }
    if not SELF_VALIDATION_ENABLED:
        append_validation_result(wb, result)
        return result

    dash_ws = ensure_dashboard_sheet(wb)
    eligible = []
    for ri in range(2, dash_ws.max_row + 1):
        symbol = str(dash_ws.cell(row=ri, column=DC["Symbol"] + 1).value or "").strip()
        in_screener = str(dash_ws.cell(row=ri, column=DC["In Screener?"] + 1).value or "").strip()
        signal = str(dash_ws.cell(row=ri, column=DC["Signal"] + 1).value or "").strip()
        if in_screener != "Yes" or not symbol or symbol.startswith(SENTINEL):
            continue
        if signal in ("No Data", "Symbol Not Found", "Error"):
            continue
        eligible.append((ri, symbol))
    sample = _select_validation_sample(eligible, SELF_VALIDATION_SAMPLE_SIZE)
    if not sample:
        append_validation_result(wb, result)
        return result

    compare_keys = [
        "Current Price", "1D%", "1W%",
        "RSI 14", "ADX 14", "+DI 14", "-DI 14", "ATR 14", "NATR 14",
        "Signal", "Setup Signal", "Core Signal",
        "Signal Quality", "Signal Regime",
        "RS Tag", "RS vs NIFTY 1M%", "RS vs NIFTY 3M%", "RS vs Sector 1M%", "RS vs Sector 3M%",
        "Avg Traded Value 20D Cr", "Liquidity Tag",
    ]
    details = []
    latest_sessions = []

    for ri, symbol in sample:
        row = {h: dash_ws.cell(row=ri, column=DC[h] + 1).value for h in DASHBOARD_HEADERS}
        hist = fetch_history_fresh(symbol)
        result["Checked Rows"] += 1
        if not hist:
            result["Unresolved Rows"] += 1
            if len(details) < SELF_VALIDATION_MAX_DETAILS:
                details.append(f"{symbol}: fresh fetch unavailable")
            continue
        last_ts = _hist_last_ts(hist)
        if last_ts is not None:
            latest_sessions.append(str(last_ts)[:10])
        fresh_metrics = compute_metrics(hist, symbol=symbol, benchmark_refresh=True)
        mismatches = [key for key in compare_keys if not _validation_values_match(row.get(key), fresh_metrics.get(key))]
        if mismatches:
            result["Mismatch Rows"] += 1
            if len(details) < SELF_VALIDATION_MAX_DETAILS:
                details.append(f"{symbol}: {', '.join(mismatches[:4])}")
        else:
            result["Matched Rows"] += 1

    if latest_sessions:
        result["Latest Session"] = max(latest_sessions)
    result["Status"] = "PASS" if result["Mismatch Rows"] == 0 else "WARN"
    if result["Checked Rows"] == 0:
        result["Status"] = "SKIP"
    result["Details"] = " | ".join(details) if details else (
        f"All {result['Matched Rows']}/{result['Checked Rows']} sampled active rows matched fresh recomputation."
        if result["Checked Rows"] else "No eligible rows."
    )
    append_validation_result(wb, result)
    return result

def update_dashboard(wb):
    """Scan all scanner sheets, aggregate per-stock stats, compute actionable columns, write Dashboard + history."""
    dash_ws = ensure_dashboard_sheet(wb)
    hist_ws = ensure_dashboard_history_sheet(wb)
    now = fmt_dt()
    _dashboard_float_fields = {
        "Consensus Score",
        "Capture Price", "Current Price",
        "Cam H3", "Cam H4", "Cam L3", "Cam L4",
        "Ideal Enter Price", "Possible Sell Value", "Stop Loss Value",
        "Since Capture%", "1D%", "1W%", "1M%", "3M%", "6M%", "1Y%",
        "RSI 14", "ADX 14", "+DI 14", "-DI 14", "ATR 14", "NATR 14",
        "Win Prob%", "Hist Precision%", "Exp 5D%", "Exp 10D%",
        "RS vs NIFTY 1M%", "RS vs NIFTY 3M%", "RS vs Sector 1M%", "RS vs Sector 3M%",
        "Avg Traded Value 20D Cr",
        "AI Conf%",
    }
    _dashboard_int_fields = {"Momentum Rank", "Days Tracked", "Total Appearances", "Unique Scanners", "WF Samples"}
    _dashboard_float_fields.update({
        "Market Cap Cr", "PE", "PB", "EV/EBITDA", "Dividend Yield%", "Book Value", "EPS TTM",
        "Sales TTM Cr", "Profit TTM Cr", "OPM%", "NPM%", "ROE%", "ROCE%", "ROE 3Y%", "ROE Last Year%",
        "Sales CAGR 3Y%", "Sales CAGR 5Y%", "Sales Growth TTM%",
        "Profit CAGR 3Y%", "Profit CAGR 5Y%", "Profit Growth TTM%",
        "Debt/Equity", "Borrowings Cr", "Interest Coverage", "Asset Turnover",
        "CFO TTM Cr", "FCF TTM Cr", "CFO/PAT", "CFO/OP%",
        "Debtor Days", "Inventory Days", "Cash Conversion Cycle",
        "Promoter Holding%", "Promoter Holding Change%", "Promoter Pledge%",
        "Fundamental Score", "Tech + Fundamental Score",
    })
    _dashboard_int_fields.update({"Pros Count", "Cons Count"})

    def _is_valid_date(val):
        """Check if a value looks like a date string (not a number)."""
        if val is None:
            return False
        s = str(val).strip()
        if not s:
            return False
        # Must contain date separators, not just a number
        if '-' in s and ':' in s and len(s) >= 10:
            return True
        return False

    def _normalize_dashboard_snapshot_row(row):
        normalized = {}
        for h in DASHBOARD_HEADERS:
            val = row.get(h)
            if h in _dashboard_float_fields:
                normalized[h] = _as_float(val)
            elif h in _dashboard_int_fields:
                normalized[h] = _as_int(val)
            else:
                normalized[h] = val
        return normalized

    # -- Collect per-stock data from all scanner sheets ----------------------
    agg = {}
    _AI_SCORE_MAP = {"STRONG BUY":3,"BUY":2,"ACCUMULATE":1,"HOLD":0,"REDUCE":-1,"SELL":-2,"STRONG SELL":-3}

    display_cols = [
        "Current Price","Since Capture%","1D%","1W%","1M%","3M%","6M%","1Y%",
        "RSI 14","ADX 14","+DI 14","-DI 14","ATR 14","NATR 14",
        "Signal","Setup Signal","Core Signal",
        "Signal Quality","Signal Regime","Win Prob%","Hist Precision%","Exp 5D%","Exp 10D%","WF Samples",
        "Sector","Industry","Sector Benchmark","RS Tag","RS vs NIFTY 1M%","RS vs NIFTY 3M%","RS vs Sector 1M%","RS vs Sector 3M%","Avg Traded Value 20D Cr","Liquidity Tag",
        "AI Decision","AI Conf%","Vol Ratio 20","52W High Dist%","Last Updated",
    ] + FUNDAMENTAL_SCANNER_FIELDS

    for sc in SCANNERS:
        sname = sc["name"][:31]
        sid = sc["id"]
        tf = _SCANNER_TF.get(sid, "D")
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        for ri in range(2, ws.max_row + 1):
            sym  = str(ws.cell(row=ri, column=C["Symbol"]+1).value or "").strip()
            name = str(ws.cell(row=ri, column=C["Name"]+1).value or "").strip()
            if not sym and not name:
                continue
            key = _stock_key(sym, name)
            if not key:
                continue

            in_screener = str(ws.cell(row=ri, column=C["In Screener?"]+1).value or "").strip()
            first_cap = ws.cell(row=ri, column=C["First Captured"]+1).value
            last_seen = ws.cell(row=ri, column=C["Last Seen"]+1).value
            cap_price = ws.cell(row=ri, column=C["Capture Price"]+1).value
            cur_price = ws.cell(row=ri, column=C["Current Price"]+1).value
            since_cap = ws.cell(row=ri, column=C["Since Capture%"]+1).value
            ai_dec = str(ws.cell(row=ri, column=C["AI Decision"]+1).value or "").strip().upper()
            vol_ratio = ws.cell(row=ri, column=C["Vol Ratio 20"]+1).value
            adx_val = ws.cell(row=ri, column=C["ADX 14"]+1).value
            pdi_val = ws.cell(row=ri, column=C["+DI 14"]+1).value
            mdi_val = ws.cell(row=ri, column=C["-DI 14"]+1).value
            atr_val = ws.cell(row=ri, column=C["ATR 14"]+1).value
            natr_val = ws.cell(row=ri, column=C["NATR 14"]+1).value
            dist_52w = ws.cell(row=ri, column=C["52W High Dist%"]+1).value
            rsi_val = ws.cell(row=ri, column=C["RSI 14"]+1).value

            is_active = (in_screener == "Yes")
            row_metric_values = {
                "Current Price": cur_price,
                "Since Capture%": since_cap,
                "1D%": ws.cell(row=ri, column=C["1D%"]+1).value,
                "1W%": ws.cell(row=ri, column=C["1W%"]+1).value,
                "1M%": ws.cell(row=ri, column=C["1M%"]+1).value,
                "3M%": ws.cell(row=ri, column=C["3M%"]+1).value,
                "6M%": ws.cell(row=ri, column=C["6M%"]+1).value,
                "1Y%": ws.cell(row=ri, column=C["1Y%"]+1).value,
                "RSI 14": rsi_val,
                "ADX 14": adx_val,
                "+DI 14": pdi_val,
                "-DI 14": mdi_val,
                "ATR 14": atr_val,
                "NATR 14": natr_val,
                "Signal": ws.cell(row=ri, column=C["Signal"]+1).value,
                "Setup Signal": ws.cell(row=ri, column=C["Setup Signal"]+1).value,
                "Core Signal": ws.cell(row=ri, column=C["Core Signal"]+1).value,
                "Signal Quality": ws.cell(row=ri, column=C["Signal Quality"]+1).value,
                "Signal Regime": ws.cell(row=ri, column=C["Signal Regime"]+1).value,
                "Win Prob%": ws.cell(row=ri, column=C["Win Prob%"]+1).value,
                "Hist Precision%": ws.cell(row=ri, column=C["Hist Precision%"]+1).value,
                "Exp 5D%": ws.cell(row=ri, column=C["Exp 5D%"]+1).value,
                "Exp 10D%": ws.cell(row=ri, column=C["Exp 10D%"]+1).value,
                "WF Samples": ws.cell(row=ri, column=C["WF Samples"]+1).value,
                "Sector": ws.cell(row=ri, column=C["Sector"]+1).value,
                "Industry": ws.cell(row=ri, column=C["Industry"]+1).value,
                "Sector Benchmark": ws.cell(row=ri, column=C["Sector Benchmark"]+1).value,
                "RS Tag": ws.cell(row=ri, column=C["RS Tag"]+1).value,
                "RS vs NIFTY 1M%": ws.cell(row=ri, column=C["RS vs NIFTY 1M%"]+1).value,
                "RS vs NIFTY 3M%": ws.cell(row=ri, column=C["RS vs NIFTY 3M%"]+1).value,
                "RS vs Sector 1M%": ws.cell(row=ri, column=C["RS vs Sector 1M%"]+1).value,
                "RS vs Sector 3M%": ws.cell(row=ri, column=C["RS vs Sector 3M%"]+1).value,
                "Avg Traded Value 20D Cr": ws.cell(row=ri, column=C["Avg Traded Value 20D Cr"]+1).value,
                "Liquidity Tag": ws.cell(row=ri, column=C["Liquidity Tag"]+1).value,
                "AI Decision": ws.cell(row=ri, column=C["AI Decision"]+1).value,
                "AI Conf%": ws.cell(row=ri, column=C["AI Conf%"]+1).value,
                "Vol Ratio 20": vol_ratio,
                "52W High Dist%": dist_52w,
                "Last Updated": ws.cell(row=ri, column=C["Last Updated"]+1).value,
            }
            for fund_col in FUNDAMENTAL_SCANNER_FIELDS:
                if fund_col in C:
                    row_metric_values[fund_col] = ws.cell(row=ri, column=C[fund_col]+1).value
            if not is_active:
                row_metric_values = {col: None for col in row_metric_values}

            if key not in agg:
                agg[key] = {
                    "Symbol": sym if sym and not sym.startswith(SENTINEL) else "",
                    "Name": name,
                    "First Captured": first_cap if _is_valid_date(first_cap) else None,
                    "Last Seen": last_seen,
                    "scanners": set(), "scanner_ids": set(),
                    "scanner_yes_counts": {},  # sname -> count of In Screener=Yes
                    "timeframes": set(), "active_timeframes": set(), "appearances": 0,
                    "ai_decisions": [], "ai_confs": [],
                    "Capture Price": cap_price, "Current Price": cur_price,
                    "_display_priority": 1 if is_active else 0,
                    "_is_active_any": bool(is_active),
                }
                agg[key].update(row_metric_values)
            else:
                entry = agg[key]
                if _is_valid_date(first_cap) and (not _is_valid_date(entry["First Captured"]) or str(first_cap) < str(entry["First Captured"])):
                    entry["First Captured"] = first_cap
                if last_seen and (not entry["Last Seen"] or str(last_seen) > str(entry["Last Seen"])):
                    entry["Last Seen"] = last_seen
                if cap_price and not entry["Capture Price"]:
                    entry["Capture Price"] = cap_price
                current_priority = 1 if is_active else 0
                existing_priority = int(entry.get("_display_priority", 0))
                if current_priority > existing_priority:
                    for col in display_cols:
                        entry[col] = None
                    entry["_display_priority"] = current_priority
                if current_priority >= int(entry.get("_display_priority", 0)):
                    for col in display_cols:
                        val = row_metric_values.get(col)
                        if val is not None:
                            entry[col] = val
                    entry["_display_priority"] = current_priority
                if not entry["Symbol"] and sym and not sym.startswith(SENTINEL):
                    entry["Symbol"] = sym
                entry["_is_active_any"] = bool(entry.get("_is_active_any")) or bool(is_active)

            agg[key]["scanners"].add(sname)
            agg[key]["scanner_ids"].add(sid)
            agg[key]["timeframes"].add(tf)
            if is_active:
                agg[key]["active_timeframes"].add(tf)
                agg[key]["appearances"] += 1
                agg[key]["scanner_yes_counts"][sname] = agg[key]["scanner_yes_counts"].get(sname, 0) + 1
                if ai_dec in _AI_SCORE_MAP:
                    agg[key]["ai_decisions"].append(_AI_SCORE_MAP[ai_dec])
                    conf = ws.cell(row=ri, column=C["AI Conf%"]+1).value
                    if conf is not None:
                        try: agg[key]["ai_confs"].append(float(conf))
                        except (ValueError, TypeError): pass

    # -- Load existing dashboard totals ------------------------------------
    existing_totals = {}
    existing_dashboard_rows = {}
    for ri in range(2, dash_ws.max_row + 1):
        d_sym  = str(dash_ws.cell(row=ri, column=DC["Symbol"]+1).value or "").strip()
        d_name = str(dash_ws.cell(row=ri, column=DC["Name"]+1).value or "").strip()
        d_key  = _stock_key(d_sym, d_name)
        if not d_key: continue
        raw_dashboard_row = {h: dash_ws.cell(row=ri, column=DC[h]+1).value for h in DASHBOARD_HEADERS}
        existing_dashboard_rows[d_key] = _normalize_dashboard_snapshot_row(raw_dashboard_row)
        old_total = existing_dashboard_rows[d_key].get("Total Appearances")
        old_first = existing_dashboard_rows[d_key].get("First Captured")
        old_cap   = existing_dashboard_rows[d_key].get("Capture Price")
        existing_totals[d_key] = {
            "total": old_total or 0,
            "first_captured": old_first, "capture_price": old_cap,
        }

    # -- Merge lifetime totals ---------------------------------------------
    for key, entry in agg.items():
        cur_app = entry["appearances"]
        if key in existing_totals:
            entry["appearances"] = existing_totals[key]["total"] + cur_app
            et = existing_totals[key]
            if _is_valid_date(et["first_captured"]) and (not _is_valid_date(entry["First Captured"]) or str(et["first_captured"]) < str(entry["First Captured"])):
                entry["First Captured"] = et["first_captured"]
            if et["capture_price"] and not entry["Capture Price"]:
                entry["Capture Price"] = et["capture_price"]
        else:
            entry["appearances"] = max(cur_app, 1)

    # -- Compute derived columns -------------------------------------------
    def _safe_float(v):
        if v is None: return None
        try: return float(v)
        except (ValueError, TypeError): return None

    def _price_mtf_value(entry, prev_row=None):
        sym = str(entry.get("Symbol") or "").strip()
        if sym and sym in _price_mtf_data:
            return _price_mtf_data[sym]
        if prev_row:
            prev_value = prev_row.get("Historical MTF")
            if prev_value not in (None, ""):
                return prev_value
        return ""

    def _format_mtf_alignment(timeframes):
        return " ".join(
            f"{tf}{'âœ…' if tf in timeframes else 'âŒ'}"
            for tf in ("D", "W", "M")
        )

    def _format_mtf_alignment_ascii(timeframes):
        return " ".join(
            f"{tf} {'OK' if tf in timeframes else 'NO'}"
            for tf in ("D", "W", "M")
        )

    def _format_mtf_alignment_ticks(timeframes):
        return " ".join(
            f"{tf}{chr(0x2705) if tf in timeframes else chr(0x274C)}"
            for tf in ("D", "W", "M")
        )

    def _derive_camarilla_plan(price, bb_signal, cam_info):
        plan = {
            "Cam Setup": "—",
            "Cam H3": None, "Cam H4": None, "Cam L3": None, "Cam L4": None,
            "Ideal Enter Price": None, "Possible Sell Value": None, "Stop Loss Value": None,
        }
        if not cam_info:
            return plan

        h3 = _safe_float(cam_info.get("Cam H3"))
        h4 = _safe_float(cam_info.get("Cam H4"))
        l3 = _safe_float(cam_info.get("Cam L3"))
        l4 = _safe_float(cam_info.get("Cam L4"))
        plan.update({"Cam H3": h3, "Cam H4": h4, "Cam L3": l3, "Cam L4": l4})
        if price is None or any(v is None for v in (h3, h4, l3, l4)):
            return plan

        step = h4 - h3
        buf_base = max(price, h4, l3)
        buffer = round(buf_base * 0.002, 2) if buf_base > 0 else 0.0
        bb = str(bb_signal or "").upper()

        if price >= h4:
            zone = "ABOVE H4"
        elif price >= h3:
            zone = "BETWEEN H3-H4"
        elif price <= l4:
            zone = "BELOW L4"
        elif price <= l3:
            zone = "BETWEEN L4-L3"
        else:
            zone = "INSIDE L3-H3"

        plan["Cam Setup"] = zone

        if "SQUEEZE BREAK" in bb:
            plan["Cam Setup"] = "SQUEEZE + H4 BREAK" if price >= h4 else "WATCH H4 BREAK"
            plan["Ideal Enter Price"] = h4
            plan["Possible Sell Value"] = round(h4 + step, 2) if step > 0 else None
            plan["Stop Loss Value"] = round(max(h3 - buffer, 0.0), 2)
        elif "SQUEEZE" in bb and price >= h3:
            plan["Cam Setup"] = "SQUEEZE NEAR H4"
            plan["Ideal Enter Price"] = h4
            plan["Possible Sell Value"] = round(h4 + step, 2) if step > 0 else None
            plan["Stop Loss Value"] = round(max(h3 - buffer, 0.0), 2)
        elif "BUY ZONE" in bb or "OVERSOLD" in bb:
            if price <= l4:
                plan["Cam Setup"] = "OVERSOLD BELOW L4"
            elif price <= l3:
                plan["Cam Setup"] = "OVERSOLD AT L3"
            else:
                plan["Cam Setup"] = "WATCH L3 SUPPORT"
            plan["Ideal Enter Price"] = l3
            plan["Possible Sell Value"] = h3
            plan["Stop Loss Value"] = round(max(l4 - buffer, 0.0), 2)
        elif any(tok in bb for tok in ("SELL ZONE", "STRETCHED", "NEAR HIGH")):
            if price >= h4:
                plan["Cam Setup"] = "AT/ABOVE H4 RESISTANCE"
            elif price >= h3:
                plan["Cam Setup"] = "UPPER BAND UNDER H4"

        return plan

    def _momentum_tag(rank, total):
        rank = _as_int(rank)
        total = max(_as_int(total) or 0, 1)
        if rank is None or rank <= 0:
            return ""
        pct = rank / total
        if pct <= 0.10:
            return "ELITE"
        if pct <= 0.20:
            return "STRONG"
        if pct <= 0.40:
            return "HEALTHY"
        if pct <= 0.60:
            return "NEUTRAL"
        if pct <= 0.80:
            return "WEAK"
        return "LAGGING"

    for key, e in agg.items():
        e["In Screener?"] = "Yes" if e.get("_is_active_any") else "No"
        prev_row = existing_dashboard_rows.get(key)
        if not e.get("_is_active_any") and prev_row:
            for h in DASHBOARD_HEADERS:
                if h in ("Symbol", "Name", "In Screener?"):
                    continue
                e[h] = prev_row.get(h)
            e["Historical MTF"] = _price_mtf_value(e, prev_row)
            e["_preserve_dashboard"] = True
            e["In Screener?"] = "No"
            e["_mom_score"] = _as_float(prev_row.get("Momentum Rank"))
            if e["_mom_score"] is None:
                e["_mom_score"] = 999999.0
            continue

        # ① Consensus Score (1-10): average AI score across scanners, mapped to 1-10
        if e["ai_decisions"]:
            avg_ai = sum(e["ai_decisions"]) / len(e["ai_decisions"])
            e["Consensus Score"] = round(min(10, max(1, (avg_ai + 3) / 6 * 9 + 1)), 1)
        else:
            e["Consensus Score"] = None

        # ② MTF Alignment: D✅/❌ W✅/❌ M✅/❌
        tfs = e["active_timeframes"]
        e["MTF Alignment"] = (
            ("D✅" if "D" in tfs else "D❌") + " " +
            ("W✅" if "W" in tfs else "W❌") + " " +
            ("M✅" if "M" in tfs else "M❌")
        )
        mtf_count = len(tfs)
        e["MTF Alignment"] = _format_mtf_alignment_ticks(tfs)
        e["Historical MTF"] = _price_mtf_value(e, prev_row)

        # ③ Days Tracked
        fc = e.get("First Captured")
        if fc:
            try:
                fc_dt = datetime.strptime(str(fc)[:19], "%Y-%m-%d %H:%M:%S")
                e["Days Tracked"] = (ist_now().replace(tzinfo=None) - fc_dt).days
            except Exception:
                e["Days Tracked"] = None
        else:
            e["Days Tracked"] = None

        # ④ Momentum Rank (raw score, will be ranked later)
        d1, w1, m1 = _safe_float(e.get("1D%")), _safe_float(e.get("1W%")), _safe_float(e.get("1M%"))
        rs_nifty_3m = _safe_float(e.get("RS vs NIFTY 3M%"))
        rs_sector_3m = _safe_float(e.get("RS vs Sector 3M%"))
        mom_score = 0
        if d1 is not None: mom_score += d1 * 3  # weight recent more
        if w1 is not None: mom_score += w1 * 2
        if m1 is not None: mom_score += m1 * 1
        if rs_nifty_3m is not None: mom_score += rs_nifty_3m * 0.8
        if rs_sector_3m is not None: mom_score += rs_sector_3m * 0.8
        e["_mom_score"] = mom_score

        # ⑤ Volume Buzz
        vr = _safe_float(e.get("Vol Ratio 20"))
        if vr is not None and vr > 1.5: e["Volume Buzz"] = "🔥🔥 High"
        elif vr is not None and vr > 1.2: e["Volume Buzz"] = "🔥 Above Avg"
        elif vr is not None and vr >= 0.8: e["Volume Buzz"] = "— Normal"
        elif vr is not None: e["Volume Buzz"] = "💤 Low"
        else: e["Volume Buzz"] = "—"

        # ⑥ Since Capture Trend
        sc_pct = _safe_float(e.get("Since Capture%"))
        if sc_pct is not None:
            if sc_pct >= 20: e["Since Capture Trend"] = "↑↑ Strong Gain"
            elif sc_pct >= 5: e["Since Capture Trend"] = "↑ Gaining"
            elif sc_pct >= -5: e["Since Capture Trend"] = "→ Flat"
            elif sc_pct >= -20: e["Since Capture Trend"] = "↓ Losing"
            else: e["Since Capture Trend"] = "↓↓ Heavy Loss"
        else:
            e["Since Capture Trend"] = "—"

        # ⑦ Best Scanner (scanner where this stock appears "Yes" most often)
        syc = e.get("scanner_yes_counts", {})
        if syc:
            e["Best Scanner"] = max(syc, key=syc.get)
        elif e["scanners"]:
            e["Best Scanner"] = sorted(e["scanners"])[0]  # fallback: alphabetical
        else:
            e["Best Scanner"] = ""

        # ⑧ Risk Tag
        rsi = _safe_float(e.get("RSI 14"))
        adx = _safe_float(e.get("ADX 14"))
        pdi = _safe_float(e.get("+DI 14"))
        mdi = _safe_float(e.get("-DI 14"))
        natr = _safe_float(e.get("NATR 14"))
        dist = _safe_float(e.get("52W High Dist%"))
        win_prob = _safe_float(e.get("Win Prob%"))
        hist_precision = _safe_float(e.get("Hist Precision%"))
        signal_quality = str(e.get("Signal Quality", "") or "")
        signal_regime = str(e.get("Signal Regime", "") or "")
        rs_tag = str(e.get("RS Tag", "") or "")
        avg_tv_cr = _safe_float(e.get("Avg Traded Value 20D Cr"))
        risk_flags = 0
        if rsi is not None and (rsi > 78 or rsi < 30): risk_flags += 2
        if adx is not None and adx < 16: risk_flags += 1
        if natr is not None and natr >= 8: risk_flags += 1
        if adx is not None and adx >= 20 and pdi is not None and mdi is not None and mdi > pdi: risk_flags += 1
        if dist is not None and dist < 2: risk_flags += 1  # near 52W high resistance
        if dist is not None and dist > 30: risk_flags += 2  # deep hole
        if win_prob is not None and win_prob < QUALITY_GATE_MIN_WIN_PROB: risk_flags += 1
        if hist_precision is not None and hist_precision < QUALITY_GATE_MIN_HIST_PRECISION: risk_flags += 1
        if signal_regime == "HIGH-VOL": risk_flags += 2
        elif signal_regime == "CHOPPY": risk_flags += 1
        if "REJECT" in signal_quality: risk_flags += 2
        if avg_tv_cr is not None and avg_tv_cr < 1: risk_flags += 2
        elif avg_tv_cr is not None and avg_tv_cr < 5: risk_flags += 1
        if "LAGGING" in rs_tag.upper(): risk_flags += 2
        elif "WEAK RS" in rs_tag.upper(): risk_flags += 1
        if risk_flags >= 3: e["Risk Tag"] = "🔴 HIGH"
        elif risk_flags >= 1: e["Risk Tag"] = "🟡 MED"
        else: e["Risk Tag"] = "🟢 LOW"

        # ⑨ Screener Link
        sym = e.get("Symbol", "")
        disp = sym[4:] if sym.startswith("BSE:") else sym
        e["Screener Link"] = f"https://www.screener.in/company/{disp}/" if disp else ""

        # ⑩ BB Signal (Bollinger Band composite -- Dashboard only)
        sym_lookup = e.get("Symbol", "")
        bb_info = _bb_data.get(sym_lookup, {})
        bb_pctb = bb_info.get("BB %B")
        bb_width = bb_info.get("BB Width")
        bb_width_pctl = bb_info.get("BB Width Pctl")
        bb_squeeze = bb_info.get("BB Squeeze")
        if bb_pctb is not None and rsi is not None:
            if bb_pctb > 1.0 and rsi > 70:
                e["BB Signal"] = "🔴 SELL ZONE"       # above upper + overbought → take profit
            elif bb_squeeze and bb_pctb >= 0.8:
                e["BB Signal"] = "🚀 SQUEEZE BREAK"
            elif bb_pctb > 1.0:
                e["BB Signal"] = "🟡 STRETCHED"        # above upper but momentum OK → trail stop
            elif bb_pctb >= 0.85 and rsi > 65:
                e["BB Signal"] = "⚠️ NEAR HIGH"        # approaching upper + elevated RSI → caution
            elif bb_pctb < 0.0 and rsi < 35:
                e["BB Signal"] = "💰 BUY ZONE"         # below lower + oversold → reversal candidate
            elif bb_pctb < 0.0:
                e["BB Signal"] = "📉 OVERSOLD"          # below lower band
            elif bb_squeeze or (bb_width_pctl is not None and bb_width_pctl <= 15):
                e["BB Signal"] = "🟢 SQUEEZE"           # bands tight → breakout imminent
            else:
                e["BB Signal"] = "— NORMAL"
        elif bb_pctb is not None:
            # Have BB but no RSI — still show band position
            if bb_pctb > 1.0:
                e["BB Signal"] = "🟡 STRETCHED"
            elif bb_pctb < 0.0:
                e["BB Signal"] = "📉 OVERSOLD"
            elif bb_squeeze or (bb_width_pctl is not None and bb_width_pctl <= 15):
                e["BB Signal"] = "🟢 SQUEEZE"
            else:
                e["BB Signal"] = "— NORMAL"
        else:
            e["BB Signal"] = "—"

        cam_info = _cam_data.get(sym_lookup, {})
        e.update(_derive_camarilla_plan(_safe_float(e.get("Current Price")), e.get("BB Signal", ""), cam_info))
        _apply_combined_final_signal(e, mtf_count=mtf_count)

    # -- Assign Momentum Rank (1 = best) -----------------------------------
    ranked = sorted(
        [k for k in agg.keys() if not agg[k].get("_preserve_dashboard")],
        key=lambda k: _as_float(agg[k].get("_mom_score")) or 0.0,
        reverse=True,
    )
    for rank, k in enumerate(ranked, 1):
        agg[k]["Momentum Rank"] = rank
        agg[k]["Momentum Tag"] = _momentum_tag(rank, len(ranked))
    for key, e in agg.items():
        if not e.get("Momentum Tag"):
            e["Momentum Tag"] = _momentum_tag(e.get("Momentum Rank"), len(ranked))

    # ⑩ Quick Action -- synthesized from consensus + MTF + momentum + risk
    for key, e in agg.items():
        if e.get("_preserve_dashboard"):
            e["Quick Action"] = existing_dashboard_rows.get(key, {}).get("Quick Action", "👀 WATCH")
            continue
        cs = _safe_float(e.get("Consensus Score"))
        mom_rank = _as_int(e.get("Momentum Rank")) or 999
        total = len(ranked) if ranked else 1
        mom_pct = mom_rank / total  # lower = better
        mtf_count = len(e["active_timeframes"])
        risk = e.get("Risk Tag", "")
        signal_quality = str(e.get("Signal Quality", "") or "")
        signal_regime = str(e.get("Signal Regime", "") or "")
        signal_family = _signal_family_key(e.get("Signal"))
        win_prob = _safe_float(e.get("Win Prob%"))
        rs_nifty_3m = _safe_float(e.get("RS vs NIFTY 3M%"))
        rs_sector_3m = _safe_float(e.get("RS vs Sector 3M%"))
        liquidity_tag = str(e.get("Liquidity Tag", "") or "")
        rs_tag = str(e.get("RS Tag", "") or "")
        liquidity_blocked = any(tok in liquidity_tag.upper() for tok in ("THIN", "ILLIQUID"))
        rs_blocked = any(tok in rs_tag.upper() for tok in ("LAGGING", "WEAK RS"))
        final_signal = str(e.get("Final Signal", "") or "").upper()
        fund_risk = str(e.get("Fundamental Risk Tag", "") or "").upper()
        combined_score = _safe_float(e.get("Tech + Fundamental Score"))

        if "REJECT" in signal_quality:
            if "HIGH" in risk or signal_regime == "HIGH-VOL":
                e["Quick Action"] = "🔴 AVOID"
            elif cs is not None and cs < 4:
                e["Quick Action"] = "⚠️ CAUTION"
            else:
                e["Quick Action"] = "👀 WATCH"
            continue
        elif final_signal == "AVOID" or fund_risk == "HIGH":
            e["Quick Action"] = "🔴 AVOID" if "HIGH" in risk or final_signal == "AVOID" else "⚠️ CAUTION"
        elif final_signal == "HIGH CONVICTION BUY" and not liquidity_blocked and not rs_blocked and "HIGH" not in risk:
            e["Quick Action"] = "🟢 BUY NOW"
        elif final_signal == "BUY" and not liquidity_blocked and "HIGH" not in risk and (combined_score is None or combined_score >= 60):
            e["Quick Action"] = "🟡 ACCUMULATE"
        elif final_signal == "EARLY ENTRY" and not liquidity_blocked and "HIGH" not in risk:
            e["Quick Action"] = "🟡 ACCUMULATE" if mom_pct <= 0.5 else "👀 WATCH"
        elif final_signal in ("WAIT FOR TECHNICALS", "WATCH"):
            e["Quick Action"] = "👀 WATCH"
        elif final_signal == "TRADE ONLY":
            e["Quick Action"] = "⚠️ CAUTION" if "HIGH" in risk else "👀 WATCH"
        elif liquidity_blocked or rs_blocked:
            e["Quick Action"] = "⚠️ CAUTION" if "HIGH" not in risk else "🔴 AVOID"
        elif signal_family in ("BREAKOUT", "STRONG BUY") and cs is not None and cs >= 7 and mtf_count >= 2 and mom_pct <= 0.3 and "HIGH" not in risk and (win_prob is None or win_prob >= 60) and (rs_nifty_3m is None or rs_nifty_3m >= 0) and (rs_sector_3m is None or rs_sector_3m >= 0):
            e["Quick Action"] = "🟢 BUY NOW"
        elif signal_family in ("BREAKOUT", "STRONG BUY", "BUY") and cs is not None and cs >= 6 and mtf_count >= 2 and "HIGH" not in risk and signal_regime != "HIGH-VOL" and not liquidity_blocked:
            e["Quick Action"] = "🟡 ACCUMULATE"
        elif signal_family in ("BUY", "PULLBACK", "OVERSOLD") and cs is not None and cs >= 5 and "HIGH" not in risk:
            e["Quick Action"] = "👀 WATCH"
        elif cs is not None and cs < 4 and "HIGH" in risk:
            e["Quick Action"] = "🔴 AVOID"
        elif cs is not None and cs < 4:
            e["Quick Action"] = "⚠️ CAUTION"
        else:
            e["Quick Action"] = "👀 WATCH"

    # -- Sort by Consensus Score desc -> Total Appearances desc -------------
    sorted_stocks = sorted(
        agg.items(),
        key=lambda x: (
            1 if x[1].get("In Screener?") == "Yes" else 0,
            _as_float(x[1].get("Consensus Score")) or 0.0,
            _as_float(x[1].get("Tech + Fundamental Score")) or 0.0,
            _as_int(x[1].get("appearances")) or 0,
            len(x[1]["scanners"]),
        ),
        reverse=True,
    )

    # -- Write Dashboard ---------------------------------------------------
    for ri in range(2, dash_ws.max_row + 1):
        for ci in range(1, min(max(dash_ws.max_column or 0, len(DASHBOARD_HEADERS)) + 5, 200)):
            cell = dash_ws.cell(row=ri, column=ci)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()

    for idx, (key, entry) in enumerate(sorted_stocks, start=2):
        scanner_list = ", ".join(sorted(entry["scanners"]))
        row_data = {
            "Symbol": entry.get("Symbol", ""), "Name": entry.get("Name", ""),
            "In Screener?": entry.get("In Screener?", "No"),
            "Quick Action": entry.get("Quick Action", ""),
            "Consensus Score": entry.get("Consensus Score"),
            "MTF Alignment": entry.get("MTF Alignment", ""),
            "Historical MTF": entry.get("Historical MTF", ""),
            "Sector": entry.get("Sector", ""),
            "Industry": entry.get("Industry", ""),
            "Sector Benchmark": entry.get("Sector Benchmark", ""),
            "RS Tag": entry.get("RS Tag", ""),
            "RS vs NIFTY 1M%": entry.get("RS vs NIFTY 1M%"),
            "RS vs NIFTY 3M%": entry.get("RS vs NIFTY 3M%"),
            "RS vs Sector 1M%": entry.get("RS vs Sector 1M%"),
            "RS vs Sector 3M%": entry.get("RS vs Sector 3M%"),
            "Avg Traded Value 20D Cr": entry.get("Avg Traded Value 20D Cr"),
            "Liquidity Tag": entry.get("Liquidity Tag", ""),
            "Momentum Rank": entry.get("Momentum Rank"),
            "Momentum Tag": entry.get("Momentum Tag") or _momentum_tag(entry.get("Momentum Rank"), len(sorted_stocks)),
            "Risk Tag": entry.get("Risk Tag", ""),
            "BB Signal": entry.get("BB Signal", "—"),
            "Cam Setup": entry.get("Cam Setup", "—"),
            "Volume Buzz": entry.get("Volume Buzz", ""),
            "Since Capture Trend": entry.get("Since Capture Trend", ""),
            "First Captured": entry.get("First Captured"),
            "Days Tracked": entry.get("Days Tracked"),
            "Last Seen": entry.get("Last Seen"),
            "Total Appearances": entry["appearances"],
            "Unique Scanners": len(entry["scanners"]),
            "Scanner List": scanner_list,
            "Best Scanner": entry.get("Best Scanner", ""),
            "Capture Price": entry.get("Capture Price"),
            "Current Price": entry.get("Current Price"),
            "Cam H3": entry.get("Cam H3"),
            "Cam H4": entry.get("Cam H4"),
            "Cam L3": entry.get("Cam L3"),
            "Cam L4": entry.get("Cam L4"),
            "Ideal Enter Price": entry.get("Ideal Enter Price"),
            "Possible Sell Value": entry.get("Possible Sell Value"),
            "Stop Loss Value": entry.get("Stop Loss Value"),
            "Since Capture%": entry.get("Since Capture%"),
            "1D%": entry.get("1D%"), "1W%": entry.get("1W%"), "1M%": entry.get("1M%"),
            "3M%": entry.get("3M%"), "6M%": entry.get("6M%"), "1Y%": entry.get("1Y%"),
            "RSI 14": entry.get("RSI 14"),
            "ADX 14": entry.get("ADX 14"),
            "+DI 14": entry.get("+DI 14"),
            "-DI 14": entry.get("-DI 14"),
            "ATR 14": entry.get("ATR 14"),
            "NATR 14": entry.get("NATR 14"),
            "Signal": entry.get("Signal"),
            "Setup Signal": entry.get("Setup Signal"),
            "Core Signal": entry.get("Core Signal"),
            "Signal Quality": entry.get("Signal Quality"),
            "Signal Regime": entry.get("Signal Regime"),
            "Win Prob%": entry.get("Win Prob%"),
            "Hist Precision%": entry.get("Hist Precision%"),
            "Exp 5D%": entry.get("Exp 5D%"),
            "Exp 10D%": entry.get("Exp 10D%"),
            "WF Samples": entry.get("WF Samples"),
            "AI Decision": entry.get("AI Decision"),
            "AI Conf%": entry.get("AI Conf%"),
            "Screener Link": entry.get("Screener Link", ""),
            "Last Updated": (now if entry.get("In Screener?") == "Yes" else entry.get("Last Updated")),
        }
        for fund_col in FUNDAMENTAL_RAW_FIELDS + FUNDAMENTAL_DERIVED_FIELDS:
            row_data[fund_col] = entry.get(fund_col)
        for h, ci in DC.items():
            cell = dash_ws.cell(row=idx, column=ci+1)
            cell.value = row_data.get(h)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", wrap_text=(h in ("Scanner List","Screener Link")))

        # -- Quick Action colour ------------------------------------------
        qa = str(row_data.get("Quick Action", ""))
        qa_cell = dash_ws.cell(row=idx, column=DC["Quick Action"]+1)
        if "BUY NOW" in qa:
            qa_cell.fill = PatternFill("solid", fgColor="1B5E20"); qa_cell.font = Font(bold=True, color="FFFFFF")
        elif "ACCUMULATE" in qa:
            qa_cell.fill = PatternFill("solid", fgColor="4CAF50"); qa_cell.font = Font(bold=True, color="FFFFFF")
        elif "WATCH" in qa:
            qa_cell.fill = PatternFill("solid", fgColor="FFF9C4"); qa_cell.font = Font(bold=True, color="F57F17")
        elif "CAUTION" in qa:
            qa_cell.fill = PatternFill("solid", fgColor="FFCC80"); qa_cell.font = Font(bold=True, color="E65100")
        elif "AVOID" in qa:
            qa_cell.fill = PatternFill("solid", fgColor="B71C1C"); qa_cell.font = Font(bold=True, color="FFFFFF")

        # -- Consensus Score colour ---------------------------------------
        cs = _as_float(entry.get("Consensus Score"))
        cs_cell = dash_ws.cell(row=idx, column=DC["Consensus Score"]+1)
        if cs is not None:
            if cs >= 8: cs_cell.fill = PatternFill("solid", fgColor="1B5E20"); cs_cell.font = Font(bold=True, color="FFFFFF")
            elif cs >= 6: cs_cell.fill = PatternFill("solid", fgColor="A5D6A7"); cs_cell.font = Font(bold=True, color="1B5E20")
            elif cs >= 4: cs_cell.fill = PatternFill("solid", fgColor="FFF9C4")
            else: cs_cell.fill = PatternFill("solid", fgColor="FFCDD2")

        # -- Risk Tag colour ----------------------------------------------
        rt = str(entry.get("Risk Tag", ""))
        rt_cell = dash_ws.cell(row=idx, column=DC["Risk Tag"]+1)
        if "HIGH" in rt: rt_cell.fill = PatternFill("solid", fgColor="FFCDD2"); rt_cell.font = Font(bold=True, color="B71C1C")
        elif "MED" in rt: rt_cell.fill = PatternFill("solid", fgColor="FFF9C4"); rt_cell.font = Font(color="F57F17")
        elif "LOW" in rt: rt_cell.fill = PatternFill("solid", fgColor="C8E6C9"); rt_cell.font = Font(color="1B5E20")

        fr = str(row_data.get("Fundamental Risk Tag", "") or "").upper()
        if "Fundamental Risk Tag" in DC and fr:
            fr_cell = dash_ws.cell(row=idx, column=DC["Fundamental Risk Tag"]+1)
            if fr == "HIGH":
                fr_cell.fill = PatternFill("solid", fgColor="FFCDD2"); fr_cell.font = Font(bold=True, color="B71C1C")
            elif fr == "MED":
                fr_cell.fill = PatternFill("solid", fgColor="FFF9C4"); fr_cell.font = Font(color="F57F17")
            elif fr == "LOW":
                fr_cell.fill = PatternFill("solid", fgColor="C8E6C9"); fr_cell.font = Font(color="1B5E20")

        fsig = str(row_data.get("Final Signal", "") or "").upper()
        if "Final Signal" in DC and fsig:
            fs_cell = dash_ws.cell(row=idx, column=DC["Final Signal"]+1)
            if fsig == "HIGH CONVICTION BUY":
                fs_cell.fill = PatternFill("solid", fgColor="1B5E20"); fs_cell.font = Font(bold=True, color="FFFFFF")
            elif fsig in ("BUY", "EARLY ENTRY", "ACCUMULATE"):
                fs_cell.fill = PatternFill("solid", fgColor="C8E6C9"); fs_cell.font = Font(bold=True, color="1B5E20")
            elif fsig in ("TRADE ONLY", "WAIT FOR TECHNICALS", "WATCH", "TECHNICAL ONLY"):
                fs_cell.fill = PatternFill("solid", fgColor="FFF9C4"); fs_cell.font = Font(color="F57F17")
            elif fsig == "AVOID":
                fs_cell.fill = PatternFill("solid", fgColor="FFCDD2"); fs_cell.font = Font(bold=True, color="B71C1C")

        guard = str(row_data.get("Decision Guardrail", ""))
        if "Decision Guardrail" in DC and guard:
            guard_cell = dash_ws.cell(row=idx, column=DC["Decision Guardrail"]+1)
            if guard.startswith("PASS"):
                guard_cell.fill = PatternFill("solid", fgColor="C8E6C9"); guard_cell.font = Font(color="1B5E20")
            elif guard.startswith("CAUTION"):
                guard_cell.fill = PatternFill("solid", fgColor="FFF9C4"); guard_cell.font = Font(color="F57F17")
            elif guard.startswith("BLOCK"):
                guard_cell.fill = PatternFill("solid", fgColor="FFCDD2"); guard_cell.font = Font(bold=True, color="B71C1C")

        conf = str(row_data.get("Final Confidence Tag", ""))
        if "Final Confidence Tag" in DC and conf:
            conf_cell = dash_ws.cell(row=idx, column=DC["Final Confidence Tag"]+1)
            if conf == "VERY HIGH":
                conf_cell.fill = PatternFill("solid", fgColor="1B5E20"); conf_cell.font = Font(bold=True, color="FFFFFF")
            elif conf == "HIGH":
                conf_cell.fill = PatternFill("solid", fgColor="C8E6C9"); conf_cell.font = Font(color="1B5E20")
            elif conf == "MED":
                conf_cell.fill = PatternFill("solid", fgColor="FFF9C4"); conf_cell.font = Font(color="F57F17")
            elif conf in ("LOW", "BLOCKED"):
                conf_cell.fill = PatternFill("solid", fgColor="FFCDD2"); conf_cell.font = Font(bold=True, color="B71C1C")

        fscore = _as_float(row_data.get("Fundamental Score"))
        if "Fundamental Score" in DC and fscore is not None:
            fc = dash_ws.cell(row=idx, column=DC["Fundamental Score"]+1)
            if fscore >= 75:
                fc.fill = PatternFill("solid", fgColor="1B5E20"); fc.font = Font(bold=True, color="FFFFFF")
            elif fscore >= 60:
                fc.fill = PatternFill("solid", fgColor="C8E6C9"); fc.font = Font(color="1B5E20")
            elif fscore >= 45:
                fc.fill = PatternFill("solid", fgColor="FFF9C4"); fc.font = Font(color="F57F17")
            else:
                fc.fill = PatternFill("solid", fgColor="FFCDD2"); fc.font = Font(bold=True, color="B71C1C")

        rs_tag = str(row_data.get("RS Tag", ""))
        if "RS Tag" in DC:
            rs_cell = dash_ws.cell(row=idx, column=DC["RS Tag"]+1)
            if "STRONG VS BOTH" in rs_tag.upper():
                rs_cell.fill = PatternFill("solid", fgColor="1B5E20"); rs_cell.font = Font(bold=True, color="FFFFFF")
            elif "RS LEADER" in rs_tag.upper():
                rs_cell.fill = PatternFill("solid", fgColor="4CAF50"); rs_cell.font = Font(bold=True, color="FFFFFF")
            elif "MIXED" in rs_tag.upper():
                rs_cell.fill = PatternFill("solid", fgColor="FFF9C4"); rs_cell.font = Font(color="F57F17")
            elif "WEAK RS" in rs_tag.upper():
                rs_cell.fill = PatternFill("solid", fgColor="FFCC80"); rs_cell.font = Font(bold=True, color="E65100")
            elif "LAGGING" in rs_tag.upper():
                rs_cell.fill = PatternFill("solid", fgColor="FFCDD2"); rs_cell.font = Font(bold=True, color="B71C1C")

        liquidity = str(row_data.get("Liquidity Tag", ""))
        if "Liquidity Tag" in DC:
            liq_cell = dash_ws.cell(row=idx, column=DC["Liquidity Tag"]+1)
            if "DEEP" in liquidity.upper():
                liq_cell.fill = PatternFill("solid", fgColor="1B5E20"); liq_cell.font = Font(bold=True, color="FFFFFF")
            elif "LIQUID" in liquidity.upper():
                liq_cell.fill = PatternFill("solid", fgColor="4CAF50"); liq_cell.font = Font(bold=True, color="FFFFFF")
            elif "ADEQUATE" in liquidity.upper():
                liq_cell.fill = PatternFill("solid", fgColor="FFF9C4"); liq_cell.font = Font(color="F57F17")
            elif "THIN" in liquidity.upper():
                liq_cell.fill = PatternFill("solid", fgColor="FFCC80"); liq_cell.font = Font(bold=True, color="E65100")
            elif "ILLIQUID" in liquidity.upper():
                liq_cell.fill = PatternFill("solid", fgColor="FFCDD2"); liq_cell.font = Font(bold=True, color="B71C1C")

        ins = str(row_data.get("In Screener?", ""))
        ins_cell = dash_ws.cell(row=idx, column=DC["In Screener?"]+1)
        if ins == "Yes":
            ins_cell.fill = PatternFill("solid", fgColor="C8E6C9")
            ins_cell.font = Font(bold=True, color="1B5E20")
        elif ins == "No":
            ins_cell.fill = PatternFill("solid", fgColor="FFCDD2")
            ins_cell.font = Font(bold=True, color="B71C1C")

        # -- BB Signal colour ---------------------------------------------
        bb = str(row_data.get("BB Signal", ""))
        bb_cell = dash_ws.cell(row=idx, column=DC["BB Signal"]+1)
        if "SELL ZONE" in bb:
            bb_cell.fill = PatternFill("solid", fgColor="B71C1C"); bb_cell.font = Font(bold=True, color="FFFFFF")
        elif "STRETCHED" in bb:
            bb_cell.fill = PatternFill("solid", fgColor="FFB74D"); bb_cell.font = Font(bold=True, color="E65100")
        elif "NEAR HIGH" in bb:
            bb_cell.fill = PatternFill("solid", fgColor="FFF9C4"); bb_cell.font = Font(bold=True, color="F57F17")
        elif "BUY ZONE" in bb:
            bb_cell.fill = PatternFill("solid", fgColor="1B5E20"); bb_cell.font = Font(bold=True, color="FFFFFF")
        elif "OVERSOLD" in bb:
            bb_cell.fill = PatternFill("solid", fgColor="E3F2FD"); bb_cell.font = Font(bold=True, color="1565C0")
        elif "SQUEEZE" in bb:
            bb_cell.fill = PatternFill("solid", fgColor="4CAF50"); bb_cell.font = Font(bold=True, color="FFFFFF")

        cam_setup = str(row_data.get("Cam Setup", ""))
        cam_cell = dash_ws.cell(row=idx, column=DC["Cam Setup"]+1)
        if "SQUEEZE + H4 BREAK" in cam_setup or "OVERSOLD AT L3" in cam_setup:
            cam_cell.fill = PatternFill("solid", fgColor="1B5E20"); cam_cell.font = Font(bold=True, color="FFFFFF")
        elif "WATCH" in cam_setup or "NEAR H4" in cam_setup or "BETWEEN" in cam_setup or "INSIDE" in cam_setup:
            cam_cell.fill = PatternFill("solid", fgColor="FFF9C4"); cam_cell.font = Font(color="F57F17")
        elif "RESISTANCE" in cam_setup or "UNDER H4" in cam_setup or "BELOW L4" in cam_setup:
            cam_cell.fill = PatternFill("solid", fgColor="FFCC80"); cam_cell.font = Font(bold=True, color="E65100")

        # -- Volume Buzz colour -------------------------------------------
        vb = str(entry.get("Volume Buzz", ""))
        vb_cell = dash_ws.cell(row=idx, column=DC["Volume Buzz"]+1)
        if "🔥🔥" in vb: vb_cell.fill = PatternFill("solid", fgColor="FF6F00"); vb_cell.font = Font(bold=True, color="FFFFFF")
        elif "🔥" in vb: vb_cell.fill = PatternFill("solid", fgColor="FFB74D")

        # -- Total Appearances colour -------------------------------------
        ta = _as_int(entry.get("appearances")) or 0
        ta_cell = dash_ws.cell(row=idx, column=DC["Total Appearances"]+1)
        if ta >= 50: ta_cell.fill = PatternFill("solid", fgColor="1B5E20"); ta_cell.font = Font(bold=True, color="FFFFFF")
        elif ta >= 20: ta_cell.fill = PatternFill("solid", fgColor="4CAF50"); ta_cell.font = Font(bold=True, color="FFFFFF")
        elif ta >= 10: ta_cell.fill = PatternFill("solid", fgColor="A5D6A7"); ta_cell.font = Font(bold=True, color="1B5E20")
        elif ta >= 5: ta_cell.fill = PatternFill("solid", fgColor="C8E6C9")

        # -- Unique Scanners colour ---------------------------------------
        us = len(entry["scanners"])
        us_cell = dash_ws.cell(row=idx, column=DC["Unique Scanners"]+1)
        if us >= 10: us_cell.fill = PatternFill("solid", fgColor="1565C0"); us_cell.font = Font(bold=True, color="FFFFFF")
        elif us >= 5: us_cell.fill = PatternFill("solid", fgColor="42A5F5"); us_cell.font = Font(bold=True, color="FFFFFF")
        elif us >= 3: us_cell.fill = PatternFill("solid", fgColor="90CAF9")

        # -- Signal / AI colours ------------------------------------------
        for signal_col in ("Signal", "Setup Signal", "Core Signal"):
            sf, ff = _rule_signal_style(row_data.get(signal_col, ""))
            if sf and signal_col in DC:
                sc = dash_ws.cell(row=idx, column=DC[signal_col] + 1)
                sc.fill = sf
                sc.font = ff
        ai_dec = row_data.get("AI Decision", "")
        if ai_dec:
            af, afnt = _ai_style(ai_dec)
            if af:
                dash_ws.cell(row=idx, column=DC["AI Decision"]+1).fill = af
                dash_ws.cell(row=idx, column=DC["AI Decision"]+1).font = afnt
        ai_conf = _as_float(row_data.get("AI Conf%"))
        if ai_conf is not None:
            try:
                cof = _ai_conf_color(int(ai_conf))
                if cof: dash_ws.cell(row=idx, column=DC["AI Conf%"]+1).fill = cof
            except (ValueError, TypeError): pass

        sq = str(row_data.get("Signal Quality", ""))
        if "REJECT" in sq and "Signal Quality" in DC:
            qc = dash_ws.cell(row=idx, column=DC["Signal Quality"]+1)
            qc.fill = PatternFill("solid", fgColor="FFCDD2")
            qc.font = Font(bold=True, color="B71C1C")
        elif "HIGH" in sq and "Signal Quality" in DC:
            qc = dash_ws.cell(row=idx, column=DC["Signal Quality"]+1)
            qc.fill = PatternFill("solid", fgColor="C8E6C9")
            qc.font = Font(bold=True, color="1B5E20")
        elif "MED" in sq and "Signal Quality" in DC:
            qc = dash_ws.cell(row=idx, column=DC["Signal Quality"]+1)
            qc.fill = PatternFill("solid", fgColor="E8F5E9")
            qc.font = Font(color="2E7D32")
        elif "LOW" in sq and "Signal Quality" in DC:
            qc = dash_ws.cell(row=idx, column=DC["Signal Quality"]+1)
            qc.fill = PatternFill("solid", fgColor="FFF9C4")
            qc.font = Font(color="F57F17")

        sr = str(row_data.get("Signal Regime", ""))
        if "Signal Regime" in DC:
            rc = dash_ws.cell(row=idx, column=DC["Signal Regime"]+1)
            if sr == "TRENDING":
                rc.fill = PatternFill("solid", fgColor="C8E6C9")
                rc.font = Font(bold=True, color="1B5E20")
            elif sr == "HIGH-VOL":
                rc.fill = PatternFill("solid", fgColor="FFCC80")
                rc.font = Font(bold=True, color="E65100")
            elif sr == "CHOPPY":
                rc.fill = PatternFill("solid", fgColor="FFF9C4")
                rc.font = Font(color="F57F17")

        for prob_col in ("Win Prob%", "Hist Precision%"):
            if prob_col not in DC:
                continue
            pv = _as_float(row_data.get(prob_col))
            if pv is None:
                continue
            pc = dash_ws.cell(row=idx, column=DC[prob_col]+1)
            if pv >= 65:
                pc.fill = PatternFill("solid", fgColor="C8E6C9")
                pc.font = Font(bold=True, color="1B5E20")
            elif pv >= 55:
                pc.fill = PatternFill("solid", fgColor="E8F5E9")
                pc.font = Font(color="2E7D32")
            else:
                pc.fill = PatternFill("solid", fgColor="FFCDD2")
                pc.font = Font(bold=True, color="B71C1C")

        # -- Momentum Rank colour (top 20% green) ------------------------
        mr = _as_int(entry.get("Momentum Rank")) or 999
        mr_cell = dash_ws.cell(row=idx, column=DC["Momentum Rank"]+1)
        total = len(sorted_stocks) or 1
        if mr <= total * 0.1: mr_cell.fill = PatternFill("solid", fgColor="1B5E20"); mr_cell.font = Font(bold=True, color="FFFFFF")
        elif mr <= total * 0.2: mr_cell.fill = PatternFill("solid", fgColor="4CAF50"); mr_cell.font = Font(bold=True, color="FFFFFF")
        elif mr <= total * 0.4: mr_cell.fill = PatternFill("solid", fgColor="C8E6C9")

        mt = str(entry.get("Momentum Tag", "") or "").upper()
        if "Momentum Tag" in DC and mt:
            mt_cell = dash_ws.cell(row=idx, column=DC["Momentum Tag"]+1)
            if mt == "ELITE":
                mt_cell.fill = PatternFill("solid", fgColor="1B5E20"); mt_cell.font = Font(bold=True, color="FFFFFF")
            elif mt == "STRONG":
                mt_cell.fill = PatternFill("solid", fgColor="4CAF50"); mt_cell.font = Font(bold=True, color="FFFFFF")
            elif mt == "HEALTHY":
                mt_cell.fill = PatternFill("solid", fgColor="C8E6C9"); mt_cell.font = Font(color="1B5E20")
            elif mt == "NEUTRAL":
                mt_cell.fill = PatternFill("solid", fgColor="FFF9C4"); mt_cell.font = Font(color="F57F17")
            elif mt == "WEAK":
                mt_cell.fill = PatternFill("solid", fgColor="FFCC80"); mt_cell.font = Font(color="E65100")
            elif mt == "LAGGING":
                mt_cell.fill = PatternFill("solid", fgColor="FFCDD2"); mt_cell.font = Font(bold=True, color="B71C1C")

        # -- Since Capture% + return period colours (green/red) -----------
        for pct_col in ["Since Capture%", "1D%", "1W%", "1M%", "3M%", "6M%", "1Y%"]:
            if pct_col not in DC: continue
            pv_f = _as_float(row_data.get(pct_col))
            if pv_f is None: continue
            pc = dash_ws.cell(row=idx, column=DC[pct_col]+1)
            if pv_f >= 20: pc.font = Font(bold=True, color="1B5E20")
            elif pv_f >= 5: pc.font = Font(color="2E7D32")
            elif pv_f >= 0: pc.font = Font(color="4CAF50")
            elif pv_f >= -5: pc.font = Font(color="E65100")
            elif pv_f >= -20: pc.font = Font(color="D32F2F")
            else: pc.font = Font(bold=True, color="B71C1C")

        # -- Alternate row shading ----------------------------------------
        if idx % 2 == 0:
            for ci in range(1, len(DASHBOARD_HEADERS) + 1):
                c = dash_ws.cell(row=idx, column=ci)
                if c.fill == PatternFill():
                    c.fill = PatternFill("solid", fgColor="F5F5F5")

    # -- Append to Dashboard History (append-only) -------------------------
    for key, entry in sorted_stocks:
        scanner_list = ", ".join(sorted(entry["scanners"]))
        hist_row = [
            now, entry.get("Symbol", ""), entry.get("Name", ""),
            entry.get("In Screener?", "No"),
            entry.get("Quick Action", ""), entry.get("Consensus Score"),
            entry.get("MTF Alignment", ""), entry.get("Historical MTF", ""),
            entry.get("Sector", ""), entry.get("Industry", ""), entry.get("Sector Benchmark", ""), entry.get("RS Tag", ""),
            entry.get("RS vs NIFTY 1M%"), entry.get("RS vs NIFTY 3M%"),
            entry.get("RS vs Sector 1M%"), entry.get("RS vs Sector 3M%"),
            entry.get("Avg Traded Value 20D Cr"), entry.get("Liquidity Tag", ""),
            entry.get("Momentum Rank"),
            entry.get("Risk Tag", ""), entry.get("Cam Setup", "—"),
            entry["appearances"], len(entry["scanners"]), scanner_list,
            entry.get("Capture Price"), entry.get("Current Price"),
            entry.get("Cam H3"), entry.get("Cam H4"), entry.get("Cam L3"), entry.get("Cam L4"),
            entry.get("Ideal Enter Price"), entry.get("Possible Sell Value"), entry.get("Stop Loss Value"),
            entry.get("Since Capture%"), entry.get("1D%"), entry.get("1W%"),
            entry.get("1M%"), entry.get("RSI 14"), entry.get("ADX 14"),
            entry.get("+DI 14"), entry.get("-DI 14"), entry.get("ATR 14"), entry.get("NATR 14"),
            entry.get("Signal"),
            entry.get("Setup Signal"),
            entry.get("Core Signal"),
            entry.get("Signal Quality"), entry.get("Signal Regime"),
            entry.get("Win Prob%"), entry.get("Hist Precision%"),
            entry.get("Exp 5D%"), entry.get("Exp 10D%"), entry.get("WF Samples"),
            entry.get("AI Decision"), entry.get("AI Conf%"),
        ]
        hist_row.extend(entry.get(h) for h in FUNDAMENTAL_SUMMARY_FIELDS)
        ri = hist_ws.max_row + 1
        for ci, v in enumerate(hist_row, 1):
            c = hist_ws.cell(row=ri, column=ci, value=v)
            c.border = _thin_border(); c.alignment = Alignment(horizontal="center")

    # -- Re-apply auto-filter over full data range (Google Sheets needs this) --
    last_dash_col = get_column_letter(len(DASHBOARD_HEADERS))
    last_dash_row = max(dash_ws.max_row, 1)
    dash_ws.auto_filter.ref = f"A1:{last_dash_col}{last_dash_row}"

    last_hist_col = get_column_letter(len(DASHBOARD_HISTORY_HEADERS))
    last_hist_row = max(hist_ws.max_row, 1)
    hist_ws.auto_filter.ref = f"A1:{last_hist_col}{last_hist_row}"

    print(f"  [STAT] Dashboard updated: {len(sorted_stocks)} stocks, "
          f"{hist_ws.max_row - 1} total history rows")
    return len(sorted_stocks)

def init_workbook():
    result = s3_download_excel(retries=3)
    if result is None:
        # Transient S3 failure -- DO NOT create fresh workbook (would wipe data)
        raise RuntimeError(
            "[FAIL] FATAL: Cannot download workbook from S3 after 3 retries. "
            "Refusing to start with empty workbook to prevent data loss. "
            "Check S3 connectivity and try again."
        )
    if result == _S3_FIRST_RUN:
        # Genuinely first run -- create fresh workbook
        print("  📂 Creating brand new workbook (first run)")
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
    else:
        wb = result
    ensure_price_history_sheet(wb)
    ensure_dashboard_sheet(wb)
    ensure_dashboard_history_sheet(wb)
    ensure_validation_sheet(wb)
    for sc in SCANNERS: ensure_scanner_sheet(wb, sc)
    return wb

def _stock_key(sym, name):
    sym = (str(sym) if sym else "").strip()
    name = (str(name) if name else "").strip()
    if sym and not sym.startswith("~NOFOUND"): return sym
    if name: return " ".join(name.upper().split())
    return None

def load_sheet_stocks(ws) -> dict:
    stocks = {}
    for ri in range(2, ws.max_row + 1):
        sym  = ws.cell(row=ri, column=C["Symbol"]+1).value
        name = ws.cell(row=ri, column=C["Name"]+1).value
        key = _stock_key(sym, name)
        if not key: continue
        stocks[key] = {"row": ri, "data": {h: ws.cell(row=ri, column=ci+1).value for h, ci in C.items()}}
    return stocks

def _rule_signal_style(signal: str):
    s = str(signal or "")
    if "BREAKOUT" in s or "STRONG BUY" in s:
        return PatternFill("solid",fgColor="C8E6C9"), Font(bold=True,color="1B5E20")
    if "BUY"     in s: return PatternFill("solid",fgColor="DCEDC8"), Font(bold=True,color="33691E")
    if "SELL"    in s: return PatternFill("solid",fgColor="FFCDD2"), Font(bold=True,color="B71C1C")
    if "OVERSOLD" in s: return PatternFill("solid",fgColor="E1F5FE"), Font(color="0277BD")
    if any(x in s for x in ("HOLD","PULLBACK","WEAK")):
        return PatternFill("solid",fgColor="FFF9C4"), Font(color="F57F17")
    return None, None

def write_stock_row(ws, ri: int, data: dict, bg_color: str = None, apply_static_style: bool = False):
    fill = _row_fill(bg_color)
    for h, ci in C.items():
        cell = ws.cell(row=ri, column=ci+1)
        cell.value = data.get(h)
        if apply_static_style:
            cell.border = _THIN_BORDER
            cell.alignment = _ALIGN_CENTER_WRAP if h == "AI Reason" else _ALIGN_CENTER
            if fill and h not in ("AI Decision","AI Conf%"):
                cell.fill = fill
    # Rule signal colours
    for signal_col in ("Signal", "Setup Signal", "Core Signal"):
        if signal_col not in C:
            continue
        sf, ff = _rule_signal_style(data.get(signal_col, ""))
        if sf:
            sc = ws.cell(row=ri, column=C[signal_col] + 1)
            sc.fill = sf
            sc.font = ff
    # AI Decision colour
    ai_dec = data.get("AI Decision","")
    if ai_dec:
        af, afnt = _ai_style(ai_dec)
        if af:
            ws.cell(row=ri, column=C["AI Decision"]+1).fill = af
            ws.cell(row=ri, column=C["AI Decision"]+1).font = afnt
    # AI Confidence colour
    ai_conf = data.get("AI Conf%")
    if ai_conf is not None:
        cof = _ai_conf_color(int(ai_conf))
        if cof: ws.cell(row=ri, column=C["AI Conf%"]+1).fill = cof
    # Combined final signal / fundamental risk colours
    final_signal = str(data.get("Final Signal", "") or "").upper()
    if final_signal and "Final Signal" in C:
        fc = ws.cell(row=ri, column=C["Final Signal"]+1)
        if "HIGH CONVICTION" in final_signal:
            fc.fill = PatternFill("solid", fgColor="1B5E20"); fc.font = Font(bold=True, color="FFFFFF")
        elif final_signal in ("BUY", "EARLY ENTRY", "ACCUMULATE"):
            fc.fill = PatternFill("solid", fgColor="C8E6C9"); fc.font = Font(bold=True, color="1B5E20")
        elif final_signal in ("TRADE ONLY", "WAIT FOR TECHNICALS", "WATCH"):
            fc.fill = PatternFill("solid", fgColor="FFF9C4"); fc.font = Font(color="F57F17")
        elif final_signal == "AVOID":
            fc.fill = PatternFill("solid", fgColor="FFCDD2"); fc.font = Font(bold=True, color="B71C1C")
    fund_risk = str(data.get("Fundamental Risk Tag", "") or "").upper()
    if fund_risk and "Fundamental Risk Tag" in C:
        rc = ws.cell(row=ri, column=C["Fundamental Risk Tag"]+1)
        if fund_risk == "HIGH":
            rc.fill = PatternFill("solid", fgColor="FFCDD2"); rc.font = Font(bold=True, color="B71C1C")
        elif fund_risk == "MED":
            rc.fill = PatternFill("solid", fgColor="FFF9C4"); rc.font = Font(color="F57F17")
        elif fund_risk == "LOW":
            rc.fill = PatternFill("solid", fgColor="C8E6C9"); rc.font = Font(color="1B5E20")
    # AI Reason -- left-aligned, italic
    if data.get("AI Reason"):
        rc = ws.cell(row=ri, column=C["AI Reason"]+1)
        rc.font = Font(italic=True, size=9); rc.alignment = _ALIGN_LEFT_WRAP

def _clear_runtime_cells(ws, ri: int, signal_value: str, clear_metrics: bool = True, core_signal_value: str | None = None):
    transient_cols = [
        "Current Price", "1D%", "1W%", "1M%", "3M%", "6M%", "1Y%", "2Y%", "3Y%",
        "Avg Weekly%", "Avg Monthly%", "Avg 3M%", "Avg 6M%", "Avg 1Y%",
        "RSI 14", "MA 20", "MA 50", "MA 200", "ADX 14", "Vol Ratio 20",
        "MACD Line", "MACD Hist", "52W High Dist%", "20D Breakout%",
        "ATR 14", "NATR 14", "+DI 14", "-DI 14",
        "Setup Signal", "Signal Quality", "Signal Regime", "Win Prob%", "Hist Precision%", "Exp 5D%", "Exp 10D%", "WF Samples",
        "Sector", "Industry", "Sector Benchmark", "RS Tag",
        "RS vs NIFTY 1M%", "RS vs NIFTY 3M%", "RS vs Sector 1M%", "RS vs Sector 3M%",
        "Avg Traded Value 20D Cr", "Liquidity Tag",
        "AI Decision", "AI Reason", "AI Conf%", "Last Updated"
    ] + FUNDAMENTAL_SCANNER_FIELDS
    if clear_metrics:
        for col in transient_cols:
            if col not in C:
                continue
            cell = ws.cell(row=ri, column=C[col] + 1)
            cell.value = None
            cell.fill = _BLANK_FILL
            cell.font = _DEFAULT_FONT

    signal_values = {
        "Signal": signal_value,
        "Setup Signal": signal_value,
        "Core Signal": core_signal_value if core_signal_value is not None else signal_value,
    }
    for signal_col, signal_text in signal_values.items():
        if signal_col not in C:
            continue
        sig_cell = ws.cell(row=ri, column=C[signal_col] + 1)
        sig_cell.value = signal_text
        sig_cell.fill = _BLANK_FILL
        sig_cell.font = _DEFAULT_FONT
        sf, ff = _rule_signal_style(signal_text)
        if sf:
            sig_cell.fill = sf
            sig_cell.font = ff

def append_price_history(ws, scanner_name: str, data: dict):
    ri = ws.max_row + 1
    row_data = {
        "Snapshot At": fmt_dt(),
        "Scanner": scanner_name,
        "Symbol": data.get("Symbol"),
        "Name": data.get("Name"),
        "In Screener?": data.get("In Screener?"),
        "Capture Price": data.get("Capture Price"),
        "Current Price": data.get("Current Price"),
        "Since Capture%": data.get("Since Capture%"),
        "1D%": data.get("1D%"),
        "1W%": data.get("1W%"),
        "1M%": data.get("1M%"),
        "3M%": data.get("3M%"),
        "6M%": data.get("6M%"),
        "1Y%": data.get("1Y%"),
        "Cam H3": data.get("Cam H3"),
        "Cam H4": data.get("Cam H4"),
        "Cam L3": data.get("Cam L3"),
        "Cam L4": data.get("Cam L4"),
        "RSI 14": data.get("RSI 14"),
        "ADX 14": data.get("ADX 14"),
        "+DI 14": data.get("+DI 14"),
        "-DI 14": data.get("-DI 14"),
        "ATR 14": data.get("ATR 14"),
        "NATR 14": data.get("NATR 14"),
        "Vol Ratio 20": data.get("Vol Ratio 20"),
        "MACD Line": data.get("MACD Line"),
        "MACD Hist": data.get("MACD Hist"),
        "52W High Dist%": data.get("52W High Dist%"),
        "20D Breakout%": data.get("20D Breakout%"),
        "Signal": data.get("Signal"),
        "Setup Signal": data.get("Setup Signal"),
        "Core Signal": data.get("Core Signal"),
        "Signal Quality": data.get("Signal Quality"),
        "Signal Regime": data.get("Signal Regime"),
        "Win Prob%": data.get("Win Prob%"),
        "Hist Precision%": data.get("Hist Precision%"),
        "Exp 5D%": data.get("Exp 5D%"),
        "Exp 10D%": data.get("Exp 10D%"),
        "WF Samples": data.get("WF Samples"),
        "Sector": data.get("Sector"),
        "Industry": data.get("Industry"),
        "Sector Benchmark": data.get("Sector Benchmark"),
        "RS Tag": data.get("RS Tag"),
        "RS vs NIFTY 1M%": data.get("RS vs NIFTY 1M%"),
        "RS vs NIFTY 3M%": data.get("RS vs NIFTY 3M%"),
        "RS vs Sector 1M%": data.get("RS vs Sector 1M%"),
        "RS vs Sector 3M%": data.get("RS vs Sector 3M%"),
        "Avg Traded Value 20D Cr": data.get("Avg Traded Value 20D Cr"),
        "Liquidity Tag": data.get("Liquidity Tag"),
        "AI Decision": data.get("AI Decision"),
        "AI Conf%": data.get("AI Conf%"),
    }
    for h in FUNDAMENTAL_SUMMARY_FIELDS:
        row_data[h] = data.get(h)
    for ci, h in enumerate(PRICE_HISTORY_HEADERS, 1):
        v = row_data.get(h)
        c = ws.cell(row=ri, column=ci, value=v)
        c.border = _THIN_BORDER; c.alignment = _ALIGN_CENTER

# -----------------------------------------------------------------------------
# SCREENER.IN FETCH
# -----------------------------------------------------------------------------
def fetch_screener(url: str) -> list:
    """
    Fetch ALL stocks from a screener page, auto-detecting total pages.

    Strategy:
      1. Fetch page=1, parse total stock count + stocks-per-page from HTML.
      2. Compute total_pages = ceil(total / per_page).
      3. Fetch pages 2..total_pages in parallel and merge, deduplicating by name.

    This prevents the silent duplication that occurs when page= exceeds the
    actual page count (screener.in clamps out-of-range page numbers to the last page).
    """
    m = re.search(r'/screens/(\d+)(/[^?]*)?', url)
    if not m: return []
    screen_id = m.group(1)
    slug = (m.group(2) or "").rstrip("/")
    has_proxy = bool(str(PROXY_URL or "").strip())

    direct_hdrs = {"User-Agent": "Mozilla/5.0"}
    proxy_hdrs  = {"User-Agent": "Mozilla/5.0", "X-Screener-Cookie": SCREENER_COOKIE}

    attempts = []
    if has_proxy:
        attempts.append((proxy_hdrs, True))
    attempts.append((direct_hdrs, False))

    def _parse_html(text):
        """Extract stocks + pagination metadata from HTML page."""
        if "0 results found" in text:
            return [], 0, 0
        stocks, seen = [], set()
        for m2 in re.finditer(
            r'href="/company/([^/"]+)/?(?:consolidated/)?"[^>]*>([^<]+)</a>', text
        ):
            raw  = _clean_sym(m2.group(1))
            name = _clean_name(m2.group(2))
            if not name or name in seen: continue
            seen.add(name)
            is_num = raw.isdigit()
            stocks.append({
                "symbol": "" if is_num else raw,
                "name": name,
                "bseCode": raw if is_num else "",
                "screener_slug": "" if is_num else raw.lower(),
            })
        # Parse "57 results found: Showing page 1 of 3"
        pi = re.search(r'(\d+) results found.*?page\s+(\d+)\s+of\s+(\d+)', text)
        total_count = int(pi.group(1)) if pi else 0
        total_pages = int(pi.group(3)) if pi else (1 if stocks else 0)
        return stocks, total_count, total_pages

    def _get_page(page_num):
        """Fetch one HTML page, return list of stocks."""
        base_url = (
            f"https://www.screener.in/screens/{screen_id}{slug}/"
            f"?page={page_num}"
        )
        for hdrs, use_proxy in attempts:
            try:
                target = _proxy_url(base_url) if use_proxy else base_url
                r = requests.get(target, headers=hdrs, timeout=25)
                if r.status_code == 200:
                    stocks, _, _ = _parse_html(r.text)
                    if stocks:
                        return stocks
            except Exception:
                pass
        return []

    # --- Step 1: fetch page 1 to discover total pages ---
    stocks_p1, total_count, total_pages = [], 0, 0
    base_p1 = f"https://www.screener.in/screens/{screen_id}{slug}/?page=1"
    for hdrs, use_proxy in attempts:
        try:
            target = _proxy_url(base_p1) if use_proxy else base_p1
            r = requests.get(target, headers=hdrs, timeout=25)
            if r.status_code == 200:
                stocks_p1, total_count, total_pages = _parse_html(r.text)
                if stocks_p1 or total_count == 0:
                    break
        except Exception:
            pass

    if not stocks_p1:
        return []

    if LOG_SCREEN_PAGE_COUNTS:
        print(f"    [SCREEN] {screen_id}: {total_count} stocks across {total_pages} page(s)")

    if total_pages <= 1:
        return stocks_p1

    # --- Step 2: fetch remaining pages in parallel ---
    all_stocks = list(stocks_p1)
    seen_names = {s["name"] for s in all_stocks}

    with ThreadPoolExecutor(max_workers=min(total_pages - 1, SCREENER_FETCH_WORKERS)) as pool:
        futures = {pool.submit(_get_page, pg): pg for pg in range(2, total_pages + 1)}
        for f in as_completed(futures):
            for s in f.result():
                if s["name"] not in seen_names:
                    seen_names.add(s["name"])
                    all_stocks.append(s)

    return all_stocks

# -----------------------------------------------------------------------------
# YAHOO FINANCE -- DIRECT CHART API (mirrors GAS tickerExists, no logging)
# -----------------------------------------------------------------------------
_YF_CHART_HOSTS = ["query2.finance.yahoo.com", "query1.finance.yahoo.com"]

def _yf_chart_api(ticker, range_="5d", interval="1d"):
    for host in _YF_CHART_HOSTS:
        try:
            url = (f"https://{host}/v8/finance/chart/"
                   f"{url_encode(ticker, safe='')}?range={range_}&interval={interval}")
            r = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=8)
            if r.status_code != 200: continue
            result = r.json().get("chart",{}).get("result",[None])[0]
            if result and result.get("meta",{}).get("symbol"): return result
        except Exception: pass
    return None

def _yf_chart_history(ticker):
    for host in _YF_CHART_HOSTS:
        try:
            url = (f"https://{host}/v8/finance/chart/"
                   f"{url_encode(ticker, safe='')}?range=3y&interval=1d")
            r = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=15)
            if r.status_code != 200: continue
            result = r.json().get("chart",{}).get("result",[None])[0]
            if not result: continue
            q = result.get("indicators",{}).get("quote",[{}])[0]
            raw_ts = result.get("timestamp") or []
            raw_c = q.get("close") or []
            raw_h = q.get("high") or []
            raw_l = q.get("low") or []
            raw_v = q.get("volume") or []

            n_raw = min(len(raw_ts), len(raw_c), len(raw_h), len(raw_l), len(raw_v))
            dt_aligned, cl_aligned, hi_aligned, lo_aligned, vo_aligned = [], [], [], [], []

            for i in range(n_raw):
                ts, c, h, l, v = raw_ts[i], raw_c[i], raw_h[i], raw_l[i], raw_v[i]
                if c is not None and h is not None and l is not None:
                    dt_aligned.append(pd.to_datetime(int(ts), unit="s"))
                    cl_aligned.append(float(c))
                    hi_aligned.append(float(h))
                    lo_aligned.append(float(l))
                    vo_aligned.append(float(v) if v is not None else 0.0)

            if len(cl_aligned) >= 5:
                return {"dates": dt_aligned, "closes": cl_aligned, "highs": hi_aligned,
                        "lows": lo_aligned, "volumes": vo_aligned}
        except Exception: pass
    return None

_yf_cache: OrderedDict = OrderedDict()

def _yf_fetch_fallback(symbol, period="3y", interval="1d", use_cache=True):
    key = f"{symbol}:{period}:{interval}"
    if use_cache and key in _yf_cache:
        _yf_cache.move_to_end(key); return _yf_cache[key]
    try:
        df = yf.Ticker(symbol).history(period=period,interval=interval,auto_adjust=True)
        if df is not None and not df.empty and len(df) >= 5:
            df = df.dropna(subset=["Close"])
            if use_cache:
                _yf_cache[key] = df; _yf_cache.move_to_end(key)
                while len(_yf_cache) > YF_CACHE_MAX: _yf_cache.popitem(last=False)
            return df
    except Exception: pass
    return None

def _hist_from_df(df):
    if df is None or len(df) < 5:
        return None
    return {
        "dates": df.index.tolist(),
        "closes": df["Close"].tolist(),
        "highs": df["High"].tolist(),
        "lows": df["Low"].tolist(),
        "volumes": df["Volume"].tolist(),
    }

def _hist_last_ts(hist):
    try:
        dates = hist.get("dates") or []
        if not dates:
            return None
        ts = pd.Timestamp(dates[-1])
        return ts.tz_localize(None) if ts.tzinfo is not None else ts
    except Exception:
        return None

def _pick_fresher_history(primary, fallback):
    if primary is None:
        return fallback
    if fallback is None:
        return primary
    p_ts = _hist_last_ts(primary)
    f_ts = _hist_last_ts(fallback)
    if p_ts is not None and f_ts is not None:
        if f_ts > p_ts:
            return fallback
        if p_ts > f_ts:
            return primary
    p_len = len(primary.get("closes") or [])
    f_len = len(fallback.get("closes") or [])
    if f_len > p_len:
        return fallback
    return primary

def _yf_exists(ticker): return _yf_chart_api(ticker, range_="5d") is not None

def _yf_search(query):
    try:
        r = requests.get("https://query2.finance.yahoo.com/v1/finance/search",
                         params={"q":query,"region":"IN","quotesCount":10},
                         headers={"User-Agent":"Mozilla/5.0"}, timeout=10)
        return r.json().get("quotes",[]) if r.status_code == 200 else []
    except Exception: return []

def _first_equity(quotes, suffix):
    hit = next((q for q in quotes
                if q.get("quoteType")=="EQUITY" and
                str(q.get("symbol","")).upper().endswith(suffix.upper())), None)
    return hit["symbol"] if hit else ""

def _ticker_for_symbol(symbol: str) -> str:
    sym = str(symbol or "").strip()
    if not sym or sym.startswith(SENTINEL):
        return ""
    return sym[4:] + ".BO" if sym.startswith("BSE:") else sym + YF_SUFFIX

def _fetch_history_for_tickers(tickers, force_refresh=False):
    for tk in tickers:
        if not tk:
            continue
        hist = _yf_chart_history(tk)
        df = _yf_fetch_fallback(tk, period="3y", interval="1d", use_cache=not force_refresh)
        hist = _pick_fresher_history(hist, _hist_from_df(df))
        if hist:
            return hist
    return None

# -----------------------------------------------------------------------------
# SYMBOL RESOLUTION
# -----------------------------------------------------------------------------
def resolve_symbol(name, bse_code="", screener_slug=""):
    ns = YF_SUFFIX; bo = ".BO"
    if bse_code and str(bse_code).strip().isdigit():
        code = str(bse_code).strip()
        if _yf_exists(code + bo):
            quotes = _yf_search(code); ns_t = _first_equity(quotes, ns)
            if ns_t: return ns_t.replace(ns,"")
            return f"BSE:{code}"
    for slug_try in [screener_slug, name.split()[0] if name else ""]:
        slug_up = slug_try.strip(".").upper() if slug_try else ""
        if slug_up and len(slug_up) > 1 and not slug_up.isdigit():
            if _yf_exists(slug_up + ns): return slug_up
            if _yf_exists(slug_up + bo): return f"BSE:{slug_up}"
    if bse_code and str(bse_code).strip().isdigit():
        code = str(bse_code).strip(); quotes = _yf_search(code)
        ns_t = _first_equity(quotes, ns)
        if ns_t: return ns_t.replace(ns,"")
        bo_t = _first_equity(quotes, bo)
        if bo_t: return f"BSE:{bo_t.replace(bo,'')}"
    words = name.split(); cleaned = _name_for_search(name)
    queries = []
    if name:    queries.append(name)
    if cleaned and cleaned != name and len(cleaned) > 2: queries.append(cleaned)
    if len(words) >= 2:
        two = " ".join(w for w in words[:2] if _valid_query_word(w))
        if two and two not in queries: queries.append(two)
    first = next((w for w in words if _valid_query_word(w)), "")
    if first and first not in queries: queries.append(first)
    queries = list(dict.fromkeys(q for q in queries if q and len(q) > 2))
    for q in queries:
        q_safe = re.sub(r'[&]','',q).strip()
        if not q_safe: continue
        for search_q in [q_safe+" NSE", q_safe]:
            quotes = _yf_search(search_q); ns_t = _first_equity(quotes, ns)
            if ns_t: return ns_t.replace(ns,"")
        for search_q in [q_safe+" BSE", q_safe]:
            quotes = _yf_search(search_q); bo_t = _first_equity(quotes, bo)
            if bo_t: return f"BSE:{bo_t.replace(bo,'')}"
    return ""

# -----------------------------------------------------------------------------
# HISTORY FETCH
# -----------------------------------------------------------------------------
def fetch_history(symbol):
    if not symbol or symbol.startswith(SENTINEL): return None
    tickers = [symbol[4:]+".BO"] if symbol.startswith("BSE:") else [symbol+YF_SUFFIX, symbol+".BO"]
    return _fetch_history_for_tickers(tickers, force_refresh=False)

def fetch_history_fresh(symbol):
    if not symbol or str(symbol).startswith(SENTINEL):
        return None
    tickers = [symbol[4:]+".BO"] if str(symbol).startswith("BSE:") else [symbol+YF_SUFFIX, symbol+".BO"]
    return _fetch_history_for_tickers(tickers, force_refresh=True)

_SECTOR_BENCHMARK_PREFERENCES = {
    "Financial Services": ["^CNXFINSERVICE", "^NSEBANK"],
    "Technology": ["^CNXIT"],
    "Healthcare": ["^CNXPHARMA"],
    "Energy": ["^CNXENERGY"],
    "Consumer Defensive": ["^CNXFMCG", "^CNXCONSUMPTION"],
    "Consumer Cyclical": ["^CNXAUTO", "^CNXCONSUMPTION", "^CNXSERVICE"],
    "Industrials": ["^CNXINFRA"],
    "Materials": ["^CNXMETAL"],
    "Real Estate": ["^CNXREALTY"],
    "Utilities": ["^CNXENERGY", "^CNXINFRA"],
    "Communication Services": ["^CNXMEDIA", "^CNXSERVICE"],
    "Services": ["^CNXSERVICE"],
}

def _normalize_sector_name(sector: str, industry: str = "") -> str:
    raw = " ".join(x for x in (sector, industry) if x).strip()
    if not raw:
        return ""
    s = re.sub(r"[^a-z0-9]+", " ", raw.lower()).strip()
    if any(tok in s for tok in ("bank", "financial", "insurance", "nbfc", "lending")):
        return "Financial Services"
    if any(tok in s for tok in ("technology", "software", "it ", " it", "semiconductor", "digital")):
        return "Technology"
    if any(tok in s for tok in ("health", "pharma", "biotech", "medical", "hospital")):
        return "Healthcare"
    if any(tok in s for tok in ("energy", "oil", "gas", "petro", "refin")):
        return "Energy"
    if any(tok in s for tok in ("utility", "utilities", "transmission", "distribution power")):
        return "Utilities"
    if any(tok in s for tok in ("real estate", "realty", "property")):
        return "Real Estate"
    if any(tok in s for tok in ("metal", "mining", "material", "cement", "chemical", "commodity")):
        return "Materials"
    if any(tok in s for tok in ("auto", "automobile", "retail", "travel", "leisure", "hotel", "apparel", "consumer cyclical")):
        return "Consumer Cyclical"
    if any(tok in s for tok in ("fmcg", "staples", "consumer defensive", "food", "beverage", "household")):
        return "Consumer Defensive"
    if any(tok in s for tok in ("industrial", "infrastructure", "engineering", "capital goods", "construction", "logistics", "shipping")):
        return "Industrials"
    if any(tok in s for tok in ("communication", "telecom", "media", "entertainment", "broadcast")):
        return "Communication Services"
    if "service" in s:
        return "Services"
    return sector.strip() or raw.title()

def _fetch_yahoo_asset_profile(ticker: str) -> dict:
    if not ticker:
        return {}
    for host in _YF_CHART_HOSTS:
        try:
            url = (
                f"https://{host}/v10/finance/quoteSummary/"
                f"{url_encode(ticker, safe='')}?modules=assetProfile"
            )
            r = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=SECTOR_PROFILE_TIMEOUT_SEC)
            if r.status_code != 200:
                continue
            result = ((r.json().get("quoteSummary") or {}).get("result") or [None])[0] or {}
            profile = result.get("assetProfile") or {}
            if profile:
                return profile
        except Exception:
            pass
    return {}

def _get_benchmark_history(ticker: str, force_refresh: bool = False):
    ticker = str(ticker or "").strip()
    if not ticker:
        return None
    if not force_refresh:
        with _benchmark_hist_lock:
            cached = _benchmark_hist_cache.get(ticker)
            if cached is not None:
                return cached
    hist = _fetch_history_for_tickers([ticker], force_refresh=force_refresh)
    if not force_refresh:
        with _benchmark_hist_lock:
            _benchmark_hist_cache[ticker] = hist
    return hist

def _resolve_sector_benchmark_ticker(sector: str) -> str:
    norm = str(sector or "").strip()
    if not norm:
        return ""
    with _symbol_meta_lock:
        if norm in _sector_benchmark_ticker_cache:
            return _sector_benchmark_ticker_cache[norm]
    ticker = ""
    for candidate in _SECTOR_BENCHMARK_PREFERENCES.get(norm, []):
        if _get_benchmark_history(candidate):
            ticker = candidate
            break
    with _symbol_meta_lock:
        _sector_benchmark_ticker_cache[norm] = ticker
    return ticker

def _get_symbol_meta(symbol: str) -> dict:
    sym = str(symbol or "").strip()
    if not sym:
        return {"Sector": "", "Industry": "", "Sector Benchmark": ""}
    with _symbol_meta_lock:
        cached = _symbol_meta_cache.get(sym)
        if cached is not None:
            return dict(cached)
    ticker = _ticker_for_symbol(sym)
    profile = _fetch_yahoo_asset_profile(ticker)
    sector_raw = str(profile.get("sector") or profile.get("sectorDisp") or "").strip()
    industry_raw = str(profile.get("industry") or profile.get("industryDisp") or "").strip()
    sector_norm = _normalize_sector_name(sector_raw, industry_raw)
    sector_benchmark = _resolve_sector_benchmark_ticker(sector_norm)
    meta = {
        "Sector": sector_norm,
        "Industry": industry_raw or sector_raw,
        "Sector Benchmark": sector_benchmark,
    }
    with _symbol_meta_lock:
        _symbol_meta_cache[sym] = dict(meta)
    return meta

def _benchmark_relative_returns(hist: dict, benchmark_ticker: str, force_refresh: bool = False) -> dict:
    out = {"RS vs NIFTY 1M%": None, "RS vs NIFTY 3M%": None, "RS vs Sector 1M%": None, "RS vs Sector 3M%": None}
    stock_1m = period_ret(hist.get("closes") or [], 21)
    stock_3m = period_ret(hist.get("closes") or [], 63)
    nifty_hist = _get_benchmark_history(BENCHMARK_TICKER, force_refresh=force_refresh)
    nifty_1m = period_ret(nifty_hist.get("closes") or [], 21) if nifty_hist else None
    nifty_3m = period_ret(nifty_hist.get("closes") or [], 63) if nifty_hist else None
    if stock_1m is not None and nifty_1m is not None:
        out["RS vs NIFTY 1M%"] = round(stock_1m - nifty_1m, 2)
    if stock_3m is not None and nifty_3m is not None:
        out["RS vs NIFTY 3M%"] = round(stock_3m - nifty_3m, 2)
    if benchmark_ticker:
        sector_hist = _get_benchmark_history(benchmark_ticker, force_refresh=force_refresh)
        sector_1m = period_ret(sector_hist.get("closes") or [], 21) if sector_hist else None
        sector_3m = period_ret(sector_hist.get("closes") or [], 63) if sector_hist else None
        if stock_1m is not None and sector_1m is not None:
            out["RS vs Sector 1M%"] = round(stock_1m - sector_1m, 2)
        if stock_3m is not None and sector_3m is not None:
            out["RS vs Sector 3M%"] = round(stock_3m - sector_3m, 2)
    return out

def _relative_strength_tag(rs_nifty_1m, rs_nifty_3m, rs_sector_1m, rs_sector_3m) -> str:
    values = [v for v in (rs_nifty_1m, rs_nifty_3m, rs_sector_1m, rs_sector_3m) if v is not None]
    if not values:
        return "—"
    score = 0
    for idx, value in enumerate((rs_nifty_1m, rs_nifty_3m, rs_sector_1m, rs_sector_3m)):
        if value is None:
            continue
        weight = 2 if idx in (1, 3) else 1
        score += weight if value >= 0 else -weight
    if score >= 5:
        return "🟢 Strong vs Both"
    if score >= 2:
        return "🟢 RS Leader"
    if score <= -5:
        return "🔴 Lagging"
    if score < 0:
        return "🟠 Weak RS"
    return "🟡 Mixed"

# -----------------------------------------------------------------------------
# TECHNICAL INDICATORS
# -----------------------------------------------------------------------------
def _rma(arr, period):
    out = np.full(len(arr), np.nan)
    if len(arr) < period: return out
    out[period-1] = arr[:period].mean(); a = 1.0/period
    for i in range(period, len(arr)):
        v = arr[i]; out[i] = a*v+(1-a)*out[i-1] if not np.isnan(v) else out[i-1]
    return out

def _clean(lst): return [c for c in lst if c is not None and not (isinstance(c,float) and math.isnan(c))]

def calc_ma(closes, period):
    v = _clean(closes)
    if len(v) < period: return None
    return round(sum(v[-period:]) / period, 2)

def calc_rsi(closes, period=14):
    cl = np.array(_clean(closes), dtype=float)
    if len(cl) < period+1: return None
    d = np.diff(cl); g, l = np.where(d>0,d,0.0), np.where(d<0,-d,0.0)
    ag, al = g[:period].mean(), l[:period].mean()
    for i in range(period, len(g)):
        ag=(ag*(period-1)+g[i])/period; al=(al*(period-1)+l[i])/period
    if ag < 1e-10 and al < 1e-10: return 50.0
    if al < 1e-10: return 100.0
    if ag < 1e-10: return 0.0
    return round(100-100/(1+ag/al), 1)

def _last_finite(arr):
    vals = np.asarray(arr, dtype=float)
    vals = vals[np.isfinite(vals)]
    return float(vals[-1]) if len(vals) else None

def calc_dmi_metrics(highs, lows, closes, period=14):
    h=np.array(highs,dtype=float); l=np.array(lows,dtype=float); c=np.array(closes,dtype=float)
    if min(len(h), len(l), len(c)) < period*2:
        return {"ADX 14": None, "+DI 14": None, "-DI 14": None, "ATR 14": None, "NATR 14": None}
    trs, pdms, mdms = [], [], []
    for i in range(1,len(c)):
        trs.append(max(h[i]-l[i],abs(h[i]-c[i-1]),abs(l[i]-c[i-1])))
        up,dn = h[i]-h[i-1],l[i-1]-l[i]
        pdms.append(up if up>dn and up>0 else 0.0)
        mdms.append(dn if dn>up and dn>0 else 0.0)
    trs=np.array(trs); pdms=np.array(pdms); mdms=np.array(mdms)
    ts=_rma(trs,period); ps=_rma(pdms,period); ms=_rma(mdms,period)
    with np.errstate(invalid="ignore",divide="ignore"):
        pdi=np.where(ts>1e-10,100*ps/ts,np.nan)
        mdi=np.where(ts>1e-10,100*ms/ts,np.nan)
        sm=pdi+mdi; dx=np.where(sm>1e-10,100*np.abs(pdi-mdi)/sm,np.nan)
    adx = _rma(dx[period-1:], period) if len(dx) >= period else np.array([])
    adx_val = _last_finite(adx)
    atr_val = _last_finite(ts)
    cur = float(c[-1]) if len(c) else None
    natr_val = (atr_val / cur * 100.0) if atr_val is not None and cur is not None and abs(cur) > 1e-10 else None
    pdi_val = _last_finite(pdi)
    mdi_val = _last_finite(mdi)
    return {
        "ADX 14": round(adx_val, 1) if adx_val is not None else None,
        "+DI 14": round(pdi_val, 1) if pdi_val is not None else None,
        "-DI 14": round(mdi_val, 1) if mdi_val is not None else None,
        "ATR 14": round(atr_val, 2) if atr_val is not None else None,
        "NATR 14": round(natr_val, 2) if natr_val is not None else None,
    }

def calc_adx(highs, lows, closes, period=14):
    return calc_dmi_metrics(highs, lows, closes, period).get("ADX 14")

def calc_macd(closes, fast=12, slow=26, sig=9):
    cl=pd.Series(_clean(closes),dtype=float)
    if len(cl)<slow+sig: return None, None
    m=(cl.ewm(span=fast,adjust=False,min_periods=fast).mean()
      -cl.ewm(span=slow,adjust=False,min_periods=slow).mean())
    s=m.ewm(span=sig,adjust=False,min_periods=sig).mean()
    ml=m.dropna(); hl=(m-s).dropna()
    return (round(float(ml.iloc[-1]),2) if len(ml) else None,
            round(float(hl.iloc[-1]),2) if len(hl) else None)

def calc_vol_ratio(volumes, period=20):
    v=[]
    for x in volumes:
        if x is None:
            continue
        try:
            fx = float(x)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(fx):
            continue
        v.append(max(fx, 0.0))
    if len(v)<period+1: return None
    avg=sum(v[-period-1:-1])/period
    return round(v[-1]/avg,2) if avg>1e-10 else None

def calc_avg_traded_value_crore(closes, volumes, period=20):
    cl = _clean(closes)
    vals = []
    n = min(len(closes), len(volumes))
    if n < period:
        return None
    for c, v in zip(closes[-period:], volumes[-period:]):
        try:
            fc = float(c)
            fv = max(float(v), 0.0)
        except (TypeError, ValueError):
            continue
        if not (math.isfinite(fc) and math.isfinite(fv)):
            continue
        vals.append(fc * fv)
    if len(vals) < max(5, period // 2):
        return None
    return round((sum(vals) / len(vals)) / 1e7, 2)

def liquidity_tag(avg_traded_value_cr):
    v = _as_float(avg_traded_value_cr)
    if v is None:
        return "—"
    if v >= 100:
        return "🟢 Deep"
    if v >= 20:
        return "🟢 Liquid"
    if v >= 5:
        return "🟡 Adequate"
    if v >= 1:
        return "🟠 Thin"
    return "🔴 Illiquid"

def calc_dist_52w(highs, closes):
    if len(highs) < 252 or not closes: return None
    v=_clean(highs[-252:])
    if len(v) < 252: return None
    h=max(v); cur=closes[-1]
    return round((h-cur)/h*100,2) if h>1e-10 else None

def calc_breakout_20d(highs, closes):
    if len(highs)<21 or not closes: return None
    ph=_clean(highs[-21:-1])
    if len(ph) < 20: return None
    prev=max(ph); cur=closes[-1]
    return round((cur-prev)/prev*100,2) if prev>1e-10 else None

def period_ret(closes, periods):
    v=_clean(closes)
    if len(v)<=periods: return None
    past=v[-(periods+1)]; cur=v[-1]
    return round((cur-past)/past*100,2) if past and abs(past)>1e-10 else None

def avg_period_ret(closes, period, lookback):
    v=_clean(closes[-lookback:]); n=len(v)
    if n<period*2: return None
    rets=[]
    for i in range(period,n,period):
        s,e=v[i-period],v[i]
        if s and abs(s)>1e-10: rets.append((e-s)/s*100)
    return round(sum(rets)/len(rets),2) if rets else None

def calc_bollinger(closes, period=20, num_std=2, rank_lookback=126):
    """Compute Bollinger context.
    Returns current %B, current width, width percentile rank, and squeeze flag.
    """
    v = _clean(closes)
    if len(v) < period:
        return None, None, None, None

    widths = []
    bb_pctb = None
    bb_width = None
    for end in range(period, len(v) + 1):
        window = v[end - period:end]
        ma = sum(window) / period
        if ma < 1e-10:
            continue
        std = (sum((x - ma) ** 2 for x in window) / period) ** 0.5
        upper = ma + num_std * std
        lower = ma - num_std * std
        width = (upper - lower) / ma * 100
        widths.append(width)
        if end == len(v):
            cur = v[-1]
            bb_pctb = round((cur - lower) / (upper - lower), 3) if (upper - lower) > 1e-10 else None
            bb_width = round(width, 2)

    valid_widths = [w for w in widths[-rank_lookback:] if math.isfinite(w)]
    bb_width_pctl = None
    bb_squeeze = None
    if bb_width is not None and len(valid_widths) >= 20:
        bb_width_pctl = round(sum(1 for w in valid_widths if w <= bb_width) / len(valid_widths) * 100, 1)
        bb_squeeze = bb_width_pctl <= 15.0
    return bb_pctb, bb_width, bb_width_pctl, bb_squeeze

def calc_camarilla(highs, lows, closes, multiplier=1.1):
    """Compute daily Camarilla H3/H4/L3/L4 from the previous completed bar."""
    n = min(len(highs), len(lows), len(closes))
    if n < 2:
        return {"Cam H3": None, "Cam H4": None, "Cam L3": None, "Cam L4": None}

    try:
        ref_h = float(highs[n-2])
        ref_l = float(lows[n-2])
        ref_c = float(closes[n-2])
    except (TypeError, ValueError):
        return {"Cam H3": None, "Cam H4": None, "Cam L3": None, "Cam L4": None}

    if not all(math.isfinite(x) for x in (ref_h, ref_l, ref_c)):
        return {"Cam H3": None, "Cam H4": None, "Cam L3": None, "Cam L4": None}

    rng = ref_h - ref_l
    if rng <= 1e-10:
        return {"Cam H3": None, "Cam H4": None, "Cam L3": None, "Cam L4": None}

    step = rng * multiplier / 4.0
    return {
        "Cam H3": round(ref_c + step, 2),
        "Cam H4": round(ref_c + step * 2.0, 2),
        "Cam L3": round(ref_c - step, 2),
        "Cam L4": round(ref_c - step * 2.0, 2),
    }

# -----------------------------------------------------------------------------
# GAS v4.0 SIGNAL ENGINE
# -----------------------------------------------------------------------------
def calc_signal(price,ma20,ma50,ma200,rsi,adx,vol_ratio,macd_line,macd_hist,dist_52w,breakout_20d):
    cfg=SIGNAL_PROFILES.get(SIGNAL_PROFILE,SIGNAL_PROFILES["balanced"])
    precision_mode = bool(cfg.get("HIGH_PRECISION"))
    if price is None: return "No Data"
    a20=ma20 is not None and price>ma20; a50=ma50 is not None and price>ma50
    hm2=ma200 is not None; a200=hm2 and price>ma200
    hr=rsi is not None
    ob=hr and rsi>=cfg["RSI_OB"]; os_=hr and rsi<=cfg["RSI_OVERSOLD"]
    rb=hr and cfg["RSI_BULL_MIN"]<=rsi<cfg["RSI_OB"]
    rn=hr and cfg["RSI_NEU_MIN"]<=rsi<cfg["RSI_BULL_MIN"]
    ts=adx is not None and adx>=cfg["ADX_STRONG"]
    tm=adx is not None and adx>=cfg["ADX_WEAK"]
    hv=vol_ratio is not None and vol_ratio>=cfg["VOL_HIGH"]
    mb=macd_line is not None and macd_hist is not None and macd_line>0 and macd_hist>0
    md=macd_line is not None and macd_hist is not None and macd_line<0 and macd_hist<0
    n52=dist_52w is not None and dist_52w<=cfg["DIST_52W_MAX"]
    bk=breakout_20d is not None and breakout_20d>=cfg["BREAKOUT_MIN"]
    pb=hm2 and a20 and a50 and a200
    sell_rsi_max = cfg.get("SELL_RSI_MAX", cfg["RSI_NEU_MIN"])
    sell_breakdown_min = cfg.get("SELL_BREAKDOWN_MIN", cfg["BREAKOUT_MIN"])
    if precision_mode:
        if pb and n52 and bk and ts and hv and mb:                           return "🚀 BREAKOUT"
        if pb and os_ and ts and hv and mb:                                  return "[OK] STRONG BUY (Oversold)"
        if pb and rb and ts and hv and mb and bk:                            return "[OK] STRONG BUY"
        if pb and rn and ts and hv and mb and n52 and bk:                   return "[OK] BUY"
        if a20 and a50 and ob:                                               return "[WARN] HOLD (Overbought)"
        if a20 and a50 and (not hm2 or a200):                                return "[WARN] HOLD"
        if a20 and a50 and hm2 and not a200:                                 return "[WARN] HOLD (Below MA200)"
        if a50 and not a20 and mb:                                           return "[WARN] PULLBACK"
        if not a20 and a50 and os_:                                          return "👀 OVERSOLD"
        if a20 and not a50:                                                  return "[+] WEAK"
        if not a20 and not a50 and os_:                                      return "👀 OVERSOLD (Watch)"
        if not a20 and not a50 and hm2 and not a200 and ts and md and (rsi is None or rsi <= sell_rsi_max) and (breakout_20d is None or breakout_20d <= -sell_breakdown_min): return "[WARN] HOLD (Below MA200)"
        if not a20 and not a50 and hm2 and not a200:                         return "[WARN] HOLD (Below MA200)"
        return "[WARN] HOLD"
    if pb and n52 and bk and ts and hv:                           return "🚀 BREAKOUT"
    if pb and os_:                                                return "[OK] STRONG BUY (Oversold)"
    if pb and rb and ts:                                          return "[OK] STRONG BUY"
    if a20 and a50 and (not hm2 or a200) and mb and (rb or rn) and tm: return "[OK] BUY"
    if a20 and a50 and ob:                                        return "[WARN] HOLD (Overbought)"
    if a20 and a50 and (not hm2 or a200):                        return "[WARN] HOLD"
    if a20 and a50 and hm2 and not a200:                         return "[WARN] HOLD (Below MA200)"
    if a50 and not a20 and mb:                                    return "[WARN] PULLBACK"
    if not a20 and a50 and os_:                                   return "👀 OVERSOLD"
    if a20 and not a50:                                           return "[+] WEAK"
    if not a20 and not a50 and os_:                               return "👀 OVERSOLD (Watch)"
    if not a20 and not a50:                                       return "[FAIL] SELL"
    return "[WARN] HOLD"

def calc_signal_enhanced(price,ma20,ma50,ma200,rsi,adx,vol_ratio,macd_line,macd_hist,dist_52w,breakout_20d,
                         pdi=None,mdi=None,atr=None,natr=None,bb_pctb=None,bb_width_pctl=None,bb_squeeze=None):
    cfg=SIGNAL_PROFILES.get(SIGNAL_PROFILE,SIGNAL_PROFILES["balanced"])
    precision_mode = bool(cfg.get("HIGH_PRECISION"))
    if price is None: return "No Data"
    a20=ma20 is not None and price>ma20; a50=ma50 is not None and price>ma50
    hm2=ma200 is not None; a200=hm2 and price>ma200
    hr=rsi is not None
    ob=hr and rsi>=cfg["RSI_OB"]; os_=hr and rsi<=cfg["RSI_OVERSOLD"]
    rb=hr and cfg["RSI_BULL_MIN"]<=rsi<cfg["RSI_OB"]
    rn=hr and cfg["RSI_NEU_MIN"]<=rsi<cfg["RSI_BULL_MIN"]
    ts=adx is not None and adx>=cfg["ADX_STRONG"]
    tm=adx is not None and adx>=cfg["ADX_WEAK"]
    hv=vol_ratio is not None and vol_ratio>=cfg["VOL_HIGH"]
    mb=macd_line is not None and macd_hist is not None and macd_line>0 and macd_hist>0
    md=macd_line is not None and macd_hist is not None and macd_line<0 and macd_hist<0
    n52=dist_52w is not None and dist_52w<=cfg["DIST_52W_MAX"]
    bk=breakout_20d is not None and breakout_20d>=cfg["BREAKOUT_MIN"]
    pb=hm2 and a20 and a50 and a200

    di_ready = pdi is not None and mdi is not None
    di_gap = (pdi - mdi) if di_ready else None
    di_bull = di_ready and pdi > mdi
    di_bear = di_ready and mdi > pdi
    di_breakout = di_gap is not None and di_gap >= cfg["DI_BULL_GAP"]
    di_sell_gap = di_gap is not None and (-di_gap) >= cfg.get("DI_SELL_GAP", cfg["DI_BULL_GAP"])
    di_buy_ok = (not di_ready) or di_bull
    di_breakout_ok = (not di_ready) or di_breakout

    natr_ok = natr is None or natr <= cfg["NATR_MAX"]
    natr_hot = natr is not None and natr >= cfg["NATR_HOT"]
    bb_push_min = cfg.get("BB_PUSH_MIN", 0.8)
    bb_stretched_min = cfg.get("BB_STRETCHED_MIN", 1.0)
    sell_rsi_max = cfg.get("SELL_RSI_MAX", cfg["RSI_NEU_MIN"])
    sell_bb_max = cfg.get("SELL_BB_MAX", 0.4)
    sell_breakdown_min = cfg.get("SELL_BREAKDOWN_MIN", cfg["BREAKOUT_MIN"])
    squeeze_ready = (bb_squeeze is True) or (bb_width_pctl is not None and bb_width_pctl <= cfg["BB_SQUEEZE_PCTL"])
    bb_push = bb_pctb is not None and bb_pctb >= bb_push_min
    bb_stretched = bb_pctb is not None and bb_pctb > bb_stretched_min

    if precision_mode:
        if pb and n52 and bk and ts and hv and mb and di_breakout_ok and natr_ok and squeeze_ready and bb_push:
            return "🚀 BREAKOUT"
        if pb and os_ and ts and hv and mb and di_breakout_ok and bb_push and not natr_hot:
            return "[OK] STRONG BUY (Oversold)"
        if pb and rb and ts and hv and mb and di_breakout_ok and natr_ok and (bk or bb_push):
            return "[OK] STRONG BUY"
        if pb and rn and ts and hv and mb and di_breakout_ok and natr_ok and squeeze_ready and bb_push:
            if squeeze_ready and (bk or bb_push):
                return "[OK] BUY (Squeeze)"
            return "[OK] BUY"
        if a20 and a50 and ob and (di_bear or bb_stretched):
            return "[WARN] HOLD (Overbought)"
        if a20 and a50 and di_bear and tm:
            return "[WARN] HOLD (DI Weakness)"
        if a20 and a50 and (not hm2 or a200):                         return "[WARN] HOLD"
        if a20 and a50 and hm2 and not a200:                          return "[WARN] HOLD (Below MA200)"
        if a50 and not a20 and mb and di_buy_ok and not natr_hot:     return "[WARN] PULLBACK"
        if not a20 and a50 and os_:                                   return "👀 OVERSOLD"
        if a20 and not a50 and di_bull and not natr_hot:              return "[+] WEAK"
        if not a20 and not a50 and os_:                               return "👀 OVERSOLD (Watch)"
        if not a20 and not a50 and hm2 and not a200 and di_bear and di_sell_gap and ts and md and (rsi is None or rsi <= sell_rsi_max) and (bb_pctb is None or bb_pctb <= sell_bb_max) and (breakout_20d is None or breakout_20d <= -sell_breakdown_min):
            return "[WARN] HOLD (Below MA200)"
        if not a20 and not a50 and hm2 and not a200:
            return "[WARN] HOLD (Below MA200)"
        return "[WARN] HOLD"

    if pb and n52 and bk and ts and hv and di_breakout_ok and natr_ok and (squeeze_ready or bb_push):
        return "🚀 BREAKOUT"
    if pb and os_ and di_buy_ok and not natr_hot:
        return "[OK] STRONG BUY (Oversold)"
    if pb and rb and ts and di_buy_ok and not natr_hot:
        return "[OK] STRONG BUY"
    if a20 and a50 and (not hm2 or a200) and mb and (rb or rn) and tm and di_buy_ok and not natr_hot:
        if squeeze_ready and bk:
            return "[OK] BUY (Squeeze)"
        return "[OK] BUY"
    if a20 and a50 and ob and (di_bear or bb_stretched):
        return "[WARN] HOLD (Overbought)"
    if a20 and a50 and di_bear and tm:
        return "[WARN] HOLD (DI Weakness)"
    if a20 and a50 and (not hm2 or a200):                         return "[WARN] HOLD"
    if a20 and a50 and hm2 and not a200:                          return "[WARN] HOLD (Below MA200)"
    if a50 and not a20 and mb and di_buy_ok and not natr_hot:     return "[WARN] PULLBACK"
    if not a20 and a50 and os_:                                   return "👀 OVERSOLD"
    if a20 and not a50 and di_bull and not natr_hot:              return "[+] WEAK"
    if not a20 and not a50 and os_:                               return "👀 OVERSOLD (Watch)"
    if not a20 and not a50 and di_bear and tm:                    return "[FAIL] SELL"
    if not a20 and not a50:                                       return "[FAIL] SELL"
    return "[WARN] HOLD"

_VALID_SIGNAL_ENGINES = {"legacy", "enhanced"}

def _signal_engine_name() -> str:
    engine = str(SIGNAL_ENGINE or "enhanced").strip().lower()
    return engine if engine in _VALID_SIGNAL_ENGINES else "enhanced"

def _core_signal_from_metrics(metrics: dict):
    return calc_signal(
        metrics.get("Current Price"),
        metrics.get("MA 20"), metrics.get("MA 50"), metrics.get("MA 200"),
        metrics.get("RSI 14"), metrics.get("ADX 14"), metrics.get("Vol Ratio 20"),
        metrics.get("MACD Line"), metrics.get("MACD Hist"),
        metrics.get("52W High Dist%"), metrics.get("20D Breakout%"),
    )

def _enhanced_signal_from_metrics(metrics: dict):
    return calc_signal_enhanced(
        metrics.get("Current Price"),
        metrics.get("MA 20"), metrics.get("MA 50"), metrics.get("MA 200"),
        metrics.get("RSI 14"), metrics.get("ADX 14"), metrics.get("Vol Ratio 20"),
        metrics.get("MACD Line"), metrics.get("MACD Hist"),
        metrics.get("52W High Dist%"),
        metrics.get("20D Breakout%"),
        pdi=metrics.get("+DI 14"),
        mdi=metrics.get("-DI 14"),
        atr=metrics.get("ATR 14"),
        natr=metrics.get("NATR 14"),
        bb_pctb=metrics.get("BB %B"),
        bb_width_pctl=metrics.get("BB Width Pctl"),
        bb_squeeze=metrics.get("BB Squeeze"),
    )

def _setup_signal_from_metrics(metrics: dict):
    if _signal_engine_name() == "legacy":
        return _core_signal_from_metrics(metrics)
    return _enhanced_signal_from_metrics(metrics)

def _normalize_signal_label(signal: str) -> str:
    text = str(signal or "").strip()
    if not text:
        return ""
    text = re.sub(r'^\[[^\]]+\]\s*', '', text)
    text = re.sub(r'^[^A-Za-z0-9]+', '', text)
    text = " ".join(text.split())
    return text.upper()

def _signal_family_key(signal: str) -> str:
    norm = _normalize_signal_label(signal)
    if norm.startswith("BUY (SQUEEZE"):
        return "BUY"
    if norm.startswith("STRONG BUY (OVERSOLD"):
        return "STRONG BUY"
    if norm.startswith("HOLD"):
        return "HOLD"
    if norm.startswith("OVERSOLD"):
        return "OVERSOLD"
    return norm

def _is_bullish_setup_signal(signal: str) -> bool:
    return _signal_family_key(signal) in {"BREAKOUT", "STRONG BUY", "BUY", "PULLBACK", "OVERSOLD"}

def _classify_signal_regime(metrics: dict) -> str:
    price = _as_float(metrics.get("Current Price"))
    ma20 = _as_float(metrics.get("MA 20"))
    ma50 = _as_float(metrics.get("MA 50"))
    ma200 = _as_float(metrics.get("MA 200"))
    adx = _as_float(metrics.get("ADX 14"))
    natr = _as_float(metrics.get("NATR 14"))
    pdi = _as_float(metrics.get("+DI 14"))
    mdi = _as_float(metrics.get("-DI 14"))
    cfg = SIGNAL_PROFILES.get(SIGNAL_PROFILE, SIGNAL_PROFILES["balanced"])

    if natr is not None and natr >= cfg.get("NATR_HOT", 10.0):
        return "HIGH-VOL"

    trend_stack = (
        price is not None and
        (ma20 is None or price > ma20) and
        (ma50 is None or price > ma50) and
        (ma200 is None or price > ma200)
    )
    di_gap = abs(pdi - mdi) if pdi is not None and mdi is not None else None
    if adx is not None and adx >= cfg.get("ADX_STRONG", 20) and trend_stack and (
        di_gap is None or di_gap >= cfg.get("DI_BULL_GAP", 4)
    ):
        return "TRENDING"
    return "CHOPPY"

def _walkforward_signature(hist: dict) -> str:
    dates = hist.get("dates") or []
    closes = hist.get("closes") or []
    highs = hist.get("highs") or []
    lows = hist.get("lows") or []
    last_date = str(dates[-1])[:19] if dates else ""
    last_close = rnd(closes[-1], 4) if closes else None
    last_high = rnd(highs[-1], 4) if highs else None
    last_low = rnd(lows[-1], 4) if lows else None
    return json.dumps(
        {
            "version": WALKFORWARD_CACHE_VERSION,
            "profile": SIGNAL_PROFILE,
            "engine": _signal_engine_name(),
            "bars": len(closes),
            "last_date": last_date,
            "last_close": last_close,
            "last_high": last_high,
            "last_low": last_low,
        },
        sort_keys=True,
        separators=(",", ":"),
    )

def _load_walkforward_cache():
    global _walkforward_cache
    if _walkforward_cache is not None:
        return _walkforward_cache
    cache = {"version": WALKFORWARD_CACHE_VERSION, "symbols": {}}
    try:
        if os.path.exists(WALKFORWARD_CACHE_FILE):
            with open(WALKFORWARD_CACHE_FILE, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict) and isinstance(data.get("symbols"), dict):
                cache = data
    except Exception as e:
        print(f"  [WARN] Walk-forward cache load failed: {e}")
    _walkforward_cache = cache
    return _walkforward_cache

def _save_walkforward_cache():
    global _walkforward_cache_dirty
    if not _walkforward_cache_dirty:
        return
    with _walkforward_cache_lock:
        cache = _load_walkforward_cache()
        tmp = WALKFORWARD_CACHE_FILE + ".tmp"
        try:
            with open(tmp, "w", encoding="utf-8") as fh:
                json.dump(cache, fh, ensure_ascii=True, separators=(",", ":"))
            os.replace(tmp, WALKFORWARD_CACHE_FILE)
            _walkforward_cache_dirty = False
        except Exception as e:
            print(f"  [WARN] Walk-forward cache save failed: {e}")
            try:
                if os.path.exists(tmp):
                    os.remove(tmp)
            except Exception:
                pass

def _base_signal_metrics_from_hist(hist: dict) -> dict:
    cl = hist["closes"]; hi = hist["highs"]; lo = hist["lows"]; vo = hist["volumes"]
    dmi = calc_dmi_metrics(hi, lo, cl, 14)
    ml, mh = calc_macd(cl)
    bb_pctb, bb_width, bb_width_pctl, bb_squeeze = calc_bollinger(cl, 20, 2)
    return {
        "Current Price": rnd(cl[-1] if cl else None),
        "RSI 14": calc_rsi(cl, 14),
        "MA 20": calc_ma(cl, 20), "MA 50": calc_ma(cl, 50), "MA 200": calc_ma(cl, 200),
        "ADX 14": dmi["ADX 14"],
        "Vol Ratio 20": calc_vol_ratio(vo, 20),
        "MACD Line": ml, "MACD Hist": mh,
        "52W High Dist%": calc_dist_52w(hi, cl),
        "20D Breakout%": calc_breakout_20d(hi, cl),
        "ATR 14": dmi["ATR 14"], "NATR 14": dmi["NATR 14"],
        "+DI 14": dmi["+DI 14"], "-DI 14": dmi["-DI 14"],
        "BB %B": bb_pctb, "BB Width": bb_width,
        "BB Width Pctl": bb_width_pctl, "BB Squeeze": bb_squeeze,
    }

def _new_walkforward_bucket():
    return {"samples": 0, "wins5": 0, "wins10": 0, "sum5": 0.0, "sum10": 0.0}

def _record_walkforward_outcome(bucket: dict, ret5, ret10):
    if ret5 is None or ret10 is None:
        return
    bucket["samples"] += 1
    bucket["wins5"] += 1 if ret5 > 0 else 0
    bucket["wins10"] += 1 if ret10 > 0 else 0
    bucket["sum5"] += float(ret5)
    bucket["sum10"] += float(ret10)

def _finalize_walkforward_bucket(bucket: dict) -> dict:
    samples = int(bucket.get("samples") or 0)
    if samples <= 0:
        return {
            "WF Samples": 0,
            "Win Prob%": None,
            "Hist Precision%": None,
            "Exp 5D%": None,
            "Exp 10D%": None,
        }
    win5 = bucket["wins5"] / samples * 100.0
    win10 = bucket["wins10"] / samples * 100.0
    return {
        "WF Samples": samples,
        "Win Prob%": round(win5 * 0.4 + win10 * 0.6, 1),
        "Hist Precision%": round(win10, 1),
        "Exp 5D%": round(bucket["sum5"] / samples, 2),
        "Exp 10D%": round(bucket["sum10"] / samples, 2),
    }

def _compute_walkforward_signal_stats(hist: dict) -> dict:
    closes = hist.get("closes") or []
    highs = hist.get("highs") or []
    lows = hist.get("lows") or []
    volumes = hist.get("volumes") or []
    n = min(len(closes), len(highs), len(lows), len(volumes))
    if n < 280:
        return {"exact": {}, "family": {}}

    exact_buckets = {}
    family_buckets = {}
    max_forward = 10
    start_idx = max(252, n - WALKFORWARD_EVAL_BARS - max_forward)

    for end_idx in range(start_idx, n - max_forward):
        slice_from = max(0, end_idx + 1 - WALKFORWARD_LOOKBACK_BARS)
        wf_hist = {
            "closes": closes[slice_from:end_idx + 1],
            "highs": highs[slice_from:end_idx + 1],
            "lows": lows[slice_from:end_idx + 1],
            "volumes": volumes[slice_from:end_idx + 1],
        }
        wf_metrics = _base_signal_metrics_from_hist(wf_hist)
        setup_signal = _setup_signal_from_metrics(wf_metrics)
        exact_key = _normalize_signal_label(setup_signal)
        if not exact_key or exact_key in ("NO DATA", "ERROR", "PENDING...", "SYMBOL NOT FOUND"):
            continue
        family_key = _signal_family_key(exact_key)
        cur_price = closes[end_idx]
        if cur_price is None or abs(float(cur_price)) <= 1e-10:
            continue
        ret5 = round((float(closes[end_idx + 5]) - float(cur_price)) / float(cur_price) * 100.0, 2)
        ret10 = round((float(closes[end_idx + 10]) - float(cur_price)) / float(cur_price) * 100.0, 2)
        _record_walkforward_outcome(exact_buckets.setdefault(exact_key, _new_walkforward_bucket()), ret5, ret10)
        _record_walkforward_outcome(family_buckets.setdefault(family_key, _new_walkforward_bucket()), ret5, ret10)

    return {
        "exact": {k: _finalize_walkforward_bucket(v) for k, v in exact_buckets.items()},
        "family": {k: _finalize_walkforward_bucket(v) for k, v in family_buckets.items()},
    }

def _get_walkforward_signal_stats(symbol: str, hist: dict) -> dict:
    global _walkforward_cache_dirty
    if not symbol:
        return {"exact": {}, "family": {}}
    signature = _walkforward_signature(hist)
    with _walkforward_cache_lock:
        cache = _load_walkforward_cache()
        entry = cache.setdefault("symbols", {}).get(symbol)
        if isinstance(entry, dict) and entry.get("signature") == signature and isinstance(entry.get("stats"), dict):
            return entry["stats"]
    stats = _compute_walkforward_signal_stats(hist)
    with _walkforward_cache_lock:
        cache = _load_walkforward_cache()
        cache.setdefault("symbols", {})[symbol] = {"signature": signature, "stats": stats}
        _walkforward_cache_dirty = True
    return stats

def _pick_walkforward_stats(stats: dict, setup_signal: str) -> dict:
    empty = {
        "WF Samples": 0,
        "Win Prob%": None,
        "Hist Precision%": None,
        "Exp 5D%": None,
        "Exp 10D%": None,
    }
    if not isinstance(stats, dict):
        return empty
    exact_key = _normalize_signal_label(setup_signal)
    family_key = _signal_family_key(setup_signal)
    exact_stats = (stats.get("exact") or {}).get(exact_key) or {}
    family_stats = (stats.get("family") or {}).get(family_key) or {}
    if int(exact_stats.get("WF Samples") or 0) >= WALKFORWARD_MIN_SAMPLES:
        return dict(exact_stats)
    if int(family_stats.get("WF Samples") or 0) > int(exact_stats.get("WF Samples") or 0):
        return dict(family_stats)
    if exact_stats:
        return dict(exact_stats)
    if family_stats:
        return dict(family_stats)
    return empty

def _quality_score_from_metrics(metrics: dict) -> float | None:
    win_prob = _as_float(metrics.get("Win Prob%"))
    hist_precision = _as_float(metrics.get("Hist Precision%"))
    exp10 = _as_float(metrics.get("Exp 10D%"))
    samples = _as_int(metrics.get("WF Samples")) or 0
    regime = str(metrics.get("Signal Regime") or "")
    if all(v is None for v in (win_prob, hist_precision, exp10)) and samples <= 0:
        return None
    score = 50.0
    if win_prob is not None:
        score += (win_prob - 50.0) * 1.2
    if hist_precision is not None:
        score += (hist_precision - 50.0) * 0.8
    if exp10 is not None:
        score += max(-12.0, min(12.0, exp10 * 4.0))
    score += min(8.0, samples * 0.8)
    if regime == "TRENDING":
        score += 8.0
    elif regime == "CHOPPY":
        score -= 8.0
    elif regime == "HIGH-VOL":
        score -= 12.0
    return round(max(0.0, min(100.0, score)), 1)

def _rejected_signal_label(reason: str) -> str:
    if reason == "HIGH-VOL":
        return "[WARN] HOLD (High Vol)"
    if reason == "CHOPPY":
        return "[WARN] HOLD (Choppy Regime)"
    if reason == "THIN HISTORY":
        return "[WARN] HOLD (Thin History)"
    return "[WARN] HOLD (Low Quality)"

def _quality_gate_from_metrics(metrics: dict) -> dict:
    setup_signal = str(metrics.get("Setup Signal") or "").strip()
    core_signal = str(metrics.get("Core Signal") or "").strip()
    if not setup_signal:
        live = core_signal or "No Data"
        return {"Signal": live, "Signal Quality": "UNAVAILABLE"}

    regime = str(metrics.get("Signal Regime") or "")
    samples = _as_int(metrics.get("WF Samples")) or 0
    win_prob = _as_float(metrics.get("Win Prob%"))
    hist_precision = _as_float(metrics.get("Hist Precision%"))
    exp10 = _as_float(metrics.get("Exp 10D%"))
    quality_score = _quality_score_from_metrics(metrics)

    if all(v is None for v in (win_prob, hist_precision, exp10)) and samples <= 0:
        return {
            "Signal": setup_signal,
            "Signal Quality": "PASS - UNVERIFIED",
            "_quality_score": quality_score,
        }

    if not _is_bullish_setup_signal(setup_signal):
        return {
            "Signal": setup_signal,
            "Signal Quality": "N/A - NON-BULL",
            "_quality_score": quality_score,
        }

    reject_reason = None
    if samples < WALKFORWARD_MIN_SAMPLES:
        if regime != "TRENDING":
            reject_reason = "THIN HISTORY"
    elif regime == "HIGH-VOL" and (win_prob is None or win_prob < QUALITY_GATE_MIN_WIN_PROB + 3):
        reject_reason = "HIGH-VOL"
    elif regime == "CHOPPY" and (win_prob is None or win_prob < QUALITY_GATE_MIN_WIN_PROB + 2):
        reject_reason = "CHOPPY"
    elif win_prob is not None and win_prob < QUALITY_GATE_MIN_WIN_PROB:
        reject_reason = "LOW EDGE"
    elif hist_precision is not None and hist_precision < QUALITY_GATE_MIN_HIST_PRECISION:
        reject_reason = "LOW EDGE"
    elif exp10 is not None and exp10 < QUALITY_GATE_MIN_EXP_10D:
        reject_reason = "LOW EDGE"
    elif quality_score is not None and quality_score < QUALITY_GATE_MIN_SCORE:
        reject_reason = "LOW EDGE"

    if reject_reason:
        quality_label = f"REJECT - {reject_reason}"
        return {
            "Signal": _rejected_signal_label(reject_reason),
            "Signal Quality": quality_label,
            "_quality_score": quality_score,
        }

    if quality_score is None:
        quality_label = "PASS - UNVERIFIED"
    elif quality_score >= 75:
        quality_label = "PASS - HIGH"
    elif quality_score >= 60:
        quality_label = "PASS - MED"
    else:
        quality_label = "PASS - LOW"
    return {
        "Signal": setup_signal,
        "Signal Quality": quality_label,
        "_quality_score": quality_score,
    }

def _signal_bundle_from_metrics(metrics: dict) -> dict:
    core_signal = _core_signal_from_metrics(metrics)
    setup_signal = _setup_signal_from_metrics(metrics)
    if not setup_signal:
        setup_signal = core_signal or "No Data"
    bundle = {
        "Setup Signal": setup_signal,
        "Core Signal": core_signal or setup_signal,
    }
    bundle.update(_quality_gate_from_metrics({**metrics, **bundle}))
    if not bundle.get("Signal"):
        bundle["Signal"] = setup_signal
    if not bundle.get("Signal Quality"):
        bundle["Signal Quality"] = "UNAVAILABLE"
    return bundle

def compute_metrics(hist, symbol: str = "", include_walkforward: bool = True, benchmark_refresh: bool = False):
    cl = hist["closes"]; hi = hist["highs"]; lo = hist["lows"]; vo = hist["volumes"]
    n = len(cl); cur = cl[-1] if cl else None
    base_metrics = _base_signal_metrics_from_hist(hist)
    cam = calc_camarilla(hi, lo, cl)
    avg_traded_value_cr = calc_avg_traded_value_crore(cl, vo, period=20)
    m = {
        **base_metrics,
        "Current Price": rnd(cur),
        "1D%":period_ret(cl,1),    "1W%":period_ret(cl,5),
        "1M%":period_ret(cl,21),   "3M%":period_ret(cl,63),
        "6M%":period_ret(cl,126),  "1Y%":period_ret(cl,252),
        "2Y%":period_ret(cl,504),  "3Y%":period_ret(cl,756),
        "Avg Weekly%":avg_period_ret(cl,5,252),
        "Avg Monthly%":avg_period_ret(cl,21,504),
        "Avg 3M%":avg_period_ret(cl,63,n),
        "Avg 6M%":avg_period_ret(cl,126,n),
        "Avg 1Y%":avg_period_ret(cl,252,n),
        "Cam H3": cam.get("Cam H3"), "Cam H4": cam.get("Cam H4"),
        "Cam L3": cam.get("Cam L3"), "Cam L4": cam.get("Cam L4"),
        "Sector": "",
        "Industry": "",
        "Sector Benchmark": "",
        "RS Tag": "—",
        "RS vs NIFTY 1M%": None,
        "RS vs NIFTY 3M%": None,
        "RS vs Sector 1M%": None,
        "RS vs Sector 3M%": None,
        "Avg Traded Value 20D Cr": avg_traded_value_cr,
        "Liquidity Tag": liquidity_tag(avg_traded_value_cr),
    }
    if symbol:
        meta = _get_symbol_meta(symbol)
        m.update(meta)
        rs_metrics = _benchmark_relative_returns(
            hist,
            meta.get("Sector Benchmark", ""),
            force_refresh=benchmark_refresh,
        )
        m.update(rs_metrics)
        m["RS Tag"] = _relative_strength_tag(
            rs_metrics.get("RS vs NIFTY 1M%"),
            rs_metrics.get("RS vs NIFTY 3M%"),
            rs_metrics.get("RS vs Sector 1M%"),
            rs_metrics.get("RS vs Sector 3M%"),
        )
    m["Signal Regime"] = _classify_signal_regime(m)
    if include_walkforward and symbol:
        wf_stats = _pick_walkforward_stats(_get_walkforward_signal_stats(symbol, hist), _setup_signal_from_metrics(m))
        m.update(wf_stats)
    else:
        m.update({
            "WF Samples": None,
            "Win Prob%": None,
            "Hist Precision%": None,
            "Exp 5D%": None,
            "Exp 10D%": None,
        })
    m.update(_signal_bundle_from_metrics(m))
    m["Last Updated"] = fmt_dt()
    return m


def _history_to_price_frame(hist):
    dates = hist.get("dates") or []
    closes = hist.get("closes") or []
    highs = hist.get("highs") or []
    lows = hist.get("lows") or []
    volumes = hist.get("volumes") or []
    n = min(len(dates), len(closes), len(highs), len(lows), len(volumes))
    if n < 5:
        return None
    try:
        frame = pd.DataFrame(
            {
                "Close": [float(v) for v in closes[:n]],
                "High": [float(v) for v in highs[:n]],
                "Low": [float(v) for v in lows[:n]],
                "Volume": [float(v) if v is not None else 0.0 for v in volumes[:n]],
            },
            index=pd.to_datetime(list(dates[:n])),
        )
    except Exception:
        return None
    frame = frame[~frame.index.isna()]
    if frame.empty:
        return None
    frame = frame[~frame.index.duplicated(keep="last")].sort_index()
    frame = frame.dropna(subset=["Close", "High", "Low"])
    return frame if len(frame) >= 5 else None


def _resample_history_for_mtf(hist, timeframe: str):
    frame = _history_to_price_frame(hist)
    if frame is None:
        return None
    if timeframe == "D":
        out = frame
    else:
        freq = "W-FRI" if timeframe == "W" else "M"
        out = frame.resample(freq).agg({
            "Close": "last",
            "High": "max",
            "Low": "min",
            "Volume": "sum",
        })
        out = out.dropna(subset=["Close", "High", "Low"])
    if len(out) < 5:
        return None
    return {
        "dates": list(out.index),
        "closes": out["Close"].astype(float).tolist(),
        "highs": out["High"].astype(float).tolist(),
        "lows": out["Low"].astype(float).tolist(),
        "volumes": out["Volume"].astype(float).tolist(),
    }


def _is_price_mtf_constructive(metrics: dict) -> bool:
    price = metrics.get("Current Price")
    if price is None:
        return False
    signal = str(metrics.get("Setup Signal") or metrics.get("Signal") or "").upper()
    if "SELL" in signal or "WEAK" in signal or "BELOW MA200" in signal:
        return False
    for ma_key in ("MA 20", "MA 50", "MA 200"):
        ma_value = metrics.get(ma_key)
        if ma_value is not None and price <= ma_value:
            return False
    pdi = metrics.get("+DI 14")
    mdi = metrics.get("-DI 14")
    if pdi is not None and mdi is not None and mdi > pdi:
        return False
    return any(metrics.get(ma_key) is not None for ma_key in ("MA 20", "MA 50", "MA 200"))


def compute_price_based_mtf(hist):
    hits = set()
    details = {}
    for timeframe in ("D", "W", "M"):
        tf_hist = _resample_history_for_mtf(hist, timeframe)
        if not tf_hist:
            details[timeframe] = None
            continue
        tf_metrics = compute_metrics(tf_hist, include_walkforward=False)
        details[timeframe] = tf_metrics.get("Setup Signal") or tf_metrics.get("Signal")
        if _is_price_mtf_constructive(tf_metrics):
            hits.add(timeframe)
    return {
        "ticks": _format_mtf_ticks(hits),
        "hits": hits,
        "signals": details,
    }

# -----------------------------------------------------------------------------
# TELEGRAM
# -----------------------------------------------------------------------------
def tg_send(text):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID: return False
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    for chunk in [text[i:i+4000] for i in range(0,len(text),4000)] or [""]:
        for attempt in range(5):
            try:
                r=requests.post(url,json={"chat_id":TELEGRAM_CHAT_ID,"text":chunk},timeout=15)
                if r.status_code==200: break
                if r.status_code==429:
                    time.sleep(int(r.json().get("parameters",{}).get("retry_after",30))+2)
                else: time.sleep(min(30,2**attempt))
            except Exception as e: print(f"  TG: {e}"); time.sleep(3)
        time.sleep(1.1)
    return True

def fmt_alert(scanner_name, d):
    sym  = d.get("Symbol","")
    disp = sym[4:] if sym.startswith("BSE:") else sym
    ai_line = (f"[AI] AI Decision : {d.get('AI Decision','')}  ({d.get('AI Conf%','')}% conf)\n"
               f"   Reason      : {d.get('AI Reason','')}\n") if d.get("AI Decision") else ""
    core_signal = str(d.get("Core Signal", "") or "").strip()
    setup_signal = str(d.get("Setup Signal", "") or "").strip()
    live_signal = str(d.get("Signal", "") or "").strip()
    core_line = f"[CORE] Core Signal : {core_signal}" if core_signal and core_signal != live_signal else ""
    setup_line = f"[SETUP] Setup Signal: {setup_signal}" if setup_signal and setup_signal != live_signal else ""
    quality_line = f"[EDGE] Quality={d.get('Signal Quality','')}  Regime={d.get('Signal Regime','')}  WinProb={d.get('Win Prob%','')}%  Exp10D={d.get('Exp 10D%','')}%"
    fund_line = f"[FUND] Score={d.get('Fundamental Score','')}  Risk={d.get('Fundamental Risk Tag','')}  Invest={d.get('Investability Tag','')}"
    guard_line = f"[GUARD] {d.get('Decision Guardrail','')}  Confidence={d.get('Final Confidence Tag','')}" if d.get("Decision Guardrail") else ""
    final_line = f"[FINAL] {d.get('Final Signal','')} -- {d.get('Final Signal Reason','')}" if d.get("Final Signal") else ""
    return "\n".join([
        "🆕 NEW STOCK CAPTURED",
        f"[STAT] Scanner  : {scanner_name}", f"📌 {sym} -- {d.get('Name','')}",
        f"💰 Price    : Rs {d.get('Current Price')}", f"🎯 Signal   : {d.get('Signal')}",
        setup_line.rstrip(),
        core_line.rstrip(),
        quality_line.rstrip(),
        fund_line.rstrip(),
        guard_line.rstrip(),
        final_line.rstrip(),
        ai_line.rstrip(),
        f"📐 RSI={d.get('RSI 14')}  ADX={d.get('ADX 14')}  +DI={d.get('+DI 14')}  -DI={d.get('-DI 14')}",
        f"🌊 ATR={d.get('ATR 14')}  NATR={d.get('NATR 14')}%  RVOL={d.get('Vol Ratio 20')}x",
        f"[UP] 1D={d.get('1D%')}%  1W={d.get('1W%')}%  1M={d.get('1M%')}%",
        f"🕐 {fmt_dt()}",
        f"[LINK] https://www.screener.in/company/{disp}/",
    ])

def _tg_text(value, default="NA"):
    text = str(value or "").strip()
    return text if text else default

def _tg_pct(value, digits=1):
    fv = _as_float(value)
    return f"{fv:.{digits}f}%" if fv is not None else "NA"

def _tg_num(value, digits=1):
    fv = _as_float(value)
    return f"{fv:.{digits}f}" if fv is not None else "NA"

def _tg_has_any(text, *tokens):
    upper = str(text or "").upper()
    return any(tok in upper for tok in tokens)

def _tg_signal_bucket(signal):
    family = _signal_family_key(signal)
    return {
        "BREAKOUT": 6,
        "STRONG BUY": 5,
        "BUY": 4,
        "PULLBACK": 3,
        "OVERSOLD": 2,
        "HOLD": 1,
    }.get(family, 0)

def _tg_quality_bucket(quality):
    upper = str(quality or "").upper()
    if "REJECT" in upper:
        return -2
    if "PASS - HIGH" in upper:
        return 4
    if "PASS - MED" in upper:
        return 3
    if "PASS - LOW" in upper:
        return 2
    if "PASS - UNVERIFIED" in upper:
        return 1
    if "N/A - NON-BULL" in upper:
        return 0
    return -1

def _tg_regime_bucket(regime):
    upper = str(regime or "").upper()
    if "TRENDING" in upper:
        return 2
    if "CHOPPY" in upper:
        return 1
    if "HIGH-VOL" in upper:
        return 0
    return -1

def _tg_guard_bucket(guard):
    upper = str(guard or "").upper()
    if upper.startswith("PASS"):
        return 2
    if upper.startswith("CAUTION"):
        return 1
    if upper.startswith("BLOCK"):
        return 0
    return -1

def _tg_confidence_bucket(tag):
    upper = str(tag or "").upper()
    return {
        "VERY HIGH": 4,
        "HIGH": 3,
        "MED": 2,
        "LOW": 1,
        "BLOCKED": 0,
    }.get(upper, -1)

def _tg_investability_bucket(tag):
    upper = str(tag or "").upper()
    return {
        "INVEST GRADE": 4,
        "ACCUMULATE QUALITY": 3,
        "TRADE ONLY": 1,
        "AVOID": 0,
    }.get(upper, -1)

def _tg_fund_risk_bucket(tag):
    upper = str(tag or "").upper()
    return {
        "LOW": 3,
        "MED": 2,
        "HIGH": 0,
        "UNKNOWN": -1,
    }.get(upper, -1)

def _tg_growth_bucket(tag):
    upper = str(tag or "").upper()
    return {
        "STRONG": 3,
        "GOOD": 2,
        "OK": 1,
        "AVERAGE": 1,
        "WEAK": 0,
    }.get(upper, -1)

def _tg_final_signal_bucket(signal):
    upper = str(signal or "").upper()
    return {
        "HIGH CONVICTION BUY": 6,
        "BUY": 5,
        "EARLY ENTRY": 4,
        "ACCUMULATE": 3,
        "WATCH": 2,
        "TECHNICAL ONLY": 2,
        "WAIT FOR TECHNICALS": 1,
        "TRADE ONLY": 1,
        "AVOID": 0,
    }.get(upper, -1)

def _tg_mtf_hits(value):
    text = str(value or "")
    check = chr(0x2705)
    return sum(1 for tf in ("D", "W", "M") if f"{tf}{check}" in text)

def _tg_entry_timing_label(row):
    current = _as_float(row.get("Current Price"))
    ideal = _as_float(row.get("Ideal Enter Price"))
    dist52 = _as_float(row.get("52W High Dist%"))
    bb = str(row.get("BB Signal") or "").upper()
    cam = str(row.get("Cam Setup") or "").upper()
    final_signal = str(row.get("Final Signal") or "").upper()
    early_ok = str(row.get("Early Entry OK") or "").upper()
    quality = str(row.get("Signal Quality") or "").upper()

    if "REJECT" in quality:
        return "WAIT"

    early_hits = 0
    late_hits = 0

    if current is not None and ideal is not None:
        if current <= ideal * 1.03:
            early_hits += 2
        elif current >= ideal * 1.08:
            late_hits += 2

    if _tg_has_any(bb, "BUY ZONE", "OVERSOLD", "SQUEEZE") or _tg_has_any(cam, "WATCH H4 BREAK", "SQUEEZE NEAR H4", "WATCH L3 SUPPORT"):
        early_hits += 1
    if _tg_has_any(bb, "STRETCHED", "SELL ZONE", "NEAR HIGH") or _tg_has_any(cam, "AT/ABOVE H4 RESISTANCE", "UPPER BAND UNDER H4"):
        late_hits += 1

    if dist52 is not None:
        if dist52 <= 1.5:
            late_hits += 1
        elif dist52 >= 8:
            early_hits += 1

    if final_signal == "EARLY ENTRY" or early_ok == "YES":
        early_hits += 2
    elif early_ok == "WATCH":
        early_hits += 1

    if late_hits >= early_hits + 1:
        return "LATE"
    if early_hits >= late_hits + 1:
        return "EARLY"
    return "MID"

def _tg_dashboard_active_rows(wb):
    if "Dashboard" not in wb.sheetnames:
        return []
    ws = wb["Dashboard"]
    rows = []
    for ri in range(2, ws.max_row + 1):
        sym = str(ws.cell(row=ri, column=DC["Symbol"] + 1).value or "").strip()
        name = str(ws.cell(row=ri, column=DC["Name"] + 1).value or "").strip()
        if not sym and not name:
            continue
        row = {h: ws.cell(row=ri, column=DC[h] + 1).value for h in DASHBOARD_HEADERS}
        if str(row.get("In Screener?") or "").strip().upper() != "YES":
            continue
        if str(row.get("Signal") or "").strip().upper() == "NO DATA":
            continue
        rows.append(row)
    return rows

def _tg_is_preferred_technical_candidate(row):
    if _tg_signal_bucket(row.get("Signal")) < 3:
        return False
    if _tg_quality_bucket(row.get("Signal Quality")) < 2:
        return False
    if _tg_regime_bucket(row.get("Signal Regime")) < 1:
        return False
    if _as_float(row.get("Win Prob%")) is not None and (_as_float(row.get("Win Prob%")) or 0) < 55:
        return False
    if _as_float(row.get("Hist Precision%")) is not None and (_as_float(row.get("Hist Precision%")) or 0) < 55:
        return False
    if _as_float(row.get("Exp 10D%")) is not None and (_as_float(row.get("Exp 10D%")) or 0) <= 0:
        return False
    if _tg_has_any(row.get("Risk Tag"), "HIGH"):
        return False
    if _tg_has_any(row.get("Liquidity Tag"), "THIN", "ILLIQUID"):
        return False
    return True

def _tg_is_preferred_fundamental_candidate(row):
    fresh = str(row.get("Fundamental Freshness") or "").upper()
    score = _as_float(row.get("Fundamental Score"))
    if score is None or score < 70:
        return False
    if fresh in {"INSUFFICIENT", "PARSE ERROR", "PARSER MISSING", "MISSING"}:
        return False
    if _tg_investability_bucket(row.get("Investability Tag")) < 3:
        return False
    if _tg_fund_risk_bucket(row.get("Fundamental Risk Tag")) < 2:
        return False
    if str(row.get("Early Entry OK") or "").upper() not in {"YES", "WATCH"}:
        return False
    if _tg_quality_bucket(row.get("Signal Quality")) < 0:
        return False
    if _tg_guard_bucket(row.get("Decision Guardrail")) < 1:
        return False
    return True

def _tg_is_preferred_combined_candidate(row):
    if _tg_final_signal_bucket(row.get("Final Signal")) < 3:
        return False
    if _tg_guard_bucket(row.get("Decision Guardrail")) < 1:
        return False
    if _tg_confidence_bucket(row.get("Final Confidence Tag")) < 2:
        return False
    if _tg_quality_bucket(row.get("Signal Quality")) < 0:
        return False
    if _tg_has_any(row.get("Risk Tag"), "HIGH"):
        return False
    return True

def _tg_best_technical_key(row):
    win_prob = _as_float(row.get("Win Prob%"))
    hist_precision = _as_float(row.get("Hist Precision%"))
    exp10d = _as_float(row.get("Exp 10D%"))
    consensus = _as_float(row.get("Consensus Score"))
    momentum_rank = _as_int(row.get("Momentum Rank"))
    entry_timing = _tg_entry_timing_label(row)
    return (
        _tg_signal_bucket(row.get("Signal")),
        _tg_quality_bucket(row.get("Signal Quality")),
        _tg_regime_bucket(row.get("Signal Regime")),
        2 if entry_timing == "EARLY" else 1 if entry_timing == "MID" else 0,
        _tg_mtf_hits(row.get("MTF Alignment")),
        0 if _tg_has_any(row.get("Risk Tag"), "HIGH") else 1,
        0 if _tg_has_any(row.get("Liquidity Tag"), "THIN", "ILLIQUID") else 1,
        0 if _tg_has_any(row.get("RS Tag"), "LAGGING", "WEAK RS") else 1,
        win_prob if win_prob is not None else -999.0,
        hist_precision if hist_precision is not None else -999.0,
        exp10d if exp10d is not None else -999.0,
        consensus if consensus is not None else -999.0,
        -(momentum_rank if momentum_rank is not None else 9999),
    )

def _tg_best_fundamental_key(row):
    fund_score = _as_float(row.get("Fundamental Score"))
    combined_score = _as_float(row.get("Tech + Fundamental Score"))
    win_prob = _as_float(row.get("Win Prob%"))
    hist_precision = _as_float(row.get("Hist Precision%"))
    entry_timing = _tg_entry_timing_label(row)
    early_ok = str(row.get("Early Entry OK") or "").upper()
    return (
        _tg_investability_bucket(row.get("Investability Tag")),
        _tg_fund_risk_bucket(row.get("Fundamental Risk Tag")),
        2 if early_ok == "YES" else 1 if early_ok == "WATCH" else 0,
        2 if entry_timing == "EARLY" else 1 if entry_timing == "MID" else 0,
        _tg_final_signal_bucket(row.get("Final Signal")),
        _tg_guard_bucket(row.get("Decision Guardrail")),
        _tg_confidence_bucket(row.get("Final Confidence Tag")),
        _tg_growth_bucket(row.get("Fundamental Quality Tag")),
        _tg_growth_bucket(row.get("Growth Tag")),
        _tg_growth_bucket(row.get("Profitability Tag")),
        fund_score if fund_score is not None else -999.0,
        combined_score if combined_score is not None else -999.0,
        win_prob if win_prob is not None else -999.0,
        hist_precision if hist_precision is not None else -999.0,
    )

def _tg_best_combined_key(row):
    combined_score = _as_float(row.get("Tech + Fundamental Score"))
    win_prob = _as_float(row.get("Win Prob%"))
    hist_precision = _as_float(row.get("Hist Precision%"))
    fund_score = _as_float(row.get("Fundamental Score"))
    return (
        _tg_final_signal_bucket(row.get("Final Signal")),
        _tg_guard_bucket(row.get("Decision Guardrail")),
        _tg_confidence_bucket(row.get("Final Confidence Tag")),
        combined_score if combined_score is not None else -999.0,
        fund_score if fund_score is not None else -999.0,
        win_prob if win_prob is not None else -999.0,
        hist_precision if hist_precision is not None else -999.0,
        _tg_mtf_hits(row.get("MTF Alignment")),
    )

def _tg_pick_best_row(rows, preferred_predicate, score_key):
    preferred_rows = [row for row in rows if preferred_predicate(row)]
    pool = preferred_rows or rows
    if not pool:
        return None
    return max(pool, key=score_key)

def _tg_format_pick_block(label, row):
    sym = _tg_text(row.get("Symbol"))
    name = _tg_text(row.get("Name"))
    timing = _tg_entry_timing_label(row)
    link = _tg_text(row.get("Screener Link"), "")
    win_prob = _tg_pct(row.get("Win Prob%"))
    hist_precision = _tg_pct(row.get("Hist Precision%"))
    exp10d = _tg_pct(row.get("Exp 10D%"))
    final_signal = _tg_text(row.get("Final Signal"))
    lines = [f"[{label}] {sym} - {name}"]
    if label == "OVERALL":
        lines.append(
            f"Final={final_signal} | Guard={_tg_text(row.get('Decision Guardrail'))} | Conf={_tg_text(row.get('Final Confidence Tag'))}"
        )
        lines.append(
            f"Win={win_prob} | Hist={hist_precision} | Entry={timing} | TF Score={_tg_num(row.get('Tech + Fundamental Score'))}"
        )
    elif label == "TECH":
        lines.append(
            f"Signal={_tg_text(row.get('Signal'))} | Quality={_tg_text(row.get('Signal Quality'))} | Regime={_tg_text(row.get('Signal Regime'))}"
        )
        lines.append(
            f"Win={win_prob} | Hist={hist_precision} | Exp10D={exp10d} | Entry={timing}"
        )
    else:
        lines.append(
            f"Invest={_tg_text(row.get('Investability Tag'))} | FScore={_tg_num(row.get('Fundamental Score'))} | FRisk={_tg_text(row.get('Fundamental Risk Tag'))}"
        )
        lines.append(
            f"EarlyOK={_tg_text(row.get('Early Entry OK'))} | Win={win_prob} | Final={final_signal} | Entry={timing}"
        )
    if link:
        lines.append(f"Link={link}")
    return lines

def fmt_dashboard_best_picks(wb):
    rows = _tg_dashboard_active_rows(wb)
    if not rows:
        return ""

    overall = _tg_pick_best_row(rows, _tg_is_preferred_combined_candidate, _tg_best_combined_key)
    tech = _tg_pick_best_row(rows, _tg_is_preferred_technical_candidate, _tg_best_technical_key)
    fund = _tg_pick_best_row(rows, _tg_is_preferred_fundamental_candidate, _tg_best_fundamental_key)

    if not any((overall, tech, fund)):
        return ""

    lines = ["[TOP] Best Current Dashboard Picks"]
    seen = set()
    for label, row in (("OVERALL", overall), ("TECH", tech), ("FUND", fund)):
        if not row:
            continue
        key = (str(row.get("Symbol") or "").strip().upper(), label)
        if key in seen:
            continue
        seen.add(key)
        lines.extend(_tg_format_pick_block(label, row))
    return "\n".join(lines)

# -----------------------------------------------------------------------------
# PARALLEL ITERATION ENGINE (Dedup + Parallel 2)
# -----------------------------------------------------------------------------
def process_all_scanners_parallel(wb):
    """
    Process ALL scanners in one iteration using dedup + parallel fetching.

    5-Phase pipeline:
      Phase 1: Fetch all screener pages (parallel 2)
      Phase 2: Update sheets + resolve new symbols (sequential -- needs wb)
      Phase 3: Batch YF fetch for unique symbols (parallel 2, deduped)
      Phase 4: Batch AI calls for unique symbols (parallel 2, deduped)
      Phase 5: Write results to each scanner sheet (sequential -- needs wb)

    Returns: list of (scanner_name, stock_desc) for new stocks
    """
    all_new = []
    _bb_data.clear()  # Reset BB cache for this iteration
    _cam_data.clear()  # Reset Camarilla cache for this iteration
    _price_mtf_data.clear()  # Reset price-based D/W/M cache for this iteration
    _fundamental_data.clear()  # Reset fundamental lookup for this iteration

    # ------- PHASE 1: Fetch all screener pages in parallel ----------------
    print(f"\n  [+] Phase 1: Fetching {len(SCANNERS)} screener pages (parallel {SCREENER_FETCH_WORKERS})...")
    screener_data = {}  # scanner_id -> [stocks]
    phase1_total_stocks = 0

    def _fetch_page(sc):
        try:
            stocks = fetch_screener(sc["url"])
            return sc["id"], stocks
        except Exception as e:
            print(f"    [FAIL] {sc['name']}: {e}")
            return sc["id"], []

    with ThreadPoolExecutor(max_workers=SCREENER_FETCH_WORKERS) as pool:
        futures = {pool.submit(_fetch_page, sc): sc for sc in SCANNERS}
        _phase1_done = 0
        _phase1_total = len(SCANNERS)
        for f in as_completed(futures):
            sc = futures[f]
            sid, stocks = f.result()
            screener_data[sid] = stocks
            _phase1_done += 1
            phase1_total_stocks += len(stocks)
            if stocks and LOG_SCREEN_PAGE_COUNTS:
                print(f"    [LIST] {sc['name']}: {len(stocks)} stocks")
            if _phase1_done % 10 == 0 or _phase1_done == _phase1_total:
                print(f"    [UP] Phase 1 progress: {_phase1_done}/{_phase1_total}")

    if phase1_total_stocks == 0:
        raise Phase1NoStocksError(
            "Phase 1 returned zero stocks across all screens. "
            "Workbook left unchanged for this iteration."
        )

    # ------- PHASE 2: Update sheets + resolve symbols (sequential) --------
    print(f"\n  [+] Phase 2: Updating sheets + resolving symbols...")
    scanner_meta = {}   # scanner_id -> {existing, new_stocks, ws}
    all_unique_syms = set()
    sym_to_name = {}    # sym -> name (for AI calls)
    sym_to_fund_meta = {}  # sym -> Screener metadata for fundamental fetches

    for sc in SCANNERS:
        ws = ensure_scanner_sheet(wb, sc)
        bg = sc.get("color", "FFFFFF")
        now = fmt_dt()
        new_stocks = []

        existing = load_sheet_stocks(ws)
        for entry in existing.values():
            ws.cell(row=entry["row"], column=C["In Screener?"]+1, value="No")

        for s in screener_data.get(sc["id"], []):
            sym = s.get("symbol", "").strip()
            name = s.get("name", "").strip()
            slug = s.get("screener_slug", "").strip()
            if not sym and not name:
                continue
            if not sym:
                sym = resolve_symbol(name, s.get("bseCode", ""), screener_slug=slug)
            key = _stock_key(sym, name)
            if not key:
                continue
            if sym and not sym.startswith(SENTINEL):
                sym_to_fund_meta.setdefault(sym, {
                    "name": name,
                    "screener_slug": slug,
                    "bseCode": s.get("bseCode", ""),
                })
            if key in existing:
                ri = existing[key]["row"]
                ws.cell(row=ri, column=C["Last Seen"]+1, value=now)
                ws.cell(row=ri, column=C["In Screener?"]+1, value="Yes")
                stored = str(ws.cell(row=ri, column=C["Symbol"]+1).value or "").strip()
                if (not stored or stored.startswith(SENTINEL)) and sym:
                    ws.cell(row=ri, column=C["Symbol"]+1, value=sym)
                    if not sym.startswith(SENTINEL):
                        sym_to_fund_meta.setdefault(sym, {
                            "name": name,
                            "screener_slug": slug,
                            "bseCode": s.get("bseCode", ""),
                        })
            else:
                if not sym:
                    sym = SENTINEL
                nr = ws.max_row + 1
                init = dict.fromkeys(HEADERS, None)
                init.update({"Symbol": sym, "Name": name, "First Captured": now,
                             "Last Seen": now, "In Screener?": "Yes", "Signal": "Pending..."})
                write_stock_row(ws, nr, init, bg_color=bg, apply_static_style=True)
                existing[key] = {"row": nr, "data": init}
                new_stocks.append({"Symbol": sym, "Name": name})
                if LOG_NEW_STOCK_EVENTS:
                    print(f"      [NEW] NEW: {sym} -- {name}")

        # Collect only active symbols; inactive rows stay frozen on-sheet.
        for key, entry in existing.items():
            ri = entry["row"]
            if str(ws.cell(row=ri, column=C["In Screener?"]+1).value or "").strip() != "Yes":
                continue
            sym_cell = str(ws.cell(row=ri, column=C["Symbol"]+1).value or "").strip()
            if sym_cell and not sym_cell.startswith(SENTINEL):
                all_unique_syms.add(sym_cell)
                name_cell = ws.cell(row=ri, column=C["Name"]+1).value or ""
                sym_to_name[sym_cell] = name_cell
                sym_to_fund_meta.setdefault(sym_cell, {
                    "name": name_cell,
                    "screener_slug": "",
                    "bseCode": "",
                })

        scanner_meta[sc["id"]] = {
            "existing": existing,
            "new_stocks": new_stocks,
            "ws": ws,
        }

    total_refs = sum(len(m["existing"]) for m in scanner_meta.values())
    print(f"    [STAT] {len(all_unique_syms)} unique symbols across {total_refs} scanner references")
    print(f"    [INFO] Dedup saves ~{total_refs - len(all_unique_syms)} redundant YF+AI calls")

    # ------- PHASE 2B: Fundamental fetch/cache (parallel, deduped) --------
    global_fundamentals = _collect_fundamentals_parallel(sym_to_fund_meta)

    # ------- PHASE 3: Batch YF fetch (parallel, deduped) -----------------
    print(f"\n  [+] Phase 3: Fetching YF data for {len(all_unique_syms)} unique symbols (parallel {PARALLEL_WORKERS})...")
    global_metrics = {}  # sym -> metrics dict
    _yf_done = [0]
    _yf_total = len(all_unique_syms)
    session_key = _completed_market_session_key()
    snapshot_hits = 0
    snapshot_writes = 0
    if session_key:
        _load_signal_snapshot_cache()
        print(f"  [SNAP] Off-market signal snapshot active for completed session {session_key}")

    def _fetch_one_yf(sym):
        if session_key:
            cached = _get_signal_snapshot_metrics(session_key, sym)
            if cached:
                return sym, cached, "snapshot", None
        try:
            hist = fetch_history(sym)
            if hist and len(hist["closes"]) >= 10:
                m = compute_metrics(hist, symbol=sym)
                return sym, m, "fresh", compute_price_based_mtf(hist)
        except Exception as e:
            print(f"      [FAIL] YF {sym}: {e}")
        return sym, None, "miss", None

    with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as pool:
        futures = {pool.submit(_fetch_one_yf, sym): sym for sym in sorted(all_unique_syms)}
        for f in as_completed(futures):
            sym, m, source, price_mtf = f.result()
            _yf_done[0] += 1
            if m:
                if source == "snapshot":
                    snapshot_hits += 1
                elif session_key and source == "fresh":
                    _put_signal_snapshot_metrics(session_key, sym, m)
                    snapshot_writes += 1
                if price_mtf and price_mtf.get("ticks"):
                    _price_mtf_data[sym] = price_mtf["ticks"]
                # Store BB data in module-level cache for Dashboard
                if m.get("BB %B") is not None or m.get("BB Width") is not None:
                    _bb_data[sym] = {
                        "BB %B": m.get("BB %B"),
                        "BB Width": m.get("BB Width"),
                        "BB Width Pctl": m.get("BB Width Pctl"),
                        "BB Squeeze": m.get("BB Squeeze"),
                    }
                if any(m.get(k) is not None for k in ("Cam H3", "Cam H4", "Cam L3", "Cam L4")):
                    _cam_data[sym] = {
                        "Cam H3": m.get("Cam H3"),
                        "Cam H4": m.get("Cam H4"),
                        "Cam L3": m.get("Cam L3"),
                        "Cam L4": m.get("Cam L4"),
                    }
                global_metrics[sym] = m
            if _yf_done[0] % 25 == 0 or _yf_done[0] == _yf_total:
                print(f"    [UP] YF progress: {_yf_done[0]}/{_yf_total} ({len(global_metrics)} OK)")

    _yf_missed = sorted(all_unique_syms - set(global_metrics.keys()))
    if _yf_missed:
        print(f"  ⚠️  YF data: {len(global_metrics)}/{_yf_total} symbols fetched  |  MISSED: {', '.join(_yf_missed)}")
    else:
        print(f"  ✅ YF data: {len(global_metrics)}/{_yf_total} symbols fetched (all OK)")

    if session_key:
        _save_signal_snapshot_cache()
        print(f"  [SNAP] Reused {snapshot_hits} frozen symbols, wrote {snapshot_writes} fresh symbols for {session_key}")
    _save_walkforward_cache()

    # Merge fundamentals after technicals are computed, keeping compute_metrics pure.
    for sym, metrics in global_metrics.items():
        fund = global_fundamentals.get(sym) if isinstance(global_fundamentals, dict) else None
        if fund:
            metrics.update(fund)
        _apply_combined_final_signal(metrics)

    # ------- PHASE 4: Batch AI calls (parallel, deduped) -----------------
    global_ai = {}  # sym -> ai_result dict
    if AI_ENABLED and global_metrics:
        need_ai = {sym: m for sym, m in global_metrics.items()
                   if m.get("Signal") not in ("No Data", "Error", "Pending...", "Symbol Not Found")}
        _ai_total = len(need_ai)
        if _gemini_batch_enabled_for_run(len(need_ai)):
            print(f"\n  [+] Phase 4: AI analysis for {len(need_ai)} symbols (Gemini batch mode)...")
            global_ai = _collect_ai_google_batch(need_ai, sym_to_name)
        else:
            print(f"\n  [+] Phase 4: AI analysis for {len(need_ai)} symbols (parallel {PARALLEL_WORKERS})...")
            global_ai = _collect_ai_parallel(need_ai, sym_to_name)
        print(f"  [OK] AI: {len(global_ai)}/{_ai_total} symbols analyzed")
    else:
        ai_status = "disabled" if not AI_ENABLED else "skipped (no data)"
        print(f"\n  [+] Phase 4: AI {ai_status}.")

    # ------- PHASE 5: Write results to scanner sheets (sequential) -------
    print(f"\n  [+] Phase 5: Writing results to {len(SCANNERS)} scanner sheets...")
    ph = ensure_price_history_sheet(wb)

    for sc_idx, sc in enumerate(SCANNERS):
        meta = scanner_meta.get(sc["id"])
        if not meta:
            continue
        ws = meta["ws"]
        existing = meta["existing"]
        new_stocks = meta["new_stocks"]
        new_symbols = {d.get("Symbol") for d in new_stocks if d.get("Symbol")}
        new_names = {d.get("Name") for d in new_stocks if d.get("Name")}

        stock_count = 0
        ai_written = 0
        for key, entry in existing.items():
            ri = entry["row"]
            if str(ws.cell(row=ri, column=C["In Screener?"]+1).value or "").strip() != "Yes":
                continue
            sym_cell = str(ws.cell(row=ri, column=C["Symbol"]+1).value or "").strip()

            if not sym_cell or sym_cell.startswith(SENTINEL):
                _clear_runtime_cells(ws, ri, "Symbol Not Found", clear_metrics=True)
                continue

            m = global_metrics.get(sym_cell)
            if not m:
                _clear_runtime_cells(ws, ri, "No Data", clear_metrics=True)
                continue

            # Copy so we don't mutate the global cache
            m_copy = dict(m)

            # Compute Since Capture %
            cap_val = ws.cell(row=ri, column=C["Capture Price"]+1).value
            if not cap_val or cap_val == 0:
                cap_val = m_copy["Current Price"]
                ws.cell(row=ri, column=C["Capture Price"]+1, value=cap_val)
            m_copy["Since Capture%"] = (round((m_copy["Current Price"] - cap_val) / cap_val * 100, 2)
                                        if cap_val and abs(cap_val) > 1e-10 else None)

            # Apply AI results from global cache
            ai_result = global_ai.get(sym_cell, {})
            m_copy["AI Decision"] = ai_result.get("decision", "")
            m_copy["AI Reason"]   = ai_result.get("reason", "")
            m_copy["AI Conf%"]    = ai_result.get("confidence", None)
            if ai_result:
                ai_written += 1
            _apply_combined_final_signal(m_copy)

            # Write to sheet
            row_data = dict.fromkeys(HEADERS, None)
            row_data.update({
                "Symbol": sym_cell,
                "Name": ws.cell(row=ri, column=C["Name"]+1).value,
                "First Captured": ws.cell(row=ri, column=C["First Captured"]+1).value,
                "Last Seen": ws.cell(row=ri, column=C["Last Seen"]+1).value,
                "In Screener?": ws.cell(row=ri, column=C["In Screener?"]+1).value,
                "Capture Price": cap_val,
            })
            row_data.update(m_copy)

            # Protect "First Captured" -- must be a date, never a number
            fc_val = row_data.get("First Captured")
            if fc_val is None or not (isinstance(fc_val, str) and '-' in fc_val and ':' in fc_val and len(fc_val) >= 10):
                row_data["First Captured"] = fmt_dt()  # recovery: stamp now
            write_stock_row(ws, ri, row_data)
            append_price_history(ph, sc["name"], row_data)

            # Telegram alert for new stocks
            if sym_cell in new_symbols or row_data.get("Name") in new_names:
                tg_send(fmt_alert(sc["name"], row_data))
                all_new.append((sc["name"], f"{sym_cell} ({row_data.get('Name', '')})"))

            stock_count += 1

        ai_tag = f" ({ai_written} AI'd)" if AI_ENABLED else ""
        if LOG_SCANNER_WRITE_EACH_SHEET or (sc_idx + 1) % 10 == 0 or (sc_idx + 1) == len(SCANNERS):
            print(f"    [OK] [{sc['id']}] {sc['name']}: {stock_count} stocks written{ai_tag}")

        # S3 checkpoint every N scanners
        if S3_CHECKPOINT_EVERY > 0 and (sc_idx + 1) % S3_CHECKPOINT_EVERY == 0:
            print(f"    [SAVE] S3 checkpoint ({sc_idx + 1}/{len(SCANNERS)})...")
            s3_upload_excel(
                wb,
                save_local=S3_CHECKPOINT_SAVE_LOCAL,
                backup=S3_CHECKPOINT_BACKUP,
            )

    return all_new



def main(max_iterations=None):
    _primary_provider = _ai_primary_provider()
    _active_models = _provider_model_list(_primary_provider)
    _active_model = _active_models[0] if _active_models else "N/A"
    _secondary_label = _ai_secondary_provider().upper() if AI_SECONDARY_ENABLED else "OFF"
    _signal_engine = _signal_engine_name().upper()
    print("="*62)
    print(f"  [STAT] SCREENER.IN TRACKER -- PARALLEL (×{PARALLEL_WORKERS}) + DEDUP")
    print(f"  🕐 {fmt_dt(ist_now())}  |  AI: {_primary_provider.upper()} ({_active_model}) | Secondary: {_secondary_label}")
    print("="*62)
    print(f"  [SIG] Engine: {_signal_engine}")
    if COMPACT_RUNTIME_LOGS:
        print("  [LOG] Compact runtime logging active (Colab-safe)")
    wb = init_workbook()
    start_msg = (f"[START] Screener Tracker Started (PARALLEL ×{PARALLEL_WORKERS})\n{fmt_dt(ist_now())}\n"
                 f"Scanners: {len(SCANNERS)}  |  AI: {_primary_provider.upper()} ({_active_model})  |  Secondary: {_secondary_label}  |  Signal: {_signal_engine}\n"
                 f"S3 Excel: s3://{S3_BUCKET}/{S3_EXCEL_KEY}\n"
                 f"S3 Dashboard Excel: s3://{S3_BUCKET}/{S3_DASHBOARD_EXCEL_KEY}")
    if DASHBOARD_DB_ENABLED:
        start_msg += f"\nS3 DB: s3://{S3_BUCKET}/{S3_DASHBOARD_DB_KEY}"
    tg_send(start_msg)
    iteration = 0
    try:
        while True:
            iteration += 1; start_ts = ist_now()
            print(f"\n{'='*62}")
            print(f"  Iteration #{iteration}  |  {fmt_dt(start_ts)}")
            print(f"{'='*62}")
            try:
                # Periodic workbook reload: flush accumulated RAM
                if iteration > 1 and (iteration - 1) % RELOAD_WB_EVERY == 0:
                    print("  [RELOAD] Reloading workbook from S3 (periodic memory flush)...")
                    wb_fresh = s3_download_excel(retries=3)
                    if wb_fresh and wb_fresh != _S3_FIRST_RUN:
                        wb = wb_fresh
                        ensure_price_history_sheet(wb)
                        ensure_dashboard_sheet(wb)
                        ensure_dashboard_history_sheet(wb)
                        ensure_validation_sheet(wb)
                        for sc in SCANNERS: ensure_scanner_sheet(wb, sc)
                        print("  [OK] Workbook reloaded.")
                    else:
                        print("  [WARN]  Reload failed -- continuing with existing workbook.")

                # -- PARALLEL ITERATION ----------------------------------
                all_new = process_all_scanners_parallel(wb)

                # -- Update Dashboard ------------------------------------
                print("\n  [STAT] Updating Dashboard...")
                dashboard_updated = True
                try:
                    update_dashboard(wb)
                except Exception as e:
                    dashboard_updated = False
                    print(f"    [FAIL] Dashboard update failed: {e}")

                validation_result = None
                if dashboard_updated:
                    try:
                        print("  [VALID] Running post-run freshness check...")
                        validation_result = run_post_run_validation(wb, iteration)
                        print(
                            f"  [VALID] {validation_result.get('Status')} | "
                            f"{validation_result.get('Matched Rows', 0)}/{validation_result.get('Checked Rows', 0)} matched | "
                            f"mismatch={validation_result.get('Mismatch Rows', 0)} unresolved={validation_result.get('Unresolved Rows', 0)}"
                        )
                    except Exception as e:
                        print(f"  [WARN]  Post-run validation failed: {e}")

                # -- Append Dashboard DB snapshot + upload ----------------
                db_ok = None
                db_rows = 0
                if DASHBOARD_DB_ENABLED and dashboard_updated:
                    try:
                        db_rows, db_path = append_dashboard_snapshot_to_db(
                            wb,
                            iteration=iteration,
                            snapshot_at=fmt_dt(ist_now()),
                        )
                        print(f"  [DB] Appended {db_rows} dashboard rows -> {db_path}")
                        db_ok = s3_upload_dashboard_db(db_path)
                    except Exception as e:
                        db_ok = False
                        print(f"  [WARN]  Dashboard DB append/upload failed: {e}")
                elif DASHBOARD_DB_ENABLED:
                    db_ok = False
                    print("  [WARN]  Skipping Dashboard DB append because Dashboard update failed.")

                # Final cycle uploads
                ok = s3_upload_excel(wb)
                dash_excel_ok = s3_upload_dashboard_excel(wb) if dashboard_updated else False
                elapsed = int((ist_now() - start_ts).total_seconds())

                # -- Presigned S3 download links -------------------------
                presigned_url = s3_generate_presigned_url(expiry_seconds=86400, key=S3_EXCEL_KEY) if ok else None
                dashboard_presigned_url = (
                    s3_generate_presigned_url(expiry_seconds=86400, key=S3_DASHBOARD_EXCEL_KEY)
                    if dash_excel_ok else None
                )

                lines = [f"[OK] Iteration #{iteration} done (PARALLEL ×{PARALLEL_WORKERS})",
                         f"🕐 {fmt_dt(ist_now())} ({elapsed}s)",
                         f"📁 Workbook {'saved [OK]' if ok else 'FAILED [FAIL]'}",
                         f"📊 Dashboard {'saved [OK]' if dash_excel_ok else 'FAILED [FAIL]'}",
                         f"🆕 New: {len(all_new)}"]
                if validation_result:
                    lines.append(
                        f"[VALID] {validation_result.get('Status')} "
                        f"({validation_result.get('Matched Rows', 0)}/{validation_result.get('Checked Rows', 0)} matched, "
                        f"{validation_result.get('Mismatch Rows', 0)} mismatch, "
                        f"{validation_result.get('Unresolved Rows', 0)} unresolved)"
                    )
                if DASHBOARD_DB_ENABLED:
                    lines.append(f"[DB] {'saved [OK]' if db_ok else 'FAILED [FAIL]'} ({db_rows} rows)")
                if all_new: lines += [f"  - [{sc}] {nm}" for sc, nm in all_new[:10]]
                if presigned_url:
                    lines.append(f"[DL] Workbook: {presigned_url}")
                if dashboard_presigned_url:
                    lines.append(f"[DL] Dashboard: {dashboard_presigned_url}")
                tg_send("\n".join(lines))
                if dashboard_updated and dash_excel_ok:
                    try:
                        best_picks_text = fmt_dashboard_best_picks(wb)
                        if best_picks_text:
                            tg_send(best_picks_text)
                    except Exception as e:
                        print(f"  [WARN] Telegram best-picks summary failed: {e}")
                status_line = f"\n  Workbook: {'[OK]' if ok else '[FAIL]'}  |  Dashboard: {'[OK]' if dash_excel_ok else '[FAIL]'}"
                if validation_result:
                    status_line += (
                        f"  |  VALID: {validation_result.get('Status')} "
                        f"({validation_result.get('Matched Rows', 0)}/{validation_result.get('Checked Rows', 0)})"
                    )
                if DASHBOARD_DB_ENABLED:
                    status_line += f"  |  DB: {'[OK]' if db_ok else '[FAIL]'} ({db_rows} rows)"
                status_line += f"  |  New: {len(all_new)}  |  [TIME] {elapsed}s"
                print(status_line)
                if presigned_url:
                    print(f"  [LINK] Workbook presigned URL generated (24h expiry)")
                if dashboard_presigned_url:
                    print(f"  [LINK] Dashboard presigned URL generated (24h expiry)")
                if max_iterations is not None and iteration >= max_iterations:
                    print("  🏁 Done."); break

                # -- Colab auto-restart ----------------------------------
                if (COLAB_RESTART_EVERY > 0 and
                    iteration % COLAB_RESTART_EVERY == 0 and
                    _is_colab()):
                    print(f"\n  [RELOAD] Reached {iteration} iterations -- restarting Colab runtime...")
                    tg_send(f"[RELOAD] Auto-restart after iteration #{iteration}\n"
                            f"Runtime will restart and re-run all cells automatically.")
                    colab_restart_and_run_all()

            except Phase1NoStocksError as e:
                print(f"  [WARN] Iteration #{iteration} skipped: {e}")
                tg_send(f"[WARN] Tracker iteration #{iteration} skipped: {e}")
            except Exception as e:
                print(f"  [CRASH] Iteration #{iteration} crashed: {e}")
                print("  [SAVE] Emergency S3 upload...")
                try: s3_upload_excel(wb)
                except Exception as e2: print(f"  [FAIL] Emergency upload failed: {e2}")
                tg_send(f"[WARN] Tracker iteration #{iteration} crashed: {e}\nWill retry after sleep.")
            print(f"  [WAIT] Sleeping {ROTATION_COOLDOWN_SEC}s...\n")
            time.sleep(ROTATION_COOLDOWN_SEC)

    except KeyboardInterrupt:
        print("\n  [HALT] Interrupted by user -- saving work before exit...")
        tg_send("[STOP] Tracker manually stopped. Saving final state to S3...")
        try: s3_upload_excel(wb)
        except Exception as e: print(f"  [FAIL] Final save failed: {e}")
        print("  [OK] Done. Goodbye.")

# -----------------------------------------------------------------------------
# COLAB AUTO-RESTART
# -----------------------------------------------------------------------------
def _is_colab():
    try:
        import google.colab  # noqa: F401
        return True
    except ImportError:
        return False

def colab_restart_and_run_all():
    if not _is_colab():
        print("  [WARN]  Not running in Colab -- skipping restart.")
        return False
    try:
        from IPython.display import display, Javascript
        print("  [RELOAD] Scheduling Colab restart + Run All...")
        display(Javascript('''
            (function() {
                function waitAndRunAll() {
                    if (document.querySelector('#runtime-menu-button, [data-command="runall"]')) {
                        if (window.colab && window.colab.global && window.colab.global.notebook) {
                            window.colab.global.notebook.kernel.invokeFunction('notebook.execute_all_cells', [], {});
                        } else {
                            document.querySelector('[data-command="runall"]')?.click();
                        }
                    } else {
                        setTimeout(waitAndRunAll, 2000);
                    }
                }
                setTimeout(waitAndRunAll, 5000);
            })();
        '''))
        time.sleep(2)
        os._exit(0)
    except Exception as e:
        print(f"  [FAIL] Colab restart failed: {e}")
        return False

if __name__ == "__main__":
    main()
    # main(max_iterations=1)
